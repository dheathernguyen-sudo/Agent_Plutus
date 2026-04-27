"""build_workbook.py — Declarative Excel workbook builder.

Reads a portfolio model dict and writes 2026_Portfolio_Analysis.xlsx.
Each tab is defined as an ordered list of sections. Row numbers are
auto-tracked. Named ranges are defined as cells are created.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
import datetime

# ---------------------------------------------------------------------------
# Shared styles
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HEADER_ALIGN = Alignment(horizontal='center')
TITLE_FONT = Font(name='Arial', size=14, bold=True)
SECTION_FONT = Font(name='Arial', size=12, bold=True)
BOLD_FONT = Font(name='Arial', size=10, bold=True)
BLACK_FONT = Font(name='Arial', size=10)
BLUE_FONT = Font(name='Arial', size=10, color='0000FF')
GREEN_FONT = Font(name='Arial', size=10, color='008000')
NOTE_FONT = Font(name='Arial', size=9, italic=True, color='666666')
PROSE_FONT = Font(name='Arial', size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
DOLLAR = '$#,##0.00'
PCT = '0.00%'
QTY_FMT = '#,##0.000'
N_FMT = '0.0000'


# ---------------------------------------------------------------------------
# Cell and row helpers
# ---------------------------------------------------------------------------
def _cell(ws, row, col, value=None, font=None, fmt=None):
    """Write a cell with optional font and number format."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or BLUE_FONT
    c.border = THIN_BORDER
    if fmt:
        c.number_format = fmt
    return c


def _header_row(ws, row, labels, col_start=1):
    """Write a formatted header row."""
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = THIN_BORDER


def _section_header(ws, row, title):
    """Write a section header."""
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    return row + 1


# ---------------------------------------------------------------------------
# Named range helpers
# ---------------------------------------------------------------------------
from registry import _TAB_PREFIX


def _define_name(wb, tab_name, key, col, row):
    """Define a single named range: {prefix}_{key} -> 'Tab'!$COL$ROW."""
    prefix = _TAB_PREFIX.get(tab_name)
    if not prefix:
        return
    name = f"{prefix}_{key}"
    if any(c in tab_name for c in " ()"):
        ref = f"'{tab_name}'!${col}${row}"
    else:
        ref = f"{tab_name}!${col}${row}"
    # Remove existing if present
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names.add(DefinedName(name=name, attr_text=ref))


# ---------------------------------------------------------------------------
# Section writer framework
# ---------------------------------------------------------------------------
def write_sections(wb, tab_name, title, subtitle, sections, col_widths, acct_data=None, model=None):
    """Create a sheet and write sections in order.

    Each section is (title_str, builder_func) where builder_func(ws, row, acct_data, model, wb)
    returns the next row number and a dict of {name_key: (col, row)} for named ranges.

    Returns (ws, row_map) where row_map has all named range entries.
    """
    ws = wb.create_sheet(tab_name)

    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    row = 1
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value=subtitle).font = NOTE_FONT
    row += 2

    row_map = {}

    for section_title, builder_fn in sections:
        row = _section_header(ws, row, section_title)
        row, names = builder_fn(ws, row, acct_data, model, wb)
        row_map.update(names)
        row += 1  # gap between sections

    # Define named ranges (skip internal tracking keys that start with _)
    for key, val in row_map.items():
        if not key.startswith("_") and isinstance(val, tuple) and len(val) == 2:
            col, r = val
            _define_name(wb, tab_name, key, col, r)

    ws.sheet_view.showGridLines = False
    return ws, row_map


# ---------------------------------------------------------------------------
# Account tab section builders
# ---------------------------------------------------------------------------
def _build_return_section(ws, row, acct, model, wb):
    """YTD RETURN CALCULATIONS section."""
    _header_row(ws, row, ['Metric', 'Value', 'Note'])
    row += 1
    names = {}

    names["TWR"] = ("B", row)
    _cell(ws, row, 1, 'Time-Weighted Return (YTD)')
    _cell(ws, row, 2, None, fmt=PCT)  # forward-filled
    row += 1

    names["MWRR"] = ("B", row)
    _cell(ws, row, 1, 'Money-Weighted Return (YTD)')
    _cell(ws, row, 2, acct["returns"]["mwrr"], font=BLACK_FONT, fmt=PCT)
    _cell(ws, row, 3, '(computed from monthly cash flows)', font=NOTE_FONT)
    row += 1

    names["cb_return"] = ("B", row)
    _cell(ws, row, 1, 'Cost Basis Return')
    _cell(ws, row, 2, None, fmt=PCT)  # forward-filled
    _cell(ws, row, 3, 'Unrealized G/L / Cost Basis', font=NOTE_FONT)
    row += 1

    # Cash flow labels vary by account type
    labels = acct.get("cash_flow_labels", {"add": "Additions", "sub": "Subtractions"})
    _cell(ws, row, 1, f'Total {labels["add"]}')
    _cell(ws, row, 2, None, fmt=DOLLAR)  # forward-filled
    names["total_add"] = ("B", row)
    row += 1

    _cell(ws, row, 1, f'Total {labels["sub"]}')
    _cell(ws, row, 2, None, fmt=DOLLAR)  # forward-filled
    names["total_sub"] = ("B", row)
    row += 1

    return row, names


def _build_gain_section(ws, row, acct, model, wb):
    """YTD INVESTMENT GAIN SUMMARY section."""
    _header_row(ws, row, ['Metric', 'Value', 'Note'])
    row += 1
    names = {}
    gains = acct["gains"]

    names["dividends"] = ("B", row)
    _cell(ws, row, 1, 'Dividends/Income' if not acct.get("is_margin") else 'Dividends Received')
    _cell(ws, row, 2, None, fmt=DOLLAR)  # forward-filled
    row += 1

    names["unrealized"] = ("B", row)
    _cell(ws, row, 1, 'Unrealized Gain/Loss')
    _cell(ws, row, 2, None, fmt=DOLLAR)  # forward-filled
    _cell(ws, row, 3, 'Current holdings vs. cost basis (all-time)', font=NOTE_FONT)
    row += 1

    names["realized"] = ("B", row)
    _cell(ws, row, 1, 'Realized Gain/Loss (2026)')
    _cell(ws, row, 2, None, fmt=DOLLAR)  # forward-filled
    row += 1

    names["total_ytd"] = ("B", row)
    _cell(ws, row, 1, 'Total YTD Gain', font=BOLD_FONT)
    div_r = names["dividends"][1]
    unr_r = names["unrealized"][1]
    rea_r = names["realized"][1]
    _cell(ws, row, 2, f'=B{div_r}+B{unr_r}+B{rea_r}', font=BOLD_FONT, fmt=DOLLAR)
    _cell(ws, row, 3, 'Unrealized + Realized + Dividends', font=NOTE_FONT)
    row += 1

    return row, names


def _build_holdings_section(ws, row, acct, model, wb):
    """CURRENT HOLDINGS section."""
    names = {}
    holdings = acct.get("holdings", [])
    is_rh = acct.get("is_margin", False)

    # Angel investments use "investments" instead of "holdings"
    investments = acct.get("investments", [])
    if not holdings and investments:
        # Angel-style: different columns
        _header_row(ws, row, ['Company', 'Sector', 'Year', 'Series', 'Amount Invested',
                              'Est. Current Value', 'Return Multiple'])
        row += 1
        hold_first = row
        for inv in investments:
            pm_invest = inv.get("pm_invest", 0) or 1
            pm_latest = inv.get("pm_latest", 0) or pm_invest
            est_value = inv["amount"] * (pm_latest / pm_invest) if pm_invest else inv["amount"]
            _cell(ws, row, 1, inv["company"])
            _cell(ws, row, 2, inv.get("sector", ""))
            _cell(ws, row, 3, inv.get("year", ""))
            _cell(ws, row, 4, inv.get("series", ""))
            _cell(ws, row, 5, inv["amount"], fmt=DOLLAR)
            _cell(ws, row, 6, round(est_value, 2), fmt=DOLLAR)
            _cell(ws, row, 7, f'=F{row}/E{row}', font=BLACK_FONT, fmt='0.00x')
            row += 1
        hold_last = row - 1

        # TOTAL row
        _cell(ws, row, 1, 'TOTAL', font=BOLD_FONT)
        _cell(ws, row, 5, f'=SUM(E{hold_first}:E{hold_last})', font=BOLD_FONT, fmt=DOLLAR)
        _cell(ws, row, 6, f'=SUM(F{hold_first}:F{hold_last})', font=BOLD_FONT, fmt=DOLLAR)
        _cell(ws, row, 7, f'=F{row}/E{row}', font=BOLD_FONT, fmt='0.00x')
        names["holdings_total"] = ("F", row)
        total_row = row
        row += 1

        names["_hold_first"] = hold_first
        names["_hold_last"] = hold_last
        names["_total_row"] = total_row
        return row, names

    # 401(k) live Plaid holdings: Fund Name, Current Value, Cost Basis, Gain/Loss, Return %
    if acct.get("_401k_live_plaid"):
        _header_row(ws, row, ['Security', 'Quantity', 'Price', 'Market Value', 'Cost Basis', 'Gain/Loss', 'Return %'])
        row += 1
        hold_first = row
        for h in holdings:
            qty = h.get("qty", 0) or 0
            mv = h.get("mv", 0) or 0
            price = round(mv / qty, 2) if qty else 0
            _cell(ws, row, 1, h["ticker"])  # fund name stored in ticker field
            _cell(ws, row, 2, qty, fmt=QTY_FMT)
            _cell(ws, row, 3, price, fmt=DOLLAR)
            _cell(ws, row, 4, mv, fmt=DOLLAR)
            _cell(ws, row, 5, h.get("cb", 0), fmt=DOLLAR)
            _cell(ws, row, 6, f'=D{row}-E{row}', font=BLACK_FONT, fmt=DOLLAR)
            _cell(ws, row, 7, f'=IF(E{row}=0,"",F{row}/E{row})', font=BLACK_FONT, fmt=PCT)
            row += 1
        hold_last = row - 1

        _cell(ws, row, 1, 'TOTAL', font=BOLD_FONT)
        _cell(ws, row, 4, f'=SUM(D{hold_first}:D{hold_last})', font=BOLD_FONT, fmt=DOLLAR)
        _cell(ws, row, 5, f'=SUM(E{hold_first}:E{hold_last})', font=BOLD_FONT, fmt=DOLLAR)
        _cell(ws, row, 6, f'=SUM(F{hold_first}:F{hold_last})', font=BOLD_FONT, fmt=DOLLAR)
        _cell(ws, row, 7, f'=IF(E{row}=0,"",F{row}/E{row})', font=BOLD_FONT, fmt=PCT)
        names["holdings_total"] = ("D", row)
        total_row = row
        row += 1

        names["_hold_first"] = hold_first
        names["_hold_last"] = hold_last
        names["_total_row"] = total_row
        return row, names

    # 401(k) statement holdings (fallback): Fund Name, Beginning, Ending, Gain
    if holdings and isinstance(holdings[0], dict) and "name" in holdings[0]:
        _header_row(ws, row, ['Fund', 'Beginning Value', 'Ending Value', 'Gain/Loss'])
        row += 1
        hold_first = row
        for h in holdings:
            _cell(ws, row, 1, h["name"])
            _cell(ws, row, 2, h.get("beginning", 0), fmt=DOLLAR)
            _cell(ws, row, 3, h.get("ending", 0), fmt=DOLLAR)
            _cell(ws, row, 4, h.get("gain", 0), fmt=DOLLAR)
            row += 1
        hold_last = row - 1

        _cell(ws, row, 1, 'TOTAL', font=BOLD_FONT)
        _cell(ws, row, 2, f'=SUM(B{hold_first}:B{hold_last})', font=BOLD_FONT, fmt=DOLLAR)
        _cell(ws, row, 3, f'=SUM(C{hold_first}:C{hold_last})', font=BOLD_FONT, fmt=DOLLAR)
        _cell(ws, row, 4, f'=SUM(D{hold_first}:D{hold_last})', font=BOLD_FONT, fmt=DOLLAR)
        names["holdings_total"] = ("C", row)
        total_row = row
        row += 1

        names["_hold_first"] = hold_first
        names["_hold_last"] = hold_last
        names["_total_row"] = total_row
        return row, names

    # Standard liquid account holdings
    if is_rh:
        _header_row(ws, row, ['Security', 'Quantity', 'Price', 'Market Value', 'Average Cost', 'Cost Basis', 'Gain/Loss', 'Return %'])
    else:
        _header_row(ws, row, ['Security', 'Quantity', 'Price', 'Market Value', 'Cost Basis', 'Gain/Loss', 'Return %'])
    row += 1

    hold_first = row
    for h in holdings:
        _cell(ws, row, 1, h["ticker"])
        _cell(ws, row, 2, h["qty"], fmt=QTY_FMT if not isinstance(h["qty"], int) else '#,##0')
        _cell(ws, row, 3, h["price"], fmt=DOLLAR)
        _cell(ws, row, 4, h["mv"], fmt=DOLLAR)
        if is_rh:
            _cell(ws, row, 5, h.get("avg_cost", 0), fmt=DOLLAR)
            _cell(ws, row, 6, h.get("cb", 0), fmt=DOLLAR)
            _cell(ws, row, 7, f'=D{row}-F{row}', font=BLACK_FONT, fmt=DOLLAR)
            _cell(ws, row, 8, f'=IF(F{row}=0,"",G{row}/F{row})', font=BLACK_FONT, fmt=PCT)
        else:
            _cell(ws, row, 5, h.get("cb", 0), fmt=DOLLAR)
            _cell(ws, row, 6, f'=D{row}-E{row}', font=BLACK_FONT, fmt=DOLLAR)
            _cell(ws, row, 7, f'=IF(E{row}=0,"",F{row}/E{row})', font=BLACK_FONT, fmt=PCT)
        row += 1

    hold_last = row - 1

    # Cash position row if applicable
    cash_pos = acct.get("cash_position", 0)
    if cash_pos:
        _cell(ws, row, 1, 'Cash')
        _cell(ws, row, 4, cash_pos, fmt=DOLLAR)
        row += 1
        hold_last = row - 1

    # TOTAL row
    if is_rh:
        _cell(ws, row, 1, 'TOTAL SECURITIES', font=BOLD_FONT)
        _cell(ws, row, 4, f'=SUM(D{hold_first}:D{hold_last})', font=BOLD_FONT, fmt=DOLLAR)
        _cell(ws, row, 6, f'=SUM(F{hold_first}:F{hold_last})', font=BOLD_FONT, fmt=DOLLAR)
        _cell(ws, row, 7, f'=SUM(G{hold_first}:G{hold_last})', font=BOLD_FONT, fmt=DOLLAR)
        _cell(ws, row, 8, f'=IF(F{row}=0,"",G{row}/F{row})', font=BOLD_FONT, fmt=PCT)
        names["holdings_total_mv"] = ("D", row)
        names["holdings_total_cb"] = ("F", row)
        names["holdings_total_gl"] = ("G", row)
    else:
        _cell(ws, row, 1, 'TOTAL', font=BOLD_FONT)
        _cell(ws, row, 4, f'=SUM(D{hold_first}:D{hold_last})', font=BOLD_FONT, fmt=DOLLAR)
        _cell(ws, row, 5, f'=SUM(E{hold_first}:E{hold_last})', font=BOLD_FONT, fmt=DOLLAR)
        _cell(ws, row, 6, f'=SUM(F{hold_first}:F{hold_last})', font=BOLD_FONT, fmt=DOLLAR)
        _cell(ws, row, 7, f'=IF(E{row}=0,"",F{row}/E{row})', font=BOLD_FONT, fmt=PCT)
        names["holdings_total"] = ("D", row)
    total_row = row
    row += 1

    # Margin details for Robinhood
    if acct.get("is_margin"):
        margin = acct.get("margin_debt", 0)
        _cell(ws, row, 1, 'Margin Debt')
        _cell(ws, row, 4, margin, fmt=DOLLAR)
        names["margin_debt"] = ("D", row)
        row += 1
        _cell(ws, row, 1, 'NET PORTFOLIO VALUE', font=BOLD_FONT)
        _cell(ws, row, 4, f'=D{total_row}+D{row-1}', font=BOLD_FONT, fmt=DOLLAR)
        names["net_portfolio"] = ("D", row)
        row += 1

    names["_hold_first"] = hold_first
    names["_hold_last"] = hold_last
    names["_total_row"] = total_row

    return row, names


def _build_monthly_section(ws, row, acct, model, wb):
    """MONTHLY CALCULATIONS section."""
    names = {}
    monthly = acct.get("monthly", {})
    labels = acct.get("cash_flow_labels", {"add": "Additions", "sub": "Subtractions"})

    _header_row(ws, row, ['Month', 'Beginning Value', labels["add"], labels["sub"],
                          'Dividends', 'Market Change', 'Ending Value', 'Monthly Return', 'Growth Factor'])
    row += 1

    monthly_first = row
    for m in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']:
        _cell(ws, row, 1, m)
        if m in monthly:
            d = monthly[m]
            mkt = (d.get("change", 0) or 0) - (d.get("div", 0) or 0)
            _cell(ws, row, 2, d.get("begin", 0), fmt=DOLLAR)
            _cell(ws, row, 3, d.get("add", 0), fmt=DOLLAR)
            _cell(ws, row, 4, d.get("sub", 0), fmt=DOLLAR)
            _cell(ws, row, 5, d.get("div", 0), fmt=DOLLAR)
            _cell(ws, row, 6, round(mkt, 2), fmt=DOLLAR)
            _cell(ws, row, 7, d.get("end", 0), fmt=DOLLAR)
        else:
            for col in range(2, 8):
                _cell(ws, row, col, None, fmt=DOLLAR)
        _cell(ws, row, 8, f'=IF(B{row}=0,"",((G{row}+D{row}-C{row})/B{row})-1)', font=BLACK_FONT, fmt=PCT)
        _cell(ws, row, 9, f'=IF(H{row}="","",1+H{row})', font=BLACK_FONT, fmt=N_FMT)
        row += 1

    monthly_last = row - 1
    row += 1

    # Totals
    _cell(ws, row, 1, 'Totals', font=BOLD_FONT)
    for col in [3, 4, 5, 6]:
        _cell(ws, row, col, f'=SUM({get_column_letter(col)}{monthly_first}:{get_column_letter(col)}{monthly_last})',
              font=BOLD_FONT, fmt=DOLLAR)

    names["monthly_jan"] = ("B", monthly_first)
    names["monthly_dec"] = ("B", monthly_last)
    names["monthly_totals"] = ("B", row)
    names["_monthly_first"] = monthly_first
    names["_monthly_last"] = monthly_last
    names["_monthly_totals_row"] = row
    row += 1

    return row, names


def _build_sold_section(ws, row, acct, model, wb):
    """SOLD POSITIONS section."""
    names = {}
    _header_row(ws, row, ['Security', 'Date', 'Quantity', 'Cost Basis', 'Proceeds', 'Realized Gain/Loss', 'Action'])
    row += 1

    sold = acct.get("sold", [])

    # Handle empty sold gracefully
    if not sold:
        _cell(ws, row, 1, '(none)', font=NOTE_FONT)
        row += 1
        return row, names

    is_dict = isinstance(sold, dict)

    def _write_sold_group(ws, row, year_label, positions):
        _cell(ws, row, 1, year_label, font=BOLD_FONT)
        for col in range(2, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
        for s in positions:
            _cell(ws, row, 1, s["ticker"])
            _cell(ws, row, 2, s["date"])
            _cell(ws, row, 3, s["qty"], fmt=QTY_FMT)
            if s.get("cb") is not None:
                _cell(ws, row, 4, s["cb"], fmt=DOLLAR)
            _cell(ws, row, 5, s.get("proceeds", 0), fmt=DOLLAR)
            if s.get("cb") is not None:
                _cell(ws, row, 6, f'=E{row}-D{row}', font=BLACK_FONT, fmt=DOLLAR)
            _cell(ws, row, 7, s.get("action", ""), font=NOTE_FONT)
            row += 1
        return row

    if is_dict:
        for year in sorted(sold.keys(), reverse=True):
            row = _write_sold_group(ws, row, year, sold[year])
            # Total for this year
            group_first = row - len(sold[year])
            _cell(ws, row, 1, f'{year} TOTAL', font=BOLD_FONT)
            _cell(ws, row, 5, f'=SUM(E{group_first}:E{row-1})', font=BOLD_FONT, fmt=DOLLAR)
            _cell(ws, row, 6, f'=SUM(F{group_first}:F{row-1})', font=BOLD_FONT, fmt=DOLLAR)
            if year == "2026":
                names["sold_2026_total"] = ("F", row)
            row += 2
    else:
        row = _write_sold_group(ws, row, '2026', sold)
        group_first = row - len(sold)
        _cell(ws, row, 1, '2026 TOTAL', font=BOLD_FONT)
        _cell(ws, row, 5, f'=SUM(E{group_first}:E{row-1})', font=BOLD_FONT, fmt=DOLLAR)
        _cell(ws, row, 6, f'=SUM(F{group_first}:F{row-1})', font=BOLD_FONT, fmt=DOLLAR)
        names["sold_2026_total"] = ("F", row)
        row += 1

    return row, names


def _fix_forward_refs(ws, row_map, acct):
    """Fill in forward-reference cells now that all row numbers are known."""
    # TWR = PRODUCT of growth factors
    if "_monthly_first" in row_map and "TWR" in row_map:
        mf = row_map["_monthly_first"]
        ml = row_map["_monthly_last"]
        twr_r = row_map["TWR"][1]
        ws.cell(row=twr_r, column=2, value=f'=IFERROR(PRODUCT(I{mf}:I{ml})-1,"")')
        ws.cell(row=twr_r, column=2).font = BLACK_FONT
        ws.cell(row=twr_r, column=2).number_format = PCT
    elif "TWR" in row_map:
        # No monthly section (e.g. 401k, angel) — write computed value directly
        twr_r = row_map["TWR"][1]
        twr_val = acct.get("returns", {}).get("twr")
        ws.cell(row=twr_r, column=2, value=twr_val)
        ws.cell(row=twr_r, column=2).font = BLACK_FONT
        ws.cell(row=twr_r, column=2).number_format = PCT

    # CB Return
    if "cb_return" in row_map and "_total_row" in row_map:
        cbr_r = row_map["cb_return"][1]
        tr = row_map["_total_row"]
        if acct.get("is_margin"):
            # Robinhood: G/F columns
            ws.cell(row=cbr_r, column=2, value=f'=G{tr}/F{tr}')
        elif acct.get("investments"):
            # Angel: F/E columns
            ws.cell(row=cbr_r, column=2, value=f'=F{tr}/E{tr}-1')
        elif acct.get("holdings") and acct["holdings"] and "name" in acct["holdings"][0]:
            # 401(k): computed value
            ws.cell(row=cbr_r, column=2, value=acct.get("returns", {}).get("cb_return"))
        else:
            ws.cell(row=cbr_r, column=2, value=f'=F{tr}/E{tr}')
        ws.cell(row=cbr_r, column=2).font = BLACK_FONT
        ws.cell(row=cbr_r, column=2).number_format = PCT

    # Dividends = monthly totals div column
    if "dividends" in row_map and "_monthly_totals_row" in row_map:
        div_r = row_map["dividends"][1]
        mt = row_map["_monthly_totals_row"]
        ws.cell(row=div_r, column=2, value=f'=E{mt}')
        ws.cell(row=div_r, column=2).font = BLACK_FONT
        ws.cell(row=div_r, column=2).number_format = DOLLAR
    elif "dividends" in row_map:
        # No monthly section — write computed value
        div_r = row_map["dividends"][1]
        div_val = acct.get("gains", {}).get("dividends", 0)
        ws.cell(row=div_r, column=2, value=div_val)
        ws.cell(row=div_r, column=2).font = BLACK_FONT
        ws.cell(row=div_r, column=2).number_format = DOLLAR

    # Unrealized = holdings total G/L
    if "unrealized" in row_map and "_total_row" in row_map:
        ur = row_map["unrealized"][1]
        tr = row_map["_total_row"]
        if acct.get("is_margin"):
            gl_col = "G"
        elif acct.get("investments"):
            # Angel: unrealized = total current - total invested
            ws.cell(row=ur, column=2, value=f'=F{tr}-E{tr}')
            ws.cell(row=ur, column=2).font = BLACK_FONT
            ws.cell(row=ur, column=2).number_format = DOLLAR
            ur = None  # skip the default write below
        elif acct.get("holdings") and acct["holdings"] and "name" in acct["holdings"][0]:
            # 401(k): gain column is D
            gl_col = "D"
        else:
            gl_col = "F"
        if ur is not None:
            ws.cell(row=ur, column=2, value=f'={gl_col}{tr}')
            ws.cell(row=ur, column=2).font = BLACK_FONT
            ws.cell(row=ur, column=2).number_format = DOLLAR

    # Realized = sold total
    if "realized" in row_map and "sold_2026_total" in row_map:
        rr = row_map["realized"][1]
        st = row_map["sold_2026_total"][1]
        ws.cell(row=rr, column=2, value=f'=F{st}')
        ws.cell(row=rr, column=2).font = BLACK_FONT
        ws.cell(row=rr, column=2).number_format = DOLLAR
    elif "realized" in row_map:
        # No sold section or no sold — write computed value or 0
        rr = row_map["realized"][1]
        realized_val = acct.get("gains", {}).get("realized", 0)
        ws.cell(row=rr, column=2, value=realized_val)
        ws.cell(row=rr, column=2).font = BLACK_FONT
        ws.cell(row=rr, column=2).number_format = DOLLAR

    # Total additions/subtractions
    if "total_add" in row_map and "_monthly_totals_row" in row_map:
        mt = row_map["_monthly_totals_row"]
        ws.cell(row=row_map["total_add"][1], column=2, value=f'=C{mt}')
        ws.cell(row=row_map["total_add"][1], column=2).font = BLACK_FONT
        ws.cell(row=row_map["total_add"][1], column=2).number_format = DOLLAR
    if "total_sub" in row_map and "_monthly_totals_row" in row_map:
        mt = row_map["_monthly_totals_row"]
        ws.cell(row=row_map["total_sub"][1], column=2, value=f'=D{mt}')
        ws.cell(row=row_map["total_sub"][1], column=2).font = BLACK_FONT
        ws.cell(row=row_map["total_sub"][1], column=2).number_format = DOLLAR


def build_account_tab(wb, acct, model):
    """Build a complete account tab."""
    tab_name = acct["tab_name"]
    title = f'{acct["name"]} — {model["year"]} Performance'
    subtitle = 'Blue = hardcoded from statement | Black = formula'

    # Determine which sections apply to this account
    has_monthly = bool(acct.get("monthly", {}))
    has_sold = bool(acct.get("sold", []) if not isinstance(acct.get("sold"), dict) else acct.get("sold"))
    has_holdings = bool(acct.get("holdings", []) or acct.get("investments", []))

    sections = [
        ("YTD RETURN CALCULATIONS", _build_return_section),
        ("YTD INVESTMENT GAIN SUMMARY", _build_gain_section),
    ]

    if has_holdings:
        sections.append(("CURRENT HOLDINGS", _build_holdings_section))

    if has_monthly:
        sections.append(("MONTHLY CALCULATIONS", _build_monthly_section))

    if has_sold:
        sections.append(("SOLD POSITIONS", _build_sold_section))

    col_widths = {'A': 26, 'B': 16, 'C': 16, 'D': 16, 'E': 16, 'F': 19, 'G': 16, 'H': 16, 'I': 14}

    ws, row_map = write_sections(wb, tab_name, title, subtitle, sections, col_widths, acct, model)

    # Fix forward references
    _fix_forward_refs(ws, row_map, acct)

    return ws, row_map


# ---------------------------------------------------------------------------
# Dashboard tab
# ---------------------------------------------------------------------------
def _get_acct_beginning_ending(acct):
    """Extract beginning and ending values from an account model entry.

    Beginning = start-of-year value (Jan beginning from monthly data, or Q1 beginning)
    Ending = latest account value (current holdings MV + cash position)
    Net Cash Flow = total additions - total subtractions YTD
    """
    gains = acct.get("gains", {})
    monthly = acct.get("monthly", {})
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    populated = [m for m in months if m in monthly]

    # Beginning = start-of-year (Jan beginning)
    begin = 0
    if populated:
        begin = monthly[populated[0]].get("begin", 0) or 0
    elif acct.get("quarterly"):
        begin = acct["quarterly"][0].get("beginning", 0) or 0

    # Ending = latest account value (holdings MV + cash position)
    total_mv = gains.get("total_mv", 0) or 0
    cash_pos = acct.get("cash_position", 0) or 0
    margin_debt = acct.get("margin_debt", 0) or 0
    end = total_mv + cash_pos + margin_debt  # margin_debt is negative

    # Angel — use gains totals
    if "total_invested" in gains:
        begin = gains["total_invested"]
        end = gains["total_current"]

    # Net cash flow = total additions - total subtractions
    ncf = 0
    if populated:
        ncf = sum(monthly.get(m, {}).get("add", 0) or 0 for m in populated) - \
              sum(monthly.get(m, {}).get("sub", 0) or 0 for m in populated)
    elif acct.get("quarterly"):
        ncf = sum(
            (q.get("ee_contributions", 0) or 0) + (q.get("er_contributions", 0) or 0) + (q.get("fees", 0) or 0)
            for q in acct["quarterly"]
        )

    return begin, end, ncf


def build_dashboard(wb, model):
    """Build the Dashboard tab with all summary sections."""
    ws = wb.create_sheet("Dashboard")
    tab_name = "Dashboard"

    # Column widths
    for col, w in {'A': 48, 'B': 16, 'C': 26, 'D': 37, 'E': 29,
                   'F': 30, 'G': 26, 'H': 28, 'I': 14, 'J': 12}.items():
        ws.column_dimensions[col].width = w

    row = 1
    today = datetime.date.today()
    ws.cell(row=row, column=1,
            value=f"Portfolio Dashboard — {today.strftime('%B %d, %Y')}").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="All values are computed from account data. No cross-sheet formulas.").font = NOTE_FONT
    row += 2

    names = {}

    # ==================================================================
    # 1. DAILY SUMMARY
    # ==================================================================
    row = _section_header(ws, row, "DAILY SUMMARY")
    ds = model.get("daily_summary")
    if ds:
        total_val = ds.get("total_value", 0)
        change = ds.get("change", 0)
        pct = ds.get("change_pct", 0)
        prev_date = ds.get("prev_date", "")
        prose = (f"Your liquid portfolio is valued at ${total_val:,.2f}. "
                 f"This represents a ${change:,.2f} ({pct:.2%}) "
                 f"{'gain' if change >= 0 else 'loss'} from the last trading day ({prev_date}).")
        movers = ds.get("top_movers", "")
        if movers:
            prose += f" Top movers: {movers}"
        c = ws.cell(row=row, column=1, value=prose)
        c.font = PROSE_FONT
        c.alignment = Alignment(wrap_text=True)
    else:
        ws.cell(row=row, column=1,
                value="Daily summary requires at least 2 pipeline runs.").font = PROSE_FONT
    row += 2

    # ==================================================================
    # 2. YTD BENCHMARK COMPARISON
    # ==================================================================
    row = _section_header(ws, row, "YTD BENCHMARK COMPARISON")
    bench_headers = ['Benchmark', 'Return',
                     'Alpha: Liquid Portfolio',
                     'Alpha: Fidelity Brokerage (liquid)',
                     'Alpha: Roth IRA (liquid)',
                     'Alpha: HSA (liquid)',
                     'Alpha: Robinhood (liquid)',
                     'Alpha: 401(k) (illiquid)']
    _header_row(ws, row, bench_headers)
    row += 1

    benchmarks = model.get("benchmarks", {})
    liquid_twr = model.get("liquid_twr") or 0

    # Map alpha columns to account keys and their column index
    alpha_acct_keys = [
        ("fidelity_brokerage", 3),
        ("fidelity_roth_ira", 4),
        ("fidelity_hsa", 5),
        ("robinhood", 6),
        ("k401", 7),
    ]

    bench_names = ["S&P 500", "Dow Jones", "NASDAQ"]
    for bname in bench_names:
        bret = benchmarks.get(bname)
        if bret is None:
            _cell(ws, row, 1, bname)
            _cell(ws, row, 2, "N/A", font=NOTE_FONT)
            row += 1
            continue
        _cell(ws, row, 1, bname)
        _cell(ws, row, 2, bret, fmt=PCT)
        # Liquid Portfolio alpha
        _cell(ws, row, 3, liquid_twr - bret, fmt=PCT)
        # Per-account alphas
        for acct_key, col_offset in alpha_acct_keys:
            acct = model["accounts"].get(acct_key)
            if acct:
                acct_twr = acct["returns"].get("twr")
                if acct_twr is not None:
                    _cell(ws, row, col_offset + 1, acct_twr - bret, fmt=PCT)
                else:
                    _cell(ws, row, col_offset + 1, "N/A", font=NOTE_FONT)
            else:
                _cell(ws, row, col_offset + 1, "N/A", font=NOTE_FONT)
        names[f"bench_{bname.replace(' ', '_').lower()}"] = ("B", row)
        row += 1
    row += 1

    # ==================================================================
    # 3. YTD INVESTMENT GAIN
    # ==================================================================
    row = _section_header(ws, row, "YTD INVESTMENT GAIN")
    _header_row(ws, row, ['Metric', 'Value', 'Note'])
    row += 1

    # Sum dividends, unrealized, realized across liquid accounts
    total_div = 0
    total_unr = 0
    total_rea = 0
    for key in model["liquid_accounts"]:
        g = model["accounts"][key]["gains"]
        total_div += g.get("dividends", 0) or 0
        total_unr += g.get("unrealized", 0) or 0
        total_rea += g.get("realized", 0) or 0

    _cell(ws, row, 1, "Dividends/Income (liquid)")
    _cell(ws, row, 2, round(total_div, 2), fmt=DOLLAR)
    _cell(ws, row, 3, "Sum across liquid accounts", font=NOTE_FONT)
    names["ytd_dividends"] = ("B", row)
    row += 1

    _cell(ws, row, 1, "Unrealized Gain/Loss (liquid)")
    _cell(ws, row, 2, round(total_unr, 2), fmt=DOLLAR)
    names["ytd_unrealized"] = ("B", row)
    row += 1

    _cell(ws, row, 1, "Realized Gain/Loss (liquid)")
    _cell(ws, row, 2, round(total_rea, 2), fmt=DOLLAR)
    names["ytd_realized"] = ("B", row)
    row += 1

    # 401(k) investment gain
    k401_gain = 0
    k401 = model["accounts"].get("k401")
    if k401 and k401.get("quarterly"):
        k401_gain = sum(q.get("change_in_value", 0) or 0 for q in k401["quarterly"])
    _cell(ws, row, 1, "401(k) Investment Gain")
    _cell(ws, row, 2, round(k401_gain, 2), fmt=DOLLAR)
    _cell(ws, row, 3, "From quarterly change in value", font=NOTE_FONT)
    names["ytd_k401_gain"] = ("B", row)
    row += 1

    total_ytd = total_div + total_unr + total_rea + k401_gain
    _cell(ws, row, 1, "Total YTD Investment Gain", font=BOLD_FONT)
    _cell(ws, row, 2, round(total_ytd, 2), font=BOLD_FONT, fmt=DOLLAR)
    names["ytd_total_gain"] = ("B", row)
    row += 2

    # ==================================================================
    # 4. ACCOUNT OVERVIEW
    # ==================================================================
    row = _section_header(ws, row, "ACCOUNT OVERVIEW")
    overview_headers = ['Account', 'Beginning', 'Market Value', 'Net Cash Flow',
                        'Time-Weighted Return (YTD)', 'Money-Weighted Return (YTD)',
                        'Cost Basis Return', 'Alpha (YTD)']
    _header_row(ws, row, overview_headers)
    row += 1

    sp500_ret = benchmarks.get("S&P 500", 0) or 0

    # Helper to write one account row
    def _write_overview_row(ws, row, label, begin, end, ncf, twr, mwrr, cb_ret, alpha):
        _cell(ws, row, 1, label)
        _cell(ws, row, 2, round(begin, 2) if begin else 0, fmt=DOLLAR)
        _cell(ws, row, 3, round(end, 2) if end else 0, fmt=DOLLAR)
        _cell(ws, row, 4, round(ncf, 2) if ncf else 0, fmt=DOLLAR)
        _cell(ws, row, 5, twr, fmt=PCT) if twr is not None else _cell(ws, row, 5, "N/A", font=NOTE_FONT)
        _cell(ws, row, 6, mwrr, fmt=PCT) if mwrr is not None else _cell(ws, row, 6, "N/A", font=NOTE_FONT)
        _cell(ws, row, 7, cb_ret, fmt=PCT) if cb_ret is not None else _cell(ws, row, 7, "N/A", font=NOTE_FONT)
        if alpha is not None:
            _cell(ws, row, 8, alpha, fmt=PCT)
        else:
            _cell(ws, row, 8, "N/A", font=NOTE_FONT)

    # Liquid accounts
    liquid_begin = 0
    liquid_end = 0
    liquid_ncf = 0

    acct_display_names = {
        "fidelity_brokerage": "Fidelity Brokerage",
        "fidelity_roth_ira": "Roth IRA",
        "fidelity_hsa": "HSA",
        "robinhood": "Robinhood",
        "k401": "401(k)",
        "angel": "Angel Investments",
    }

    for key in model["liquid_accounts"]:
        acct = model["accounts"][key]
        begin, end, ncf = _get_acct_beginning_ending(acct)
        twr = acct["returns"]["twr"]
        mwrr = acct["returns"]["mwrr"]
        cb_ret = acct["returns"]["cb_return"]
        alpha = (twr - sp500_ret) if twr is not None else None
        label = acct_display_names.get(key, acct["name"])
        _write_overview_row(ws, row, label, begin, end, ncf, twr, mwrr, cb_ret, alpha)
        names[f"overview_{key}"] = ("C", row)
        liquid_begin += begin
        liquid_end += end
        liquid_ncf += ncf
        row += 1

    # LIQUID SUBTOTAL
    _cell(ws, row, 1, "LIQUID SUBTOTAL", font=BOLD_FONT)
    _cell(ws, row, 2, round(liquid_begin, 2), font=BOLD_FONT, fmt=DOLLAR)
    _cell(ws, row, 3, round(liquid_end, 2), font=BOLD_FONT, fmt=DOLLAR)
    _cell(ws, row, 4, round(liquid_ncf, 2), font=BOLD_FONT, fmt=DOLLAR)
    _cell(ws, row, 5, liquid_twr, font=BOLD_FONT, fmt=PCT) if liquid_twr else _cell(ws, row, 5, "N/A", font=BOLD_FONT)
    _cell(ws, row, 6, "", font=BOLD_FONT)  # no aggregated MWRR
    _cell(ws, row, 7, "", font=BOLD_FONT)
    liquid_alpha = (liquid_twr - sp500_ret) if liquid_twr else None
    if liquid_alpha is not None:
        _cell(ws, row, 8, liquid_alpha, font=BOLD_FONT, fmt=PCT)
    else:
        _cell(ws, row, 8, "N/A", font=BOLD_FONT)
    names["overview_liquid_subtotal"] = ("C", row)
    row += 2  # skip row

    # EXTERNAL CASH
    ext_cash_total = sum(
        (v.get("total", 0) if isinstance(v, dict) else v)
        for v in model["cash"].get("external", {}).values()
    )
    _cell(ws, row, 1, "EXTERNAL CASH", font=BOLD_FONT)
    _cell(ws, row, 3, round(ext_cash_total, 2), font=BOLD_FONT, fmt=DOLLAR)
    names["overview_ext_cash"] = ("C", row)
    row += 2  # skip row

    # Illiquid accounts
    illiquid_begin = 0
    illiquid_end = 0
    illiquid_ncf = 0
    for key in model["illiquid_accounts"]:
        acct = model["accounts"][key]
        begin, end, ncf = _get_acct_beginning_ending(acct)
        twr = acct["returns"]["twr"]
        mwrr = acct["returns"]["mwrr"]
        cb_ret = acct["returns"]["cb_return"]
        alpha = (twr - sp500_ret) if twr is not None else None
        label = acct_display_names.get(key, acct["name"])
        _write_overview_row(ws, row, label, begin, end, ncf, twr, mwrr, cb_ret, alpha)
        names[f"overview_{key}"] = ("C", row)
        illiquid_begin += begin
        illiquid_end += end
        illiquid_ncf += ncf
        row += 1

    # ILLIQUID SUBTOTAL
    illiquid_cb = 0
    illiquid_mv = 0
    for key in model["illiquid_accounts"]:
        g = model["accounts"][key]["gains"]
        illiquid_cb += g.get("total_cb", 0) or 0
        illiquid_mv += g.get("total_mv", 0) or 0
    illiquid_cb_return = (illiquid_mv - illiquid_cb) / illiquid_cb if illiquid_cb else None

    _cell(ws, row, 1, "ILLIQUID SUBTOTAL", font=BOLD_FONT)
    _cell(ws, row, 2, round(illiquid_begin, 2), font=BOLD_FONT, fmt=DOLLAR)
    _cell(ws, row, 3, round(illiquid_end, 2), font=BOLD_FONT, fmt=DOLLAR)
    _cell(ws, row, 4, round(illiquid_ncf, 2), font=BOLD_FONT, fmt=DOLLAR)
    _cell(ws, row, 5, "", font=BOLD_FONT)
    _cell(ws, row, 6, "", font=BOLD_FONT)
    if illiquid_cb_return is not None:
        _cell(ws, row, 7, illiquid_cb_return, font=BOLD_FONT, fmt=PCT)
    else:
        _cell(ws, row, 7, "N/A", font=BOLD_FONT)
    _cell(ws, row, 8, "", font=BOLD_FONT)
    names["overview_illiquid_subtotal"] = ("C", row)
    row += 2  # skip row

    # TOTAL PORTFOLIO
    total_begin = liquid_begin + illiquid_begin
    total_end = liquid_end + illiquid_end + ext_cash_total
    total_ncf = liquid_ncf + illiquid_ncf
    _cell(ws, row, 1, "TOTAL PORTFOLIO", font=BOLD_FONT)
    _cell(ws, row, 2, round(total_begin, 2), font=BOLD_FONT, fmt=DOLLAR)
    _cell(ws, row, 3, round(total_end, 2), font=BOLD_FONT, fmt=DOLLAR)
    _cell(ws, row, 4, round(total_ncf, 2), font=BOLD_FONT, fmt=DOLLAR)
    names["overview_total"] = ("C", row)
    row += 2

    # ==================================================================
    # 5. SECTOR CONCENTRATION
    # ==================================================================
    row = _section_header(ws, row, "SECTOR CONCENTRATION")
    sector_acct_cols = ['Fidelity Brokerage', 'Roth IRA', '401(k)', 'HSA', 'Robinhood', 'Angel']
    sector_headers = ['Sector', 'Total Value', '% of Portfolio'] + sector_acct_cols + ['# Holdings']
    _header_row(ws, row, sector_headers)
    row += 1

    for s in model.get("sectors", []):
        _cell(ws, row, 1, s["name"])
        _cell(ws, row, 2, s["value"], fmt=DOLLAR)
        _cell(ws, row, 3, s["pct"], fmt=PCT)
        for i, acct_label in enumerate(sector_acct_cols):
            val = s.get("by_account", {}).get(acct_label, 0)
            _cell(ws, row, 4 + i, val, fmt=DOLLAR) if val else _cell(ws, row, 4 + i, 0, fmt=DOLLAR)
        _cell(ws, row, 4 + len(sector_acct_cols), s.get("count", 0))
        row += 1
    row += 1

    # ==================================================================
    # 6. GEOGRAPHIC CONCENTRATION
    # ==================================================================
    row = _section_header(ws, row, "GEOGRAPHIC CONCENTRATION")
    _header_row(ws, row, ['Region', 'Value', '% of Portfolio'])
    row += 1

    for g in model.get("geo", []):
        _cell(ws, row, 1, g["region"])
        _cell(ws, row, 2, g["value"], fmt=DOLLAR)
        _cell(ws, row, 3, g["pct"], fmt=PCT)
        row += 1
    row += 1

    # ==================================================================
    # 7. RISK METRICS
    # ==================================================================
    row = _section_header(ws, row, "RISK METRICS")
    ws.cell(row=row, column=1,
            value="Risk metrics require 12 months of monthly return data.").font = PROSE_FONT
    row += 2

    # ==================================================================
    # 8. RETURN METRIC DEFINITIONS
    # ==================================================================
    row = _section_header(ws, row, "RETURN METRIC DEFINITIONS")

    definitions = [
        ("Time-Weighted Return (TWR)",
         "Measures portfolio performance independent of cash flows. Chains monthly Modified Dietz "
         "returns: each month's return = (Ending - Beginning - Net Flow) / (Beginning + 0.5 * Net Flow). "
         "The YTD TWR is the product of (1 + monthly return) for each month, minus 1."),
        ("Money-Weighted Return (MWR)",
         "Measures the investor's actual experience including timing of deposits and withdrawals. "
         "Computed as the annualized Internal Rate of Return (IRR) of all cash flows. "
         "Higher MWR than TWR means the investor timed contributions well."),
        ("Cost Basis Return",
         "Simple return on invested capital: Unrealized Gain / Total Cost Basis. "
         "Does not account for time or cash flows. Useful for tax-lot analysis."),
        ("Alpha",
         "Excess return versus the S&P 500 benchmark. Alpha = Account TWR - S&P 500 YTD Return. "
         "Positive alpha means the account outperformed the index."),
    ]

    for label, defn in definitions:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        c = ws.cell(row=row, column=2, value=defn)
        c.font = PROSE_FONT
        c.alignment = Alignment(wrap_text=True)
        # Merge B through H for the definition
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 45
        row += 1
    row += 1

    # Define named ranges for Dashboard
    for key, val in names.items():
        if isinstance(val, tuple) and len(val) == 2:
            col, r = val
            _define_name(wb, tab_name, key, col, r)

    ws.sheet_view.showGridLines = False
    return ws


# ---------------------------------------------------------------------------
# Cash tab
# ---------------------------------------------------------------------------
def build_cash_tab(wb, model):
    """Build the Cash tab with external and embedded balances."""
    ws = wb.create_sheet("Cash")
    tab_name = "Cash"

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16

    row = 1
    ws.cell(row=row, column=1, value="Cash Accounts").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="Blue = hardcoded from Plaid | Black = formula").font = NOTE_FONT
    row += 2

    names = {}

    # Use cash data from model (already fetched once by pipeline)
    plaid_cash = {}
    ext = model["cash"].get("external", {})
    for label, data in ext.items():
        if isinstance(data, dict):
            plaid_cash[label] = data
        else:
            plaid_cash[label] = {"accounts": [{"name": label, "balance": data}], "total": data}

    # ==================================================================
    # EXTERNAL CASH ACCOUNTS
    # ==================================================================
    ws.cell(row=row, column=1, value="EXTERNAL CASH ACCOUNTS").font = SECTION_FONT
    row += 1
    _header_row(ws, row, ['Account', 'Institution', 'Balance'])
    row += 1

    plaid_labels = {'chase': 'Chase', 'marcus': 'Marcus (Goldman Sachs)'}
    ext_start = row

    for label, cdata in sorted(plaid_cash.items()):
        inst_name = plaid_labels.get(label, label)
        if isinstance(cdata, dict) and "accounts" in cdata:
            for acct in cdata["accounts"]:
                _cell(ws, row, 1, acct.get("name", label))
                _cell(ws, row, 2, inst_name)
                _cell(ws, row, 3, acct.get("balance", 0), fmt=DOLLAR)
                row += 1
        else:
            _cell(ws, row, 1, label)
            _cell(ws, row, 2, inst_name)
            _cell(ws, row, 3, cdata if isinstance(cdata, (int, float)) else 0, fmt=DOLLAR)
            row += 1

    if row == ext_start:
        ws.cell(row=row, column=1, value="No external cash data available.").font = NOTE_FONT
        row += 1

    _cell(ws, row, 1, "TOTAL EXTERNAL CASH", font=BOLD_FONT)
    _cell(ws, row, 2, "", font=BOLD_FONT)
    _cell(ws, row, 3, f'=SUM(C{ext_start}:C{row-1})', font=BOLD_FONT, fmt=DOLLAR)
    names["total_ext_cash"] = ("C", row)
    total_ext_row = row
    row += 2

    # ==================================================================
    # EMBEDDED CASH (for reference only)
    # ==================================================================
    ws.cell(row=row, column=1,
            value="EMBEDDED CASH (for reference only — included in account balances)").font = SECTION_FONT
    row += 1
    _header_row(ws, row, ['Account', 'Institution', 'Balance'])
    row += 1

    emb_labels = {
        "fidelity_brokerage": ("Fidelity Brokerage Core", "Fidelity"),
        "fidelity_roth_ira": ("Fidelity Roth IRA Core", "Fidelity"),
        "fidelity_hsa": ("Fidelity HSA Core", "Fidelity"),
    }

    emb_start = row
    embedded = model["cash"].get("embedded", {})
    for key, balance in embedded.items():
        if not balance:
            continue
        label_info = emb_labels.get(key, (key, "Unknown"))
        _cell(ws, row, 1, label_info[0])
        _cell(ws, row, 2, label_info[1])
        _cell(ws, row, 3, balance, fmt=DOLLAR)
        row += 1

    _cell(ws, row, 1, "TOTAL EMBEDDED CASH", font=BOLD_FONT)
    _cell(ws, row, 2, "", font=BOLD_FONT)
    _cell(ws, row, 3, f'=SUM(C{emb_start}:C{row-1})', font=BOLD_FONT, fmt=DOLLAR)
    names["total_emb_cash"] = ("C", row)
    total_emb_row = row
    row += 2

    # ==================================================================
    # CASH SUMMARY
    # ==================================================================
    ws.cell(row=row, column=1, value="CASH SUMMARY").font = SECTION_FONT
    row += 1
    _cell(ws, row, 1, "External Cash (counts toward portfolio)")
    _cell(ws, row, 2, f'=C{total_ext_row}', font=BLACK_FONT, fmt=DOLLAR)
    row += 1
    _cell(ws, row, 1, "Embedded Cash (already in account balances)")
    _cell(ws, row, 2, f'=C{total_emb_row}', font=BLACK_FONT, fmt=DOLLAR)
    row += 1
    _cell(ws, row, 1, "Total Cash (all sources)", font=BOLD_FONT)
    _cell(ws, row, 2, f'=B{row-2}+B{row-1}', font=BOLD_FONT, fmt=DOLLAR)
    names["total_all_cash"] = ("B", row)

    # Define named ranges
    for key, val in names.items():
        if isinstance(val, tuple) and len(val) == 2:
            col, r = val
            _define_name(wb, tab_name, key, col, r)

    ws.sheet_view.showGridLines = False
    return ws


# ---------------------------------------------------------------------------
# Angel Investments tab (dedicated layout — not a brokerage account)
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# 401(k) tab — standard monthly structure + live Plaid fund holdings
# ---------------------------------------------------------------------------
def build_401k_tab(wb, acct, model):
    """Build the 401(k) tab using standard account structure with monthly calculations.

    If live Plaid holdings are available, they're converted to the standard
    holdings format so they render in the correct position (before monthly calc).
    """
    # Convert live Plaid holdings into the standard format the generic builder expects
    live_holdings = acct.get("live_holdings", [])
    if live_holdings:
        acct = dict(acct)  # shallow copy so we don't mutate the model
        acct["holdings"] = [
            {
                "ticker": h["name"],  # fund name as ticker
                "qty": h.get("quantity", 0),
                "price": round(h["current_value"] / h["quantity"], 4) if h.get("quantity") else 0,
                "mv": h["current_value"],
                "cb": h["cost_basis"],
            }
            for h in live_holdings
        ]
        acct["_401k_live_plaid"] = True  # flag for custom header

    ws, row_map = build_account_tab(wb, acct, model)

    # Add TWR (Merrill Lynch Stated) note
    twr_stated = acct.get("twr_merrill_stated")
    if twr_stated is not None and "TWR" in row_map:
        twr_r = row_map["TWR"][1]
        _cell(ws, twr_r, 3, f'Merrill stated: {twr_stated:.2%}', font=NOTE_FONT)

    return ws, row_map


def build_angel_tab(wb, acct, model):
    """Build the Angel Investments tab with its unique investment-style layout.

    Columns: Company, Sector, Year, Series, Amount Invested,
    Post-Money at Investment, Latest Valuation, Valuation Source,
    Est. Current Value (formula), Return % (formula)
    """
    ws = wb.create_sheet("Angel Investments")
    tab_name = "Angel Investments"

    col_widths = {'A': 16, 'B': 14, 'C': 8, 'D': 12, 'E': 18,
                  'F': 22, 'G': 22, 'H': 22, 'I': 18, 'J': 12}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    row = 1
    ws.cell(row=row, column=1, value=f"Angel Investments — {model['year']} Portfolio Summary").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value="Blue = input data | Black = formula").font = NOTE_FONT
    row += 2

    _header_row(ws, row, ['Company', 'Sector', 'Year', 'Series', 'Amount Invested',
                          'Post-Money at Investment', 'Latest Valuation', 'Valuation Source',
                          'Est. Current Value', 'Return %'])
    row += 1

    investments = acct.get("investments", [])
    first = row
    for inv in investments:
        _cell(ws, row, 1, inv["company"])
        _cell(ws, row, 2, inv["sector"])
        _cell(ws, row, 3, inv["year"])
        _cell(ws, row, 4, inv["series"])
        _cell(ws, row, 5, inv["amount"], fmt=DOLLAR)
        _cell(ws, row, 6, inv["pm_invest"], fmt='$#,##0')
        _cell(ws, row, 7, inv["pm_latest"], fmt='$#,##0')
        _cell(ws, row, 8, inv["source"])
        _cell(ws, row, 9, f'=E{row}*(G{row}/F{row})', font=BLACK_FONT, fmt=DOLLAR)
        _cell(ws, row, 10, f'=(I{row}-E{row})/E{row}', font=BLACK_FONT, fmt=PCT)
        row += 1

    last = row - 1

    # TOTAL row
    _cell(ws, row, 1, 'TOTAL', font=BOLD_FONT)
    _cell(ws, row, 5, f'=SUM(E{first}:E{last})', font=BOLD_FONT, fmt=DOLLAR)
    _cell(ws, row, 9, f'=SUM(I{first}:I{last})', font=BOLD_FONT, fmt=DOLLAR)
    _cell(ws, row, 10, f'=(I{row}-E{row})/E{row}', font=BOLD_FONT, fmt=PCT)
    for c in range(1, 11):
        ws.cell(row=row, column=c).border = THIN_BORDER
    total_row = row
    row += 2

    # Valuation methodology notes
    ws.cell(row=row, column=1, value="Valuation Methodology:").font = BOLD_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="Est. Current Value = Amount Invested x (Latest Valuation / Post-Money at Investment)").font = NOTE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="Investments at cost: no subsequent priced round; Latest Valuation = Post-Money at Investment.").font = NOTE_FONT

    # Define named ranges
    _define_name(wb, tab_name, "total_invested", "E", total_row)
    _define_name(wb, tab_name, "total_current", "I", total_row)

    ws.sheet_view.showGridLines = False
    return ws


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------
def build(model, output_path="2026_Portfolio_Analysis.xlsx"):
    """Build the complete workbook from a portfolio model."""
    wb = openpyxl.Workbook()
    del wb['Sheet']  # remove default

    build_dashboard(wb, model)

    for key in model["liquid_accounts"]:
        build_account_tab(wb, model["accounts"][key], model)

    for key in model["illiquid_accounts"]:
        acct = model["accounts"][key]
        if "investments" in acct:
            build_angel_tab(wb, acct, model)
        elif "quarterly" in acct:
            build_401k_tab(wb, acct, model)
        else:
            build_account_tab(wb, acct, model)

    build_cash_tab(wb, model)

    # Gridlines off on all sheets
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    # Post-build verification (catches data gaps)
    errors = _verify_workbook(wb, model)
    if errors:
        import logging
        for e in errors:
            logging.warning(f"  Build verification: {e}")

    # Audit: verify data accuracy — only save if all checks pass
    audit_failures = _audit_workbook(wb, model)
    if audit_failures:
        import logging
        logging.error("DATA AUDIT FAILED — workbook NOT saved. Failures:")
        for f in audit_failures:
            logging.error(f"  {f}")
        raise ValueError(f"Data audit failed: {len(audit_failures)} check(s) failed. See log for details.")

    wb.save(output_path)
    return output_path


def _verify_workbook(wb, model):
    """Post-build sanity checks to catch data gaps before saving.

    Returns list of error strings. Empty list = all checks pass.
    """
    errors = []

    # 1. All expected tabs exist
    expected_tabs = (
        ["Dashboard"]
        + [model["accounts"][k]["tab_name"] for k in model["liquid_accounts"]]
        + [model["accounts"][k]["tab_name"] for k in model["illiquid_accounts"]]
        + ["Cash"]
    )
    for tab in expected_tabs:
        if tab not in wb.sheetnames:
            errors.append(f"Missing tab: {tab}")

    # 2. Dashboard has non-zero values for key rows
    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
        for r in range(1, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if not label:
                continue
            label_str = str(label)

            # External cash should have a non-zero value
            if "EXTERNAL CASH" in label_str.upper() and "ACCOUNT" not in label_str.upper():
                val = ws.cell(r, 3).value
                if val is None or (isinstance(val, (int, float)) and val == 0):
                    errors.append(f"Dashboard row {r} '{label_str}': external cash is 0 or empty")

            # Liquid/illiquid subtotals should have values
            if "SUBTOTAL" in label_str.upper():
                for c in [2, 3]:  # Beginning, Ending
                    val = ws.cell(r, c).value
                    if val is None or (isinstance(val, (int, float)) and val == 0):
                        errors.append(f"Dashboard row {r} '{label_str}' col {c}: subtotal is 0 or empty")

            # Benchmark returns should be non-zero
            if label_str in ("S&P 500", "Dow Jones", "NASDAQ"):
                val = ws.cell(r, 2).value
                if val is None or not isinstance(val, (int, float)):
                    errors.append(f"Dashboard row {r} '{label_str}': benchmark return missing")

    # 3. Each account tab has holdings
    for key in model["liquid_accounts"]:
        tab_name = model["accounts"][key]["tab_name"]
        if tab_name in wb.sheetnames:
            ws = wb[tab_name]
            has_holdings = False
            for r in range(1, ws.max_row + 1):
                v = ws.cell(r, 1).value
                if isinstance(v, str) and v == "CURRENT HOLDINGS":
                    has_holdings = True
                    break
            if not has_holdings:
                errors.append(f"Tab '{tab_name}': missing CURRENT HOLDINGS section")

    # 4. Cash tab has external cash accounts
    if "Cash" in wb.sheetnames:
        ws = wb["Cash"]
        has_external = False
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and "TOTAL EXTERNAL CASH" in v:
                val = ws.cell(r, 3).value
                if isinstance(val, str) and val.startswith("="):
                    has_external = True  # formula present, will evaluate in Excel
                elif isinstance(val, (int, float)) and val > 0:
                    has_external = True
                break
        if not has_external:
            errors.append("Cash tab: TOTAL EXTERNAL CASH is missing or zero")

    return errors


# ---------------------------------------------------------------------------
# Audit: verify data accuracy and write confirmation to each tab
# ---------------------------------------------------------------------------
AUDIT_PASS_FONT = Font(name='Arial', size=10, bold=True, color='006400')
AUDIT_FAIL_FONT = Font(name='Arial', size=10, bold=True, color='8B0000')
AUDIT_DETAIL_FONT = Font(name='Arial', size=9, color='666666')


def _write_audit_result(ws, row, passed, checks_passed, checks_total, details=None):
    """Write audit result row at the bottom of a tab."""
    row += 1  # blank spacer
    ws.cell(row=row, column=1, value="DATA AUDIT").font = SECTION_FONT
    row += 1
    if passed:
        ws.cell(row=row, column=1,
                value=f"PASSED ({checks_passed}/{checks_total} checks)").font = AUDIT_PASS_FONT
    else:
        ws.cell(row=row, column=1,
                value=f"FAILED ({checks_passed}/{checks_total} checks passed)").font = AUDIT_FAIL_FONT
    row += 1
    if details:
        for d in details:
            ws.cell(row=row, column=1, value=d).font = AUDIT_DETAIL_FONT
            row += 1
    return row


def _audit_account_tab(ws, acct, model):
    """Audit a single account tab. Returns (checks_passed, checks_total, details)."""
    checks = []
    monthly = acct.get("monthly", {})
    gains = acct.get("gains", {})
    returns = acct.get("returns", {})
    holdings = acct.get("holdings", [])
    is_margin = acct.get("is_margin", False)

    # 1. Holdings total MV = sum of individual holdings
    if holdings:
        sum_mv = sum(h.get("mv", 0) or 0 for h in holdings)
        total_mv = gains.get("total_mv", 0) or 0
        if total_mv > 0 and abs(sum_mv - total_mv) < 1.0:
            checks.append((True, f"Holdings MV total: ${sum_mv:,.2f} matches"))
        elif total_mv > 0:
            checks.append((False, f"Holdings MV total: sum=${sum_mv:,.2f} vs reported=${total_mv:,.2f}"))
        else:
            checks.append((True, "Holdings: no holdings data (OK for this account type)"))

    # 2. Holdings total CB
    if holdings:
        sum_cb = sum(h.get("cb", 0) or 0 for h in holdings if h.get("cb") is not None)
        total_cb = gains.get("total_cb", 0) or 0
        if total_cb > 0 and abs(sum_cb - total_cb) < 1.0:
            checks.append((True, f"Holdings CB total: ${sum_cb:,.2f} matches"))
        elif total_cb > 0:
            checks.append((False, f"Holdings CB total: sum=${sum_cb:,.2f} vs reported=${total_cb:,.2f}"))

    # 3. Monthly continuity: month N+1 begin = month N end
    month_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    populated = [m for m in month_labels if m in monthly]
    continuity_ok = True
    for i in range(len(populated) - 1):
        m1 = monthly[populated[i]]
        m2 = monthly[populated[i + 1]]
        end1 = m1.get("end", 0) or 0
        begin2 = m2.get("begin", 0) or 0
        if abs(end1 - begin2) > 0.02:
            continuity_ok = False
            checks.append((False, f"Continuity break: {populated[i]} end=${end1:,.2f} != {populated[i+1]} begin=${begin2:,.2f}"))
    if continuity_ok and len(populated) > 1:
        checks.append((True, f"Monthly continuity: {len(populated)} months OK"))
    elif len(populated) <= 1:
        checks.append((True, "Monthly continuity: insufficient data (OK)"))

    # 4. Monthly accounting identity: End = Begin + Add - Sub + Change
    #    (change = total change in value, which includes dividends)
    identity_ok = True
    for m in populated:
        d = monthly[m]
        begin = d.get("begin", 0) or 0
        add = d.get("add", 0) or 0
        sub = d.get("sub", 0) or 0
        change = d.get("change", 0) or 0
        end = d.get("end", 0) or 0
        expected = begin + add - sub + change
        if abs(expected - end) > 1.0 and end > 0:
            identity_ok = False
            checks.append((False, f"Accounting {m}: expected=${expected:,.2f} vs end=${end:,.2f} (diff=${expected-end:,.2f})"))
    if identity_ok and populated:
        checks.append((True, f"Accounting identity: {len(populated)} months OK"))

    # 5. Gains total = dividends + unrealized + realized
    div = gains.get("dividends", 0) or 0
    unr = gains.get("unrealized", 0) or 0
    rea = gains.get("realized", 0) or 0
    total = gains.get("total", 0) or 0
    if total != 0:
        expected_total = div + unr + rea
        if abs(expected_total - total) < 1.0:
            checks.append((True, f"Gains total: ${total:,.2f} = div+unr+rea"))
        else:
            checks.append((False, f"Gains total: expected=${expected_total:,.2f} vs reported=${total:,.2f}"))

    # 6. CB Return = unrealized / total_cb
    cb_ret = returns.get("cb_return", 0)
    total_cb = gains.get("total_cb", 0) or 0
    if total_cb > 0 and cb_ret is not None:
        expected_cbr = unr / total_cb
        if abs(expected_cbr - cb_ret) < 0.001:
            checks.append((True, f"CB Return: {cb_ret:.4f} matches unrealized/CB"))
        else:
            checks.append((False, f"CB Return: computed={expected_cbr:.4f} vs reported={cb_ret:.4f}"))

    passed = sum(1 for ok, _ in checks if ok)
    total = len(checks)
    details = [msg for ok, msg in checks]
    return passed, total, details


def _audit_dashboard(ws, model):
    """Audit Dashboard cross-tab consistency. Returns (checks_passed, checks_total, details)."""
    checks = []

    # Find key rows by label
    row_map = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v:
            row_map[str(v).strip()] = r

    # 1. Benchmark returns — only check if model has benchmark data
    if model.get("benchmarks"):
        for bench in ("S&P 500", "Dow Jones", "NASDAQ"):
            r = row_map.get(bench)
            if r:
                val = ws.cell(r, 2).value
                if isinstance(val, (int, float)) and val != 0:
                    checks.append((True, f"Benchmark {bench}: {val:.4%}"))
                else:
                    checks.append((False, f"Benchmark {bench}: missing or zero"))
            else:
                checks.append((False, f"Benchmark {bench}: row not found"))

    # 2. Account TWR matches model
    for key in model["liquid_accounts"] + model["illiquid_accounts"]:
        acct = model["accounts"][key]
        name = acct["name"]
        r = row_map.get(name)
        if not r:
            continue
        model_twr = acct["returns"].get("twr")
        dash_twr = ws.cell(r, 5).value  # col E = TWR
        if model_twr is not None and isinstance(dash_twr, (int, float)):
            if abs(model_twr - dash_twr) < 0.0001:
                checks.append((True, f"{name} TWR: {dash_twr:.4%} matches model"))
            else:
                checks.append((False, f"{name} TWR: dashboard={dash_twr:.4%} vs model={model_twr:.4%}"))
        elif model_twr is None and (dash_twr is None or dash_twr == "N/A" or dash_twr == ""):
            checks.append((True, f"{name} TWR: N/A (consistent)"))

    # 3. Liquid subtotal has non-zero values
    r = row_map.get("LIQUID SUBTOTAL")
    if r:
        dash_begin = ws.cell(r, 2).value
        dash_end = ws.cell(r, 3).value
        if isinstance(dash_end, (int, float)) and dash_end > 0:
            checks.append((True, f"Liquid subtotal: begin=${dash_begin:,.2f}, end=${dash_end:,.2f}"))
        else:
            checks.append((False, f"Liquid subtotal: ending is zero or missing"))

    # 4. Liquid TWR
    if r and model.get("liquid_twr") is not None:
        dash_liq_twr = ws.cell(r, 5).value
        model_liq_twr = model["liquid_twr"]
        if isinstance(dash_liq_twr, (int, float)):
            if abs(dash_liq_twr - model_liq_twr) < 0.0001:
                checks.append((True, f"Liquid TWR: {dash_liq_twr:.4%} matches model"))
            else:
                checks.append((False, f"Liquid TWR: dashboard={dash_liq_twr:.4%} vs model={model_liq_twr:.4%}"))

    # 5. External cash — only check if model has cash data
    if model["cash"].get("external"):
        r = row_map.get("EXTERNAL CASH")
        if r:
            val = ws.cell(r, 3).value
            if isinstance(val, (int, float)) and val > 0:
                checks.append((True, f"External cash: ${val:,.2f}"))
            else:
                checks.append((False, f"External cash: missing or zero"))

    # 6. Investment gain components sum to total
    div_r = row_map.get("Dividends/Income")
    unr_r = row_map.get("Unrealized Gain/Loss")
    rea_r = row_map.get("Realized Gain/Loss (2026)")
    k401_r = row_map.get("401(k) Investment Gain")
    total_r = row_map.get("Total YTD Investment Gain")
    if all(r is not None for r in [div_r, unr_r, rea_r, total_r]):
        div_v = ws.cell(div_r, 2).value or 0
        unr_v = ws.cell(unr_r, 2).value or 0
        rea_v = ws.cell(rea_r, 2).value or 0
        k401_v = ws.cell(k401_r, 2).value or 0 if k401_r else 0
        total_v = ws.cell(total_r, 2).value or 0
        if all(isinstance(v, (int, float)) for v in [div_v, unr_v, rea_v, total_v]):
            expected = div_v + unr_v + rea_v + k401_v
            if abs(expected - total_v) < 1.0:
                checks.append((True, f"Investment gain total: ${total_v:,.2f} = components"))
            else:
                checks.append((False, f"Investment gain total: expected=${expected:,.2f} vs reported=${total_v:,.2f}"))

    passed = sum(1 for ok, _ in checks if ok)
    total_checks = len(checks)
    details = [msg for ok, msg in checks]
    return passed, total_checks, details


def _audit_cash_tab(ws, model):
    """Audit Cash tab. Returns (checks_passed, checks_total, details)."""
    checks = []
    ext = model["cash"].get("external", {})

    # 1. External cash total matches model
    model_total = sum(
        (v.get("total", 0) if isinstance(v, dict) else v)
        for v in ext.values()
    )
    # Find TOTAL EXTERNAL CASH row
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and "TOTAL EXTERNAL CASH" in str(v):
            cell_val = ws.cell(r, 3).value
            if isinstance(cell_val, str) and cell_val.startswith("="):
                checks.append((True, f"External cash total: formula present (will evaluate in Excel)"))
            elif isinstance(cell_val, (int, float)):
                if abs(cell_val - model_total) < 1.0:
                    checks.append((True, f"External cash total: ${cell_val:,.2f} matches model"))
                else:
                    checks.append((False, f"External cash total: cell=${cell_val:,.2f} vs model=${model_total:,.2f}"))
            break

    # 2. External accounts present
    if ext:
        ext_count = sum(len(v.get("accounts", [])) if isinstance(v, dict) else 1 for v in ext.values())
        if ext_count > 0:
            checks.append((True, f"External accounts: {ext_count} from Plaid"))

    passed = sum(1 for ok, _ in checks if ok)
    total = len(checks)
    details = [msg for ok, msg in checks]
    return passed, total, details


def _audit_workbook(wb, model):
    """Run audit on all tabs. Returns list of failure strings.

    If all checks pass, writes confirmation rows to each tab and returns [].
    If any check fails, returns the failure list — caller should abort the build.
    """
    tab_results = []  # [(tab_name, passed, total, details)]

    for key in model["liquid_accounts"] + model["illiquid_accounts"]:
        acct = model["accounts"][key]
        tab_name = acct["tab_name"]
        if tab_name not in wb.sheetnames:
            tab_results.append((tab_name, 0, 1, ["Tab missing from workbook"]))
            continue

        if "investments" in acct:
            gains = acct.get("gains", {})
            invested = gains.get("total_invested", 0) or gains.get("total_cb", 0) or 0
            current = gains.get("total_current", 0) or gains.get("total_mv", 0) or 0
            passed = (1 if invested > 0 else 0) + (1 if current > 0 else 0)
            details = [f"Total invested: ${invested:,.2f}", f"Total current value: ${current:,.2f}"]
            tab_results.append((tab_name, passed, 2, details))
        else:
            passed, total, details = _audit_account_tab(wb[tab_name], acct, model)
            tab_results.append((tab_name, passed, total, details))

    if "Dashboard" in wb.sheetnames:
        passed, total, details = _audit_dashboard(wb["Dashboard"], model)
        tab_results.append(("Dashboard", passed, total, details))

    if "Cash" in wb.sheetnames:
        passed, total, details = _audit_cash_tab(wb["Cash"], model)
        tab_results.append(("Cash", passed, total, details))

    # Collect failures
    failures = []
    for tab_name, passed, total, details in tab_results:
        if passed < total:
            failures.append(f"{tab_name}: {passed}/{total} checks passed")

    # If all passed, write confirmation to each tab
    if not failures:
        for tab_name, passed, total, details in tab_results:
            if tab_name in wb.sheetnames:
                _write_audit_result(wb[tab_name], wb[tab_name].max_row + 1, True, passed, total, details)

    return failures
