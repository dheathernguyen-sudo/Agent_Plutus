# tests/advisor/test_run_daily.py
"""Integration test for advisor.run_daily — wires together observations,
state, narrator (stubbed), and writer."""
import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook


def _build_workbook(path: Path):
    wb = Workbook()
    wb.active.title = "Dashboard"
    wb.save(path)


def _model_with_one_finding():
    return {
        "as_of": "2026-04-25",
        "year": 2026,
        "accounts": {
            "fb": {
                "name": "FB", "tab_name": "FB", "type": "liquid",
                "holdings": [{"ticker": "AAPL", "mv": 5000, "cb": 4000,
                              "qty": 10, "price": 500}],
                "cash_position": 0, "margin_debt": 0,
                "gains": {"total_mv": 5000, "total_cb": 4000,
                          "dividends": 0, "unrealized": 1000, "realized": 0},
                "returns": {},
            }
        },
        "liquid_accounts": ["fb"],
        "illiquid_accounts": [],
        "benchmarks": {},
        "cash": {"external": {}, "embedded": {}},
        "sectors": [{"name": "Tech", "pct": 1.0, "value": 5000,
                      "by_account": {"FB": 5000}}],  # 100% Tech triggers urgent
    }


class StubClient:
    @property
    def messages(self): return self
    def create(self, **kwargs):
        class B: text = json.dumps({
            "headline": "One concentration concern.",
            "new": [{"category": "sector_concentration",
                     "narrative": "Tech is 100% — diversify."}],
            "standing": [],
        })
        class R: content = [B()]
        return R()


def test_run_daily_writes_recommendations_tab(tmp_path, monkeypatch):
    from advisor import run_daily
    from advisor.profile import Profile

    wb_path = tmp_path / "wb.xlsx"
    _build_workbook(wb_path)
    state_dir = tmp_path / "state"

    run_daily(
        _model_with_one_finding(),
        wb_path,
        profile=Profile(),
        state_dir=state_dir,
        client=StubClient(),
        today=date(2026, 4, 25),
    )

    wb = load_workbook(wb_path)
    assert "Recommendations" in wb.sheetnames
    text = " ".join(str(wb["Recommendations"].cell(r, 1).value or "")
                    for r in range(1, 30))
    assert "One concentration concern" in text or "Tech" in text


def test_run_daily_persists_state(tmp_path):
    from advisor import run_daily
    from advisor.profile import Profile

    wb_path = tmp_path / "wb.xlsx"
    _build_workbook(wb_path)
    state_dir = tmp_path / "state"

    run_daily(
        _model_with_one_finding(),
        wb_path,
        profile=Profile(),
        state_dir=state_dir,
        client=StubClient(),
        today=date(2026, 4, 25),
    )
    findings_file = state_dir / "findings_2026-04-25.json"
    assert findings_file.exists()
    payload = json.loads(findings_file.read_text())
    cats = [f["category"] for f in payload["findings"]]
    assert "sector_concentration" in cats


def test_run_daily_dedup_across_days(tmp_path):
    from advisor import run_daily
    from advisor.profile import Profile

    wb_path = tmp_path / "wb.xlsx"
    _build_workbook(wb_path)
    state_dir = tmp_path / "state"

    # Day 1
    run_daily(_model_with_one_finding(), wb_path, profile=Profile(),
              state_dir=state_dir, client=StubClient(),
              today=date(2026, 4, 24))
    # Day 2 — same finding, expected to be classified "standing"
    captured = {}

    class CapturingClient(StubClient):
        def create(self, **kwargs):
            captured["payload"] = kwargs["messages"][0]["content"]
            return super().create(**kwargs)

    run_daily(_model_with_one_finding(), wb_path, profile=Profile(),
              state_dir=state_dir, client=CapturingClient(),
              today=date(2026, 4, 25))
    payload = json.loads(captured["payload"])
    standing_cats = [f["category"] for f in payload["findings"]["standing"]]
    assert "sector_concentration" in standing_cats
