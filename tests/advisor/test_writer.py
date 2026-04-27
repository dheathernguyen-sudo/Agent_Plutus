"""Tests for advisor.writer — write Recommendations tab to workbook."""
from openpyxl import Workbook, load_workbook


def test_write_creates_recommendations_tab(tmp_path):
    from advisor.writer import write_recommendations_tab

    wb_path = tmp_path / "wb.xlsx"
    wb = Workbook()
    wb.active.title = "Existing"
    wb.save(wb_path)

    write_recommendations_tab(
        wb_path,
        brief_md="## Headline\n\nHello world.",
        findings=[],
    )

    wb2 = load_workbook(wb_path)
    assert "Recommendations" in wb2.sheetnames
    ws = wb2["Recommendations"]
    # Headline content is written into the Overall State section.
    found_text = " ".join(str(ws.cell(r, c).value or "")
                          for r in range(1, 15) for c in range(1, 5))
    assert "Hello world" in found_text


def test_write_overwrites_existing_recommendations_tab(tmp_path):
    from advisor.writer import write_recommendations_tab

    wb_path = tmp_path / "wb.xlsx"
    wb = Workbook()
    wb.active.title = "Existing"
    rec = wb.create_sheet("Recommendations")
    rec["A1"] = "old content"
    wb.save(wb_path)

    write_recommendations_tab(wb_path, brief_md="## Headline\n\nfresh body", findings=[])

    wb2 = load_workbook(wb_path)
    found_text = " ".join(str(wb2["Recommendations"].cell(r, c).value or "")
                          for r in range(1, 15) for c in range(1, 5))
    assert "old content" not in found_text
    assert "fresh body" in found_text


def test_writer_failure_does_not_propagate(tmp_path, caplog):
    from advisor.writer import write_recommendations_tab
    import logging

    nonexistent = tmp_path / "missing.xlsx"
    with caplog.at_level(logging.WARNING):
        # Should not raise
        write_recommendations_tab(nonexistent, brief_md="x", findings=[])
    assert "writer" in caplog.text.lower() or "missing" in caplog.text.lower()
