"""Tests for validate_workbook.py — pure logic checks."""

import sys
from pathlib import Path

# validate_workbook.py lives one level above the repo root
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import pytest
from openpyxl import Workbook

from validate_workbook import (
    _is_formula,
    _is_formula_error,
    _to_float,
    check_formula_errors,
    check_cross_sheet_refs,
)


class TestIsFormula:
    def test_excel_formula(self):
        assert _is_formula("=SUM(A1:A10)") is True

    def test_plain_string(self):
        assert _is_formula("hello") is False

    def test_number(self):
        assert _is_formula(42) is False

    def test_none(self):
        assert _is_formula(None) is False

    def test_empty_string(self):
        assert _is_formula("") is False


class TestIsFormulaError:
    def test_ref_error(self):
        assert _is_formula_error("#REF!") is True

    def test_div_zero(self):
        assert _is_formula_error("#DIV/0!") is True

    def test_value_error(self):
        assert _is_formula_error("#VALUE!") is True

    def test_name_error(self):
        assert _is_formula_error("#NAME?") is True

    def test_na_error(self):
        assert _is_formula_error("#N/A") is True

    def test_normal_string(self):
        assert _is_formula_error("hello") is False

    def test_number(self):
        assert _is_formula_error(42) is False

    def test_none(self):
        assert _is_formula_error(None) is False


class TestToFloat:
    def test_integer(self):
        assert _to_float(42) == 42.0

    def test_float(self):
        assert _to_float(3.14) == 3.14

    def test_none(self):
        assert _to_float(None) is None

    def test_string(self):
        assert _to_float("hello") is None

    def test_formula_string(self):
        assert _to_float("=SUM(A1:A10)") is None


class TestCheckFormulaErrors:
    def test_clean_workbook_passes(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = 100
        ws["A2"] = "hello"
        ws["A3"] = "=A1*2"
        findings = check_formula_errors(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0

    def test_ref_error_detected(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "#REF!"
        findings = check_formula_errors(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 1
        assert "#REF!" in errors[0].message

    def test_multiple_errors_detected(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "#REF!"
        ws["B2"] = "#DIV/0!"
        ws["C3"] = "#VALUE!"
        findings = check_formula_errors(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 3


class TestCheckCrossSheetRefs:
    def test_valid_cross_ref_passes(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Dashboard"
        ws2 = wb.create_sheet("Details")
        ws2["A1"] = 42
        ws1["A1"] = "='Details'!A1"
        findings = check_cross_sheet_refs(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0

    def test_missing_tab_ref_detected(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Dashboard"
        ws1["A1"] = "='NonExistent'!A1"
        findings = check_cross_sheet_refs(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 1
        assert "NonExistent" in errors[0].message

    def test_empty_row_ref_warns(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Dashboard"
        ws2 = wb.create_sheet("Details")
        ws1["A1"] = "='Details'!A99"
        findings = check_cross_sheet_refs(wb)
        warns = [f for f in findings if f.severity == "WARN"]
        assert len(warns) >= 1
