"""Golden integration test.

Pins a synthetic raw extract + account data and asserts that the full
model → builder pipeline produces exact Dashboard values. This catches
bugs that unit tests miss: path drift, output-location bugs, cross-section
aggregation errors, validator/builder drift, silent field renames.

Why synthetic (not a real extract):
  - Avoids committing real financial data.
  - Small, readable, deterministic.
  - Exercises the same code paths as production (same model, same builder).

If the builder's column/row layout shifts, this test fails loudly with a
clear diff — that's the signal to update both the builder and the expected
values in tandem, rather than silently accumulating drift.
"""
import json
import os
import tempfile
from pathlib import Path

import pytest


# ---------------------------------------------------------------------------
# Synthetic fixture — the "golden" inputs
# ---------------------------------------------------------------------------
# Account JSONs that would normally live in data/*.json
_GOLDEN_ACCOUNTS = {
    "fidelity_brokerage.json": {
        "account": {"name": "Fidelity Brokerage", "type": "liquid",
                    "provider": "snaptrade", "tab_name": "Fidelity Brokerage",
                    "number": "Z23-889908"},
        "holdings": [  # stale — should be overridden by live_extraction
            {"ticker": "STALE", "qty": 1, "price": 1, "mv": 1, "cb": 1},
        ],
        "cash_position": 999.0,  # stale — should be zeroed when live source wins
        "monthly": {
            "Jan": {"begin": 1000, "end": 1100, "add": 0, "sub": 0,
                    "div": 0, "change": 100},
        },
        "sector_map": {
            "AAPL":  {"sector": "Technology", "country": "United States"},
            "FCASH": {"sector": "Cash",       "country": "United States"},
        },
    },
    "fidelity_roth_ira.json": {
        "account": {"name": "Fidelity Roth IRA", "type": "liquid",
                    "provider": "snaptrade", "tab_name": "Fidelity Roth IRA",
                    "number": "266-209863"},
        "holdings": [], "cash_position": 0,
        "monthly": {}, "sector_map": {},
    },
    "fidelity_hsa.json": {
        "account": {"name": "Fidelity HSA", "type": "liquid",
                    "provider": "snaptrade", "tab_name": "Fidelity HSA",
                    "number": "249-509651"},
        "holdings": [], "cash_position": 0,
        "monthly": {}, "sector_map": {},
    },
    "robinhood.json": {
        "account": {"name": "Robinhood", "type": "liquid",
                    "provider": "snaptrade", "tab_name": "Robinhood"},
        "holdings": [
            {"ticker": "NVDA", "qty": 10, "price": 200.0, "mv": 2000.0,
             "cb": 1000.0, "avg_cost": 100.0},
        ],
        "cash_position": 0,
        "monthly": {},
        "sector_map": {"NVDA": {"sector": "Technology", "country": "United States"}},
    },
    "k401.json": {
        "account": {"name": "401(k)", "type": "illiquid",
                    "provider": "merrill", "tab_name": "401(k)"},
        "quarterly": [{
            "period": "Q1", "beginning": 4000, "ee_contributions": 500,
            "er_contributions": 250, "fees": 0, "change_in_value": 250, "ending": 5000,
        }],
        "holdings": [],
        "monthly": {},
        "sector_map": {},
    },
    "angel.json": {
        "account": {"name": "Angel Investments", "type": "illiquid",
                    "provider": "manual", "tab_name": "Angel Investments"},
        "investments": [
            {"company": "Acme", "sector": "SaaS", "year": 2024, "series": "Seed",
             "amount": 10000, "pm_invest": 5_000_000, "pm_latest": 15_000_000,
             "source": "Seed, 2024"},
        ],
        "holdings": [],
    },
    "cash.json": {
        "account": {"name": "Cash", "type": "cash", "tab_name": "Cash"},
    },
}

# Live SnapTrade-format fid_data — redacted keys (the post-April-2026 format)
_GOLDEN_LIVE_EXTRACTION = {
    "fidelity_*****9908": {  # Fidelity Brokerage
        "AAPL":  {"qty": 5,   "price": 200.0, "mv": 1000.0, "cb": 500.0},
        "FCASH": {"qty": 100, "price": 1.0,   "mv":  100.0, "cb": 100.0},
    },
    "fidelity_*****9863": {  # Roth IRA
        "VOO":   {"qty": 3,  "price": 500.0, "mv": 1500.0, "cb": 1200.0},
        "SPAXX": {"qty": 50, "price": 1.0,   "mv":   50.0, "cb":   50.0},
    },
    "fidelity_*****9651": {  # HSA
        "CEG":   {"qty": 2,  "price": 300.0, "mv": 600.0, "cb": 400.0},
        "FDRXX": {"qty": 25, "price": 1.0,   "mv":  25.0, "cb":  25.0},
    },
}

# Plaid Merrill raw for 401(k) live holdings
_GOLDEN_RAW_EXTRACTION = {
    "merrill": {
        "securities": [
            {"security_id": "s1", "name": "Russell 1000 Index", "ticker_symbol": "FXAIX"},
            {"security_id": "s2", "name": "Intl Equity Index",   "ticker_symbol": "FSPSX"},
        ],
        "holdings": [
            {"security_id": "s1", "institution_value": 3000.0,
             "cost_basis": 2500.0, "quantity": 30.0},
            {"security_id": "s2", "institution_value": 2000.0,
             "cost_basis": 1800.0, "quantity": 100.0},
        ],
    },
}

_GOLDEN_BENCHMARKS = {"S&P 500": 0.05, "Dow Jones": 0.03, "NASDAQ": 0.07}

_GOLDEN_CASH_DATA = {
    "chase":  {"accounts": [{"name": "Checking", "balance": 500.0}],  "total": 500.0},
    "marcus": {"accounts": [{"name": "Savings",  "balance": 1000.0}], "total": 1000.0},
}

# Expected output — what the Dashboard should contain after build
# Derived by hand from the synthetic inputs above:
#   Fidelity Brokerage  MV = AAPL 1000 + FCASH 100              = 1100
#   Roth IRA            MV = VOO 1500 + SPAXX 50                = 1550
#   HSA                 MV = CEG 600 + FDRXX 25                 = 625
#   Robinhood           MV = NVDA 2000                          = 2000
#   LIQUID SUBTOTAL                                             = 5275
#   401(k)              MV = FXAIX 3000 + FSPSX 2000            = 5000
#   Angel               MV = 10000 * (15M / 5M)                 = 30000
#   ILLIQUID SUBTOTAL                                           = 35000
#   EXTERNAL CASH          = chase 500 + marcus 1000            = 1500
#   TOTAL PORTFOLIO        = 5275 + 35000 + 1500                = 41775
_EXPECTED_DASHBOARD = {
    ("Fidelity Brokerage",  "Market Value"):  1100.0,
    ("Roth IRA",            "Market Value"):  1550.0,
    ("HSA",                 "Market Value"):   625.0,
    ("Robinhood",           "Market Value"):  2000.0,
    ("LIQUID SUBTOTAL",     "Market Value"):  5275.0,
    ("Angel Investments",   "Market Value"): 30000.0,
    ("401(k)",              "Market Value"):  5000.0,
    ("ILLIQUID SUBTOTAL",   "Market Value"): 35000.0,
    ("EXTERNAL CASH",       "Market Value"):  1500.0,
    ("TOTAL PORTFOLIO",     "Market Value"): 41775.0,
}


# ---------------------------------------------------------------------------
# The golden test
# ---------------------------------------------------------------------------
@pytest.fixture
def golden_workbook_path(tmp_path):
    """Build a workbook from the synthetic fixtures and return the path."""
    import sys
    # Ensure flat-layout imports work (conftest already does this, but be explicit)
    root = Path(__file__).parent.parent
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))

    # Write synthetic data/*.json into the tmp dir
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    for fname, content in _GOLDEN_ACCOUNTS.items():
        (data_dir / fname).write_text(json.dumps(content))

    from portfolio_model import build_model
    from build_workbook import build

    model = build_model(
        data_dir=str(data_dir),
        live_extraction=_GOLDEN_LIVE_EXTRACTION,
        raw_extraction=_GOLDEN_RAW_EXTRACTION,
        benchmarks=_GOLDEN_BENCHMARKS,
        cash_data=_GOLDEN_CASH_DATA,
    )
    out_path = tmp_path / "golden.xlsx"
    build(model, str(out_path))
    return out_path


def _find_row(ws, label, col=1):
    """Return the row number whose column `col` contains `label` (exact match)."""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is not None and str(v).strip() == label:
            return r
    return None


def test_golden_workbook_dashboard_values(golden_workbook_path):
    """The Dashboard's Market Value column must match the expected numbers
    derived by hand from the synthetic fixtures. Any silent drift in the
    model, builder, or aggregation logic fails this test."""
    from openpyxl import load_workbook

    wb = load_workbook(golden_workbook_path, data_only=False)
    assert "Dashboard" in wb.sheetnames
    ws = wb["Dashboard"]

    # Find the Account Overview header row (row with 'Account' in col A)
    header_row = _find_row(ws, "Account")
    assert header_row is not None, "Could not find Account Overview header"

    # Locate the Market Value column by scanning the header row
    mv_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(header_row, c).value == "Market Value":
            mv_col = c
            break
    assert mv_col is not None, "Dashboard has no 'Market Value' column"

    # Check each expected row
    mismatches = []
    for (row_label, col_label), expected in _EXPECTED_DASHBOARD.items():
        r = _find_row(ws, row_label)
        if r is None:
            mismatches.append(f"row '{row_label}' not found in Dashboard")
            continue
        actual = ws.cell(r, mv_col).value
        if actual is None or abs(float(actual) - expected) > 0.01:
            mismatches.append(
                f"{row_label!r} Market Value: expected {expected:,.2f}, got {actual!r}"
            )

    assert not mismatches, "Golden workbook diverged:\n" + "\n".join(f"  - {m}" for m in mismatches)


def test_golden_workbook_has_expected_tabs(golden_workbook_path):
    """Workbook must contain exactly the expected tabs — no missing or extra."""
    from openpyxl import load_workbook

    wb = load_workbook(golden_workbook_path)
    expected = {"Dashboard", "Fidelity Brokerage", "Fidelity Roth IRA",
                "Fidelity HSA", "Robinhood", "401(k)", "Angel Investments", "Cash"}
    assert set(wb.sheetnames) == expected, (
        f"Tab mismatch.\n"
        f"  Missing: {expected - set(wb.sheetnames)}\n"
        f"  Extra:   {set(wb.sheetnames) - expected}"
    )


# Note: validator is exercised against the REAL workbook in
# test_regressions.py::test_current_workbook_passes_validator_with_zero_errors.
# The minimal synthetic fixture here omits monthly/sold sections that the
# registry expects, so validating it would produce noise that isn't drift.
