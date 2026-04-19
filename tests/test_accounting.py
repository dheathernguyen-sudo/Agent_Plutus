"""Tests for accounting checks in validate_workbook.py."""

import pytest
from openpyxl import Workbook

from registry import MONTHLY_COLUMNS, REGISTRY
from validate_workbook import check_balance_continuity, check_accounting_identity


def _build_monthly_workbook(tab_name, rows):
    """Build a minimal workbook with monthly data rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = tab_name

    reg = REGISTRY[tab_name]
    jan_row = reg["monthly_jan"][1]

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for i, month in enumerate(months):
        ws[f"A{jan_row + i}"] = month

    col_map = MONTHLY_COLUMNS[tab_name]
    field_to_col = {v: k for k, v in col_map.items()}

    for i, row_data in enumerate(rows):
        r = jan_row + i
        if "beginning" in row_data:
            ws[f"B{r}"] = row_data["beginning"]
        if "deposits" in row_data:
            ws[f"{field_to_col['deposits_additions_contributions']}{r}"] = row_data["deposits"]
        if "withdrawals" in row_data:
            ws[f"{field_to_col['withdrawals_subtractions_distributions']}{r}"] = row_data["withdrawals"]
        if "dividends" in row_data:
            ws[f"{field_to_col['dividends']}{r}"] = row_data["dividends"]
        if "market_change" in row_data:
            ws[f"{field_to_col['market_change']}{r}"] = row_data["market_change"]
        if "ending" in row_data:
            ws[f"{field_to_col['ending']}{r}"] = row_data["ending"]

    return wb


class TestBalanceContinuity:
    def test_continuous_balances_pass(self):
        rows = [
            {"beginning": 10000, "ending": 10500},
            {"beginning": 10500, "ending": 11000},
            {"beginning": 11000, "ending": 10800},
        ]
        wb = _build_monthly_workbook("Fidelity Brokerage", rows)
        findings = check_balance_continuity(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0

    def test_discontinuous_balances_fail(self):
        rows = [
            {"beginning": 10000, "ending": 10500},
            {"beginning": 10600, "ending": 11000},
        ]
        wb = _build_monthly_workbook("Fidelity Brokerage", rows)
        findings = check_balance_continuity(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 1

    def test_penny_rounding_passes(self):
        rows = [
            {"beginning": 10000, "ending": 10500.005},
            {"beginning": 10500.01, "ending": 11000},
        ]
        wb = _build_monthly_workbook("Fidelity Brokerage", rows)
        findings = check_balance_continuity(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0

    def test_single_month_no_error(self):
        rows = [{"beginning": 10000, "ending": 10500}]
        wb = _build_monthly_workbook("Fidelity Brokerage", rows)
        findings = check_balance_continuity(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0


class TestAccountingIdentity:
    def test_correct_identity_passes(self):
        rows = [{"beginning": 10000, "deposits": 500, "withdrawals": 200, "dividends": 50, "market_change": 300, "ending": 10650}]
        wb = _build_monthly_workbook("Fidelity Brokerage", rows)
        findings = check_accounting_identity(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0

    def test_incorrect_identity_fails(self):
        rows = [{"beginning": 10000, "deposits": 500, "withdrawals": 200, "dividends": 50, "market_change": 300, "ending": 12000}]
        wb = _build_monthly_workbook("Fidelity Brokerage", rows)
        findings = check_accounting_identity(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 1

    def test_negative_market_change_valid(self):
        rows = [{"beginning": 10000, "deposits": 0, "withdrawals": 0, "dividends": 0, "market_change": -500, "ending": 9500}]
        wb = _build_monthly_workbook("Fidelity Brokerage", rows)
        findings = check_accounting_identity(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0

    def test_all_zeros_passes(self):
        rows = [{"beginning": 0, "deposits": 0, "withdrawals": 0, "dividends": 0, "market_change": 0, "ending": 0}]
        wb = _build_monthly_workbook("Fidelity Brokerage", rows)
        findings = check_accounting_identity(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0

    def test_within_tolerance_passes(self):
        rows = [{"beginning": 10000, "deposits": 500, "withdrawals": 200, "dividends": 50, "market_change": 300, "ending": 10650.75}]
        wb = _build_monthly_workbook("Fidelity Brokerage", rows)
        findings = check_accounting_identity(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0

    def test_missing_field_skips_gracefully(self):
        rows = [{"beginning": 10000, "deposits": 500, "dividends": 50, "market_change": 300, "ending": 10850}]
        wb = _build_monthly_workbook("Fidelity Brokerage", rows)
        findings = check_accounting_identity(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0
