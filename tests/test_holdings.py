"""Tests for holdings total (Check 6) and YTD gain consistency (Check 7)."""

import sys
from pathlib import Path

# validate_workbook.py and registry.py live one level above the repo root
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import pytest
from openpyxl import Workbook

from registry import REGISTRY, HOLDINGS_ROWS
from validate_workbook import check_holdings_totals, check_ytd_gain


def _build_holdings_workbook(tab_name, holdings_values, total_values):
    wb = Workbook()
    ws = wb.active
    ws.title = tab_name
    config = HOLDINGS_ROWS[tab_name]
    first = config["first"]
    total_row = config["total"]
    mv_col = config["mv_col"]
    cb_col = config["cb_col"]
    gl_col = config["gl_col"]

    for i, (mv, cb, gl) in enumerate(holdings_values):
        r = first + i
        ws[f"{mv_col}{r}"] = mv
        ws[f"{cb_col}{r}"] = cb
        ws[f"{gl_col}{r}"] = gl

    ws[f"{mv_col}{total_row}"] = total_values[0]
    ws[f"{cb_col}{total_row}"] = total_values[1]
    ws[f"{gl_col}{total_row}"] = total_values[2]
    ws[f"A{total_row}"] = "TOTAL"
    return wb


class TestHoldingsTotals:
    def test_correct_totals_pass(self):
        holdings = [(1500.00, 1200.00, 300.00), (2000.00, 1800.00, 200.00)]
        totals = (3500.00, 3000.00, 500.00)
        wb = _build_holdings_workbook("Fidelity Brokerage", holdings, totals)
        findings = check_holdings_totals(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0

    def test_wrong_mv_total_fails(self):
        holdings = [(1500.00, 1200.00, 300.00), (2000.00, 1800.00, 200.00)]
        totals = (9999.00, 3000.00, 500.00)
        wb = _build_holdings_workbook("Fidelity Brokerage", holdings, totals)
        findings = check_holdings_totals(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) >= 1

    def test_single_holding_passes(self):
        holdings = [(5000.00, 4000.00, 1000.00)]
        totals = (5000.00, 4000.00, 1000.00)
        wb = _build_holdings_workbook("Fidelity Roth IRA", holdings, totals)
        findings = check_holdings_totals(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0

    def test_formula_total_warns_not_errors(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Fidelity Brokerage"
        config = HOLDINGS_ROWS["Fidelity Brokerage"]
        r = config["first"]
        ws[f"{config['mv_col']}{r}"] = 1500.00
        ws[f"{config['cb_col']}{r}"] = 1200.00
        ws[f"{config['gl_col']}{r}"] = 300.00
        tr = config["total"]
        ws[f"{config['mv_col']}{tr}"] = "=SUM(D13:D31)"
        ws[f"{config['cb_col']}{tr}"] = "=SUM(E13:E31)"
        ws[f"{config['gl_col']}{tr}"] = "=SUM(F13:F31)"
        findings = check_holdings_totals(wb)
        warns = [f for f in findings if f.severity == "WARN"]
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0
        assert len(warns) >= 1


def _build_ytd_workbook(tab_name, total_ytd, unrealized, realized, dividends):
    wb = Workbook()
    ws = wb.active
    ws.title = tab_name
    reg = REGISTRY[tab_name]

    def set_val(key, value):
        if key in reg:
            col, row, label = reg[key]
            ws[f"{col}{row}"] = value
            ws[f"A{row}"] = label

    set_val("total_ytd", total_ytd)
    set_val("unrealized", unrealized)
    set_val("realized", realized)
    set_val("dividends", dividends)
    return wb


class TestYTDGain:
    def test_consistent_ytd_passes(self):
        wb = _build_ytd_workbook("Fidelity Brokerage", 1250.00, 1000.00, 200.00, 50.00)
        findings = check_ytd_gain(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0

    def test_inconsistent_ytd_fails(self):
        wb = _build_ytd_workbook("Fidelity Brokerage", 2000.00, 1000.00, 200.00, 50.00)
        findings = check_ytd_gain(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 1

    def test_negative_values_valid(self):
        wb = _build_ytd_workbook("Fidelity Brokerage", -250.00, -500.00, 200.00, 50.00)
        findings = check_ytd_gain(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0

    def test_all_zeros_passes(self):
        wb = _build_ytd_workbook("Fidelity Brokerage", 0, 0, 0, 0)
        findings = check_ytd_gain(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0

    def test_formula_values_warns(self):
        wb = _build_ytd_workbook("Fidelity Brokerage", "=B7+B8+B6", "=SUM(F13:F31)", 200.00, 50.00)
        findings = check_ytd_gain(wb)
        warns = [f for f in findings if f.severity == "WARN"]
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0
        assert len(warns) >= 1

    def test_within_tolerance_passes(self):
        wb = _build_ytd_workbook("Fidelity Brokerage", 1250.80, 1000.00, 200.00, 50.00)
        findings = check_ytd_gain(wb)
        errors = [f for f in findings if f.severity == "ERROR"]
        assert len(errors) == 0
