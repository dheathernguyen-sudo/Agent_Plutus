"""Phase 1 test: compare PDF-driven vs JSON-driven rebuild, using the CURRENT
rebuild_rh_tab.py code in both runs. This isolates the question "does the JSON
path reproduce the PDF path?" from drift between the live workbook (built with
older code) and the code on disk today.
"""
import json
import os
import shutil
import sys
import importlib
from pathlib import Path

PROJECT = Path(__file__).resolve().parent.parent
TEST_DIR = PROJECT / "test_api_rebuild"
LIVE_WB = PROJECT / "2026_Portfolio_Analysis.xlsx"
WB_PDF = TEST_DIR / "wb_pdf.xlsx"
WB_JSON = TEST_DIR / "wb_json.xlsx"

shutil.copy(LIVE_WB, WB_PDF)
shutil.copy(LIVE_WB, WB_JSON)
print(f"Seeded: {WB_PDF.name}, {WB_JSON.name}")

# Build JSON-derived stmts dict (same shape the PDF parser emits)
data = json.loads((PROJECT / "rh_monthly_returns.json").read_text())
stmts_from_json = {
    r["month"]: {
        "start": r.get("start_date", ""),
        "end": r.get("end_date", ""),
        "opening": r["opening"],
        "closing": r["closing"],
        "deposits": r["deposits"],
        "withdrawals": r["withdrawals"],
        "dividends": r["dividends"],
    }
    for r in data["monthly_returns"]
}
print(f"JSON 2026 keys: {sorted(k for k in stmts_from_json if '/2026' in k)}")

os.chdir(PROJECT)
sys.path.insert(0, str(PROJECT))

# Neutralize the side effects that would touch live state or pollute stdout
import registry
registry.update_registry = lambda *a, **kw: None
import validate_workbook
validate_workbook.validate_structural = lambda *a, **kw: []
validate_workbook.format_findings = lambda *a, **kw: ""


def run_rebuild(target_wb, patched_loader=None):
    """Import rebuild_rh_tab fresh, point it at target_wb, optionally patch loader, run main()."""
    if "rebuild_rh_tab" in sys.modules:
        del sys.modules["rebuild_rh_tab"]
    import rebuild_rh_tab
    rebuild_rh_tab.XLSX = target_wb
    if patched_loader is not None:
        rebuild_rh_tab.load_all_statements = patched_loader
    rebuild_rh_tab.main()
    return rebuild_rh_tab


print("\n--- Run 1: PDF-driven (unchanged load_all_statements) ---")
mod_pdf = run_rebuild(WB_PDF)
stmts_from_pdf = mod_pdf.load_all_statements()
pdf_keys_2026 = sorted(k for k in stmts_from_pdf if "/2026" in k)
print(f"PDF 2026 keys: {pdf_keys_2026}")

print("\n--- Run 2: JSON-driven (patched load_all_statements) ---")
run_rebuild(WB_JSON, patched_loader=lambda: stmts_from_json)

print("\n--- Comparing 'Robinhood' tab: wb_pdf vs wb_json ---")
import openpyxl
a = openpyxl.load_workbook(WB_PDF, data_only=False)["Robinhood"]
b = openpyxl.load_workbook(WB_JSON, data_only=False)["Robinhood"]

mismatches = []
max_row = max(a.max_row, b.max_row)
max_col = max(a.max_column, b.max_column)
for r in range(1, max_row + 1):
    for c in range(1, max_col + 1):
        va = a.cell(row=r, column=c).value
        vb = b.cell(row=r, column=c).value
        if va != vb:
            mismatches.append((r, c, va, vb))

print(f"Cells compared: {max_row * max_col}")
print(f"Mismatches: {len(mismatches)}")
if mismatches:
    print("\nFirst 30 mismatches (cell, pdf, json):")
    for r, c, va, vb in mismatches[:30]:
        col = openpyxl.utils.get_column_letter(c)
        print(f"  {col}{r}: pdf={va!r}  json={vb!r}")
else:
    print("\nRESULT: PDF-driven and JSON-driven rebuilds produce IDENTICAL Robinhood tabs.")
    print("Phase 1 PASSED — the rebuild code works from JSON without needing PDFs.")

# Also dump a side-by-side of the raw stmts dicts (for the 2026 months) so you
# can eyeball whether any per-field differences exist between PDF- and JSON-sourced data
print("\n--- Source data diff (2026 only) ---")
for key in sorted(set(pdf_keys_2026) | set(k for k in stmts_from_json if "/2026" in k)):
    p = stmts_from_pdf.get(key, {})
    j = stmts_from_json.get(key, {})
    print(f"\n  {key}")
    for field in ("opening", "closing", "deposits", "withdrawals", "dividends"):
        pv = p.get(field)
        jv = j.get(field)
        flag = "" if pv == jv else "  <-- DIFF"
        print(f"    {field:12s} pdf={pv}  json={jv}{flag}")
