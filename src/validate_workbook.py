"""
validate_workbook.py — Core validator for 2026_Portfolio_Analysis.xlsx

Runs 7 structural and numerical checks against the workbook using
openpyxl (no live Excel required). Exits with code 1 if any ERRORs found.
"""

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.utils import column_index_from_string

from registry import REGISTRY, MONTHLY_COLUMNS, HOLDINGS_ROWS

# ---------------------------------------------------------------------------
# Finding dataclass
# ---------------------------------------------------------------------------

@dataclass
class Finding:
    severity: str   # "ERROR", "WARN", "PASS"
    tab: str
    check: str
    cell: str
    message: str


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

FORMULA_ERRORS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#N/A", "#NUM!")


def _cell_ref(col: str, row: int) -> str:
    return f"{col}{row}"


def _raw(cell) -> object:
    """Return the cell value as stored (formula string or numeric/string)."""
    return cell.value


def _is_formula(val) -> bool:
    return isinstance(val, str) and val.startswith("=")


def _is_formula_error(val) -> bool:
    if not isinstance(val, str):
        return False
    for err in FORMULA_ERRORS:
        if val.startswith(err):
            return True
    return False


def _to_float(val) -> float | None:
    """Convert a cell value to float, or None if not possible."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    return None


def _row_has_data(ws, row: int) -> bool:
    """Return True if any cell in the row has a non-None value."""
    for cell in ws[row]:
        if cell.value is not None:
            return True
    return False


# ---------------------------------------------------------------------------
# Check 1 — Label matching
# ---------------------------------------------------------------------------

def check_labels(wb) -> List[Finding]:
    findings = []
    for tab_name, keys in REGISTRY.items():
        if tab_name not in wb.sheetnames:
            findings.append(Finding(
                severity="ERROR",
                tab=tab_name,
                check="Check1-Labels",
                cell="—",
                message=f"Tab '{tab_name}' not found in workbook",
            ))
            continue
        ws = wb[tab_name]
        pass_count = 0
        for key_name, (col, row, expected_label) in keys.items():
            a_cell = ws[f"A{row}"]
            actual = str(a_cell.value) if a_cell.value is not None else ""
            # Strip leading apostrophe that openpyxl sometimes returns
            actual_clean = actual.lstrip("'")
            if expected_label.lower() not in actual_clean.lower():
                findings.append(Finding(
                    severity="ERROR",
                    tab=tab_name,
                    check="Check1-Labels",
                    cell=f"A{row}",
                    message=(
                        f"Label mismatch for '{key_name}': "
                        f"expected '{expected_label}', got '{actual_clean}'"
                    ),
                ))
            else:
                pass_count += 1
        findings.append(Finding(
            severity="PASS",
            tab=tab_name,
            check="Check1-Labels",
            cell="—",
            message=f"{pass_count}/{len(keys)} labels matched",
        ))
    return findings


# ---------------------------------------------------------------------------
# Check 2 — Formula error scan
# ---------------------------------------------------------------------------

def check_formula_errors(wb) -> List[Finding]:
    findings = []
    for tab_name in wb.sheetnames:
        ws = wb[tab_name]
        error_count = 0
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if _is_formula_error(val):
                    findings.append(Finding(
                        severity="ERROR",
                        tab=tab_name,
                        check="Check2-FormulaErrors",
                        cell=cell.coordinate,
                        message=f"Formula error: {val}",
                    ))
                    error_count += 1
        if error_count == 0:
            findings.append(Finding(
                severity="PASS",
                tab=tab_name,
                check="Check2-FormulaErrors",
                cell="—",
                message="No formula errors found",
            ))
    return findings


# ---------------------------------------------------------------------------
# Check 3 — Cross-sheet reference integrity
# ---------------------------------------------------------------------------

_XREF_RE = re.compile(r"='?([^'!]+)'?!([A-Z]+)(\d+)")


def check_cross_sheet_refs(wb) -> List[Finding]:
    findings = []
    sheet_names = set(wb.sheetnames)
    ref_count = 0
    error_count = 0

    for tab_name in wb.sheetnames:
        ws = wb[tab_name]
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not isinstance(val, str):
                    continue
                for match in _XREF_RE.finditer(val):
                    ref_tab, ref_col, ref_row_str = match.group(1), match.group(2), match.group(3)
                    ref_row = int(ref_row_str)
                    ref_count += 1

                    if ref_tab not in sheet_names:
                        findings.append(Finding(
                            severity="ERROR",
                            tab=tab_name,
                            check="Check3-CrossSheetRefs",
                            cell=cell.coordinate,
                            message=(
                                f"References non-existent tab '{ref_tab}' "
                                f"(formula: {val[:60]})"
                            ),
                        ))
                        error_count += 1
                        continue

                    # Check if referenced row has any data
                    ref_ws = wb[ref_tab]
                    if not _row_has_data(ref_ws, ref_row):
                        findings.append(Finding(
                            severity="WARN",
                            tab=tab_name,
                            check="Check3-CrossSheetRefs",
                            cell=cell.coordinate,
                            message=(
                                f"References '{ref_tab}'!{ref_col}{ref_row} "
                                f"but that row appears empty"
                            ),
                        ))

    if error_count == 0:
        findings.append(Finding(
            severity="PASS",
            tab="(all tabs)",
            check="Check3-CrossSheetRefs",
            cell="—",
            message=f"All {ref_count} cross-sheet references point to valid tabs",
        ))
    return findings


# ---------------------------------------------------------------------------
# Check 4 — Balance continuity
# ---------------------------------------------------------------------------

def check_balance_continuity(wb) -> List[Finding]:
    """Ending value of month N must equal beginning of month N+1 (within $0.01)."""
    findings = []

    for tab_name, col_map in MONTHLY_COLUMNS.items():
        if tab_name not in wb.sheetnames:
            continue
        if tab_name not in REGISTRY:
            continue

        ws = wb[tab_name]
        reg = REGISTRY[tab_name]

        # Find the ending column letter
        ending_col = None
        for col_letter, field in col_map.items():
            if field == "ending":
                ending_col = col_letter
                break
        if ending_col is None:
            continue

        # Beginning is always col B
        beginning_col = "B"

        # Find monthly rows: jan row to dec row
        jan_row = None
        dec_row = None
        for key, (col, row, _) in reg.items():
            if key == "monthly_jan":
                jan_row = row
            if key == "monthly_dec":
                dec_row = row

        if jan_row is None or dec_row is None:
            continue

        # Build list of rows that have data in both beginning and ending cols
        month_data = []
        for r in range(jan_row, dec_row + 1):
            begin_val = _to_float(_raw(ws[f"{beginning_col}{r}"]))
            ending_val = _to_float(_raw(ws[f"{ending_col}{r}"]))
            if begin_val is not None or ending_val is not None:
                month_data.append((r, begin_val, ending_val))

        errors_found = 0
        for i in range(len(month_data) - 1):
            row_n, begin_n, end_n = month_data[i]
            row_n1, begin_n1, end_n1 = month_data[i + 1]
            if end_n is None or begin_n1 is None:
                continue
            diff = abs(end_n - begin_n1)
            if diff > 0.01:
                findings.append(Finding(
                    severity="ERROR",
                    tab=tab_name,
                    check="Check4-BalanceContinuity",
                    cell=f"{ending_col}{row_n}→{beginning_col}{row_n1}",
                    message=(
                        f"Ending row {row_n} ({end_n:,.2f}) ≠ "
                        f"Beginning row {row_n1} ({begin_n1:,.2f}), "
                        f"diff={diff:,.4f}"
                    ),
                ))
                errors_found += 1

        if errors_found == 0 and month_data:
            findings.append(Finding(
                severity="PASS",
                tab=tab_name,
                check="Check4-BalanceContinuity",
                cell="—",
                message=f"Balance continuity OK across {len(month_data)} months with data",
            ))

    return findings


# ---------------------------------------------------------------------------
# Check 5 — Accounting identity
# ---------------------------------------------------------------------------

def check_accounting_identity(wb) -> List[Finding]:
    """expected_ending = beginning + inflow - outflow + dividends + market_change (tol $1)."""
    findings = []

    for tab_name, col_map in MONTHLY_COLUMNS.items():
        if tab_name not in wb.sheetnames:
            continue
        if tab_name not in REGISTRY:
            continue

        ws = wb[tab_name]
        reg = REGISTRY[tab_name]

        # Map field names to column letters
        field_to_col = {v: k for k, v in col_map.items()}

        beginning_col = field_to_col.get("beginning", "B")
        ending_col = field_to_col.get("ending")
        dividends_col = field_to_col.get("dividends")
        market_change_col = field_to_col.get("market_change")

        # Inflow: deposits/additions/contributions
        inflow_col = None
        outflow_col = None
        for field, col in field_to_col.items():
            if "deposit" in field or "addition" in field or "contribution" in field:
                inflow_col = col
            if "withdrawal" in field or "subtraction" in field or "distribution" in field:
                outflow_col = col

        if not all([ending_col, dividends_col, market_change_col, inflow_col, outflow_col]):
            findings.append(Finding(
                severity="WARN",
                tab=tab_name,
                check="Check5-AccountingIdentity",
                cell="—",
                message="Could not determine all column mappings, skipping",
            ))
            continue

        jan_row = dec_row = None
        for key, (col, row, _) in reg.items():
            if key == "monthly_jan":
                jan_row = row
            if key == "monthly_dec":
                dec_row = row

        if jan_row is None or dec_row is None:
            continue

        errors_found = 0
        checked = 0
        for r in range(jan_row, dec_row + 1):
            beginning = _to_float(_raw(ws[f"{beginning_col}{r}"]))
            ending = _to_float(_raw(ws[f"{ending_col}{r}"]))
            inflow = _to_float(_raw(ws[f"{inflow_col}{r}"]))
            outflow = _to_float(_raw(ws[f"{outflow_col}{r}"]))
            dividends = _to_float(_raw(ws[f"{dividends_col}{r}"]))
            market_change = _to_float(_raw(ws[f"{market_change_col}{r}"]))

            # Skip rows without sufficient data
            if any(v is None for v in [beginning, ending, inflow, outflow, dividends, market_change]):
                continue

            checked += 1
            expected = beginning + inflow - outflow + dividends + market_change
            diff = abs(ending - expected)
            if diff > 1.0:
                findings.append(Finding(
                    severity="ERROR",
                    tab=tab_name,
                    check="Check5-AccountingIdentity",
                    cell=f"row {r}",
                    message=(
                        f"Accounting identity fail: actual ending={ending:,.2f}, "
                        f"expected={expected:,.2f}, diff={diff:,.4f}"
                    ),
                ))
                errors_found += 1

        if errors_found == 0 and checked > 0:
            findings.append(Finding(
                severity="PASS",
                tab=tab_name,
                check="Check5-AccountingIdentity",
                cell="—",
                message=f"Accounting identity OK for {checked} months with full data",
            ))
        elif checked == 0:
            findings.append(Finding(
                severity="WARN",
                tab=tab_name,
                check="Check5-AccountingIdentity",
                cell="—",
                message="No complete monthly rows found (formula-only cells?)",
            ))

    return findings


# ---------------------------------------------------------------------------
# Check 6 — Holdings total
# ---------------------------------------------------------------------------

def check_holdings_totals(wb) -> List[Finding]:
    """Sum individual holdings and compare to total row (tol $0.01)."""
    findings = []

    for tab_name, config in HOLDINGS_ROWS.items():
        if tab_name not in wb.sheetnames:
            continue

        ws = wb[tab_name]
        first = config["first"]
        last = config["last"]
        total_row = config["total"]

        for col_key in ("mv_col", "cb_col", "gl_col"):
            col = config[col_key]
            total_cell = ws[f"{col}{total_row}"]
            total_val_raw = _raw(total_cell)

            # Skip if total is a formula — can't evaluate without Excel
            if _is_formula(total_val_raw):
                findings.append(Finding(
                    severity="WARN",
                    tab=tab_name,
                    check="Check6-HoldingsTotals",
                    cell=f"{col}{total_row}",
                    message=(
                        f"Total cell {col}{total_row} is a formula string; "
                        f"skipping {col_key} comparison (no live Excel)"
                    ),
                ))
                continue

            total_val = _to_float(total_val_raw)
            if total_val is None:
                findings.append(Finding(
                    severity="WARN",
                    tab=tab_name,
                    check="Check6-HoldingsTotals",
                    cell=f"{col}{total_row}",
                    message=f"Total cell {col}{total_row} has no numeric value (value={repr(total_val_raw)})",
                ))
                continue

            # Sum individual rows
            row_sum = 0.0
            any_formula = False
            for r in range(first, last + 1):
                v = _to_float(_raw(ws[f"{col}{r}"]))
                if v is None:
                    # Check if it's a formula
                    raw_v = _raw(ws[f"{col}{r}"])
                    if _is_formula(raw_v):
                        any_formula = True
                    # Treat None/formula as 0 for sum purposes (partial data)
                    v = 0.0
                row_sum += v

            diff = abs(row_sum - total_val)
            if diff > 0.01:
                findings.append(Finding(
                    severity="ERROR",
                    tab=tab_name,
                    check="Check6-HoldingsTotals",
                    cell=f"{col}{total_row}",
                    message=(
                        f"{col_key}: sum of rows {first}-{last} = {row_sum:,.2f}, "
                        f"total row = {total_val:,.2f}, diff={diff:,.4f}"
                        + (" (some cells are formulas, sum may be understated)" if any_formula else "")
                    ),
                ))
            else:
                findings.append(Finding(
                    severity="PASS",
                    tab=tab_name,
                    check="Check6-HoldingsTotals",
                    cell=f"{col}{total_row}",
                    message=(
                        f"{col_key}: sum={row_sum:,.2f}, total={total_val:,.2f} ✓"
                        + (" (some detail cells are formulas)" if any_formula else "")
                    ),
                ))

    return findings


# ---------------------------------------------------------------------------
# Check 7 — YTD gain consistency
# ---------------------------------------------------------------------------

YTD_TABS = ["Fidelity Brokerage", "Fidelity Roth IRA", "Fidelity HSA", "Robinhood"]


def check_ytd_gain(wb) -> List[Finding]:
    """total_ytd_gain = unrealized + realized + dividends (tol $1)."""
    findings = []

    for tab_name in YTD_TABS:
        if tab_name not in wb.sheetnames:
            continue
        if tab_name not in REGISTRY:
            continue

        ws = wb[tab_name]
        reg = REGISTRY[tab_name]

        def get_val(key):
            if key not in reg:
                return None, None
            col, row, _ = reg[key]
            cell = ws[f"{col}{row}"]
            raw = _raw(cell)
            return raw, _to_float(raw)

        total_raw, total_val = get_val("total_ytd")
        unreal_raw, unreal_val = get_val("unrealized")
        real_raw, real_val = get_val("realized")
        div_raw, div_val = get_val("dividends")

        # If any value is a formula string, WARN instead of checking
        formula_fields = []
        for fname, raw in [
            ("total_ytd", total_raw),
            ("unrealized", unreal_raw),
            ("realized", real_raw),
            ("dividends", div_raw),
        ]:
            if _is_formula(raw):
                formula_fields.append(fname)

        if formula_fields:
            findings.append(Finding(
                severity="WARN",
                tab=tab_name,
                check="Check7-YTDGain",
                cell="—",
                message=(
                    f"Formula strings found in: {', '.join(formula_fields)}; "
                    f"cannot verify YTD without live Excel evaluation"
                ),
            ))
            continue

        if any(v is None for v in [total_val, unreal_val, real_val, div_val]):
            findings.append(Finding(
                severity="WARN",
                tab=tab_name,
                check="Check7-YTDGain",
                cell="—",
                message="One or more YTD values are missing or non-numeric; skipping check",
            ))
            continue

        expected = unreal_val + real_val + div_val
        diff = abs(total_val - expected)
        if diff > 1.0:
            findings.append(Finding(
                severity="ERROR",
                tab=tab_name,
                check="Check7-YTDGain",
                cell="—",
                message=(
                    f"YTD gain mismatch: total={total_val:,.2f}, "
                    f"unrealized+realized+dividends={expected:,.2f}, diff={diff:,.4f}"
                ),
            ))
        else:
            findings.append(Finding(
                severity="PASS",
                tab=tab_name,
                check="Check7-YTDGain",
                cell="—",
                message=(
                    f"YTD gain consistent: total={total_val:,.2f} ≈ "
                    f"unrealized({unreal_val:,.2f}) + realized({real_val:,.2f}) "
                    f"+ dividends({div_val:,.2f})"
                ),
            ))

    return findings


# ---------------------------------------------------------------------------
# Output formatting
# ---------------------------------------------------------------------------

_SEV_LABEL = {"ERROR": "FAIL", "WARN": "WARN", "PASS": "PASS"}


def format_findings(findings: List[Finding]) -> str:
    """Format findings into a readable report string."""
    lines = []
    lines.append("=" * 72)
    lines.append("  WORKBOOK VALIDATION REPORT")
    lines.append("=" * 72)

    for f in findings:
        label = _SEV_LABEL.get(f.severity, f.severity)
        lines.append(f"[{label:<4}] {f.tab}: {f.check} {f.cell} -- {f.message}")

    lines.append("")
    lines.append("=" * 72)
    errors = sum(1 for f in findings if f.severity == "ERROR")
    warns  = sum(1 for f in findings if f.severity == "WARN")
    passes = sum(1 for f in findings if f.severity == "PASS")
    total  = len(findings)
    lines.append(
        f"  Summary: {total} findings -- "
        f"{errors} ERROR(S) | {warns} WARN(S) | {passes} PASS(ES)"
    )
    lines.append("=" * 72)
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# validate_full
# ---------------------------------------------------------------------------

def validate_structural(workbook_path: str, tab_name: str) -> List[Finding]:
    """Lightweight validation for a single tab -- called by rebuild scripts after saving."""
    path = Path(workbook_path)
    if not path.exists():
        return [Finding("ERROR", tab_name, "file", "", f"Workbook not found: {path}")]

    wb = openpyxl.load_workbook(str(path), data_only=False)
    findings: List[Finding] = []

    # Check 1: Label matching (this tab only)
    findings.extend([f for f in check_labels(wb) if f.tab == tab_name])

    # Check 4: Balance continuity
    if tab_name in MONTHLY_COLUMNS:
        findings.extend([f for f in check_balance_continuity(wb) if f.tab == tab_name])

    # Check 5: Accounting identity
    if tab_name in MONTHLY_COLUMNS:
        findings.extend([f for f in check_accounting_identity(wb) if f.tab == tab_name])

    # Check 6: Holdings total
    if tab_name in HOLDINGS_ROWS:
        findings.extend([f for f in check_holdings_totals(wb) if f.tab == tab_name])

    return findings


def check_com_deep_eval(workbook_path: str) -> List[Finding]:
    """Optional deep eval: open in Excel via COM, force recalculation, check for errors."""
    findings: List[Finding] = []
    try:
        import win32com.client
    except ImportError:
        findings.append(Finding("WARN", "ALL", "deep_eval", "",
                                "Deep eval skipped (requires: pip install pywin32)"))
        return findings

    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        abs_path = str(Path(workbook_path).resolve())
        wb_com = excel.Workbooks.Open(abs_path, ReadOnly=True)
        excel.Calculate()

        error_codes = {
            -2146826281: "#DIV/0!", -2146826246: "#N/A", -2146826259: "#NAME?",
            -2146826265: "#NULL!", -2146826252: "#NUM!", -2146826288: "#VALUE!",
            -2146826225: "#REF!",
        }

        for ws_com in wb_com.Worksheets:
            tab_name = ws_com.Name
            used = ws_com.UsedRange
            errors_found = 0
            for row in range(1, used.Rows.Count + 1):
                for col in range(1, used.Columns.Count + 1):
                    try:
                        cell = used.Cells(row, col)
                        val = cell.Value
                        if isinstance(val, int) and val in error_codes:
                            findings.append(Finding(
                                "ERROR", tab_name, "deep_eval",
                                f"R{row}C{col}",
                                f"Formula evaluates to {error_codes[val]}"
                            ))
                            errors_found += 1
                    except Exception:
                        pass
            if errors_found == 0:
                findings.append(Finding("PASS", tab_name, "deep_eval", "",
                                        "No formula errors after recalculation"))

        wb_com.Close(SaveChanges=False)
        excel.Quit()

    except Exception as e:
        findings.append(Finding("WARN", "ALL", "deep_eval", "",
                                f"Deep eval failed: {e}"))

    return findings


def validate_full(workbook_path: str, deep: bool = False) -> List[Finding]:
    """Load workbook (formula strings, not data_only) and run all 7 checks."""
    path = Path(workbook_path)
    if not path.exists():
        print(f"ERROR: workbook not found: {path}", file=sys.stderr)
        sys.exit(2)

    print(f"Loading workbook: {path}")
    wb = openpyxl.load_workbook(str(path), data_only=False)

    all_findings: List[Finding] = []

    print("Running Check 1 — Label matching...")
    all_findings.extend(check_labels(wb))

    print("Running Check 2 — Formula error scan...")
    all_findings.extend(check_formula_errors(wb))

    print("Running Check 3 — Cross-sheet reference integrity...")
    all_findings.extend(check_cross_sheet_refs(wb))

    print("Running Check 4 — Balance continuity...")
    all_findings.extend(check_balance_continuity(wb))

    print("Running Check 5 — Accounting identity...")
    all_findings.extend(check_accounting_identity(wb))

    print("Running Check 6 — Holdings totals...")
    all_findings.extend(check_holdings_totals(wb))

    print("Running Check 7 — YTD gain consistency...")
    all_findings.extend(check_ytd_gain(wb))

    if deep:
        print("Running Check 8 -- COM deep eval (live Excel)...")
        all_findings.extend(check_com_deep_eval(workbook_path))

    print()
    return all_findings


# ---------------------------------------------------------------------------
# CLI main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Validate 2026_Portfolio_Analysis.xlsx workbook integrity."
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        default="2026_Portfolio_Analysis.xlsx",
        help="Path to the workbook (default: 2026_Portfolio_Analysis.xlsx)",
    )
    parser.add_argument(
        "--deep",
        action="store_true",
        help="Enable deep COM-based formula evaluation (requires live Excel)",
    )
    args = parser.parse_args()

    findings = validate_full(args.workbook, deep=args.deep)
    print(format_findings(findings))

    errors = sum(1 for f in findings if f.severity == "ERROR")
    sys.exit(1 if errors > 0 else 0)


if __name__ == "__main__":
    main()
