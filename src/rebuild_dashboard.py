#!/usr/bin/env python3
"""Rebuild Dashboard tab with reordered sections."""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from daily_snapshot import load_snapshot, load_previous_snapshot, compute_daily_summary
from registry import define_named_ranges, REGISTRY, HOLDINGS_ROWS, _TAB_PREFIX
import datetime
import json
from pathlib import Path

XLSX = "2026_Portfolio_Analysis.xlsx"
EXTRACT_OUTPUT = Path(os.environ.get("PLUTUS_PIPELINE_DIR", str(Path.home() / ".portfolio_extract" / "pipeline"))) / "extract_output"

# Hardcoded fallback benchmarks (used if no cached file found)
_DEFAULT_BENCHMARKS = [
    ('S&P 500', -0.04021),
    ('Dow Jones', -0.03881),
    ('NASDAQ', -0.058378),
]


def _load_benchmarks():
    """Load benchmarks from most recent cached file, falling back to hardcoded defaults."""
    bench_files = sorted(EXTRACT_OUTPUT.glob("benchmarks_*.json"), reverse=True) if EXTRACT_OUTPUT.exists() else []
    for f in bench_files:
        try:
            data = json.loads(f.read_text())
            result = [(k, v) for k, v in data.items() if not k.startswith("_") and isinstance(v, (int, float))]
            if result:
                return result
        except Exception:
            continue
    return _DEFAULT_BENCHMARKS

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HEADER_ALIGN = Alignment(horizontal='center')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
GREEN_FONT = Font(name='Arial', size=10, color='008000')
BLACK_FONT = Font(name='Arial', size=10)
BLUE_FONT = Font(name='Arial', size=10, color='0000FF')
BOLD_FONT = Font(name='Arial', size=10, bold=True)
SECTION_FONT = Font(name='Arial', size=12, bold=True)
TITLE_FONT = Font(name='Arial', size=14, bold=True)
NOTE_FONT = Font(name='Arial', size=9, italic=True, color='666666')

DOLLAR = '$#,##0.00'
PCT = '0.00%'


def header_row(ws, row, labels):
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=1 + i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = THIN_BORDER


def cell(ws, row, col, value=None, font=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:
        c.font = font
    c.border = THIN_BORDER
    if fmt:
        c.number_format = fmt
    return c


def main():
    wb = openpyxl.load_workbook(XLSX)

    if 'Dashboard' in wb.sheetnames:
        old_idx = wb.sheetnames.index('Dashboard')
        del wb['Dashboard']
        ws = wb.create_sheet('Dashboard', old_idx)
    else:
        ws = wb.create_sheet('Dashboard', 0)

    # Column widths
    for col, w in {'A': 48, 'B': 16, 'C': 26, 'D': 37, 'E': 29, 'F': 30,
                   'G': 26, 'H': 28, 'I': 14, 'J': 12}.items():
        ws.column_dimensions[col].width = w

    # Load daily snapshots for summary
    today_str = datetime.date.today().isoformat()
    snap_today = load_snapshot(today_str)
    snap_prev = load_previous_snapshot(today_str)
    daily_summary = compute_daily_summary(snap_today, snap_prev) if snap_today and snap_prev else None

    row = 1

    # ==================================================================
    # TITLE
    # ==================================================================
    today_label = datetime.date.today().strftime('%B %d, %Y')
    ws.cell(row=row, column=1, value=f'Portfolio Analysis — {datetime.date.today().year} YTD (as of {today_label})').font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value='Green = cross-sheet reference | Black = formula | Blue = hardcoded').font = NOTE_FONT
    row += 2

    # ==================================================================
    # SECTION 1: DAILY SUMMARY
    # ==================================================================
    ws.cell(row=row, column=1, value='DAILY SUMMARY').font = SECTION_FONT
    row += 1

    prose_font = Font(name='Arial', size=10)
    if daily_summary:
        mv = snap_today['liquid_total_mv']
        chg = daily_summary['liquid_change']
        pct = daily_summary['liquid_change_pct']
        prev_date = daily_summary['prev_date']
        direction = 'gain' if chg >= 0 else 'loss'

        ws.cell(row=row, column=1,
                value=f'Your liquid portfolio is valued at ${mv:,.2f}. '
                      f'This represents a ${abs(chg):,.2f} ({abs(pct):.2%}) {direction} '
                      f'from the last trading day ({prev_date}).').font = prose_font
        row += 1

        movers = daily_summary.get('top_movers', [])
        if movers:
            names = ', '.join(f'{m["ticker"]} ({"+" if m["change_pct"] >= 0 else ""}{m["change_pct"]:.1%})'
                              for m in movers[:10])
            ws.cell(row=row, column=1,
                    value=f'Top movers today: {names}.').font = prose_font
        else:
            ws.cell(row=row, column=1,
                    value='No securities moved more than 10% today.').font = prose_font
        row += 1

        stale = snap_today.get('stale_sources', [])
        if stale:
            ws.cell(row=row, column=1,
                    value=f'Note: The following sources used fallback data: {", ".join(stale)}.').font = NOTE_FONT
            row += 1
    else:
        ws.cell(row=row, column=1,
                value='Daily summary requires at least 2 pipeline runs. Will populate after next run.').font = NOTE_FONT
        row += 1

    row += 1

    # ==================================================================
    # SECTION 2: YTD BENCHMARK COMPARISON (rows reserved, filled after Account Overview)
    # ==================================================================
    ws.cell(row=row, column=1, value='YTD BENCHMARK COMPARISON').font = SECTION_FONT
    row += 1
    # Placeholder header — overwritten with dynamic columns after Account Overview is built
    header_row(ws, row, ['Benchmark', 'Return', 'Alpha: Liquid Portfolio'])
    bench_hdr = row
    row += 1

    benchmarks = _load_benchmarks()

    sp_row = row
    bench_data_rows = {}
    for name, ret in benchmarks:
        cell(ws, row, 1, name, font=BLUE_FONT)
        cell(ws, row, 2, ret, font=BLUE_FONT, fmt=PCT)
        # Alpha columns 3-8 filled after Account Overview is built
        bench_data_rows[name] = row
        row += 1

    row += 1

    # ==================================================================
    # YTD INVESTMENT GAIN
    # ==================================================================
    # Build formulas dynamically based on which account tabs exist
    existing_tabs = set(wb.sheetnames)

    _gain_prefixes = []
    if 'Fidelity Brokerage' in existing_tabs:
        _gain_prefixes.append('fid_brok')
    if 'Fidelity Roth IRA' in existing_tabs:
        _gain_prefixes.append('roth_ira')
    if 'Fidelity HSA' in existing_tabs:
        _gain_prefixes.append('fid_hsa')
    if 'Robinhood' in existing_tabs:
        _gain_prefixes.append('robinhood')

    def _sum_formula(suffix):
        """Build a SUM-style formula from available account prefixes."""
        parts = [f'{p}_{suffix}' for p in _gain_prefixes]
        if not parts:
            return 0
        return '=' + '+'.join(parts)

    ws.cell(row=row, column=1, value='YTD INVESTMENT GAIN').font = SECTION_FONT
    row += 1
    header_row(ws, row, ['Metric', 'Value', 'Note'])
    row += 1

    cell(ws, row, 1, 'Dividends/Income')
    cell(ws, row, 2, _sum_formula('dividends'), font=GREEN_FONT, fmt=DOLLAR)
    cell(ws, row, 3, 'Sum across all accounts with dividend income', font=NOTE_FONT)
    gain_div_row = row
    row += 1

    cell(ws, row, 1, 'Unrealized Gain/Loss')
    cell(ws, row, 2, _sum_formula('unrealized'), font=GREEN_FONT, fmt=DOLLAR)
    cell(ws, row, 3, 'Current holdings vs. cost basis (all-time)', font=NOTE_FONT)
    gain_unreal_row = row
    row += 1

    cell(ws, row, 1, 'Realized Gain/Loss (2026)')
    cell(ws, row, 2, _sum_formula('realized'), font=GREEN_FONT, fmt=DOLLAR)
    cell(ws, row, 3, 'From sold positions YTD', font=NOTE_FONT)
    gain_real_row = row
    row += 1

    gain_total_parts = [f'B{gain_div_row}', f'B{gain_unreal_row}', f'B{gain_real_row}']

    if '401(k)' in existing_tabs:
        cell(ws, row, 1, '401(k) Investment Gain', font=BLACK_FONT)
        cell(ws, row, 2, "=IFERROR(k401_total_inv_gain,0)", font=GREEN_FONT, fmt=DOLLAR)
        cell(ws, row, 3, 'Change in value + fees (quarterly)', font=NOTE_FONT)
        gain_401k_row = row
        gain_total_parts.append(f'B{gain_401k_row}')
        row += 1

    cell(ws, row, 1, 'Total YTD Investment Gain', font=BOLD_FONT)
    cell(ws, row, 2, '=' + '+'.join(gain_total_parts), font=BOLD_FONT, fmt=DOLLAR)
    cell(ws, row, 3, 'Dividends + Unrealized + Realized' + (' + 401(k)' if '401(k)' in existing_tabs else ''), font=NOTE_FONT)
    row += 2

    # ==================================================================
    # ACCOUNT OVERVIEW
    # ==================================================================
    ws.cell(row=row, column=1, value='ACCOUNT OVERVIEW').font = SECTION_FONT
    row += 1
    header_row(ws, row, ['Account', 'Beginning', 'Ending', 'Net Cash Flow', 'Time-Weighted Return (YTD)', 'Money-Weighted Return (YTD)', 'Cost Basis Return', 'Alpha (YTD)'])
    row += 1

    # --- Liquid accounts (only include those with existing tabs) ---
    _all_liquid = [
        ('Fidelity Brokerage', 'fid_brok', {
            'B': "=fid_brok_monthly_begin",                          # Beginning
            'C': "=fid_brok_holdings_total_mv",                      # Ending
            'D': "=fid_brok_monthly_totals_add-fid_brok_monthly_totals_sub",  # Net Cash Flow
            'E': "=fid_brok_TWR",                                    # TWR (YTD)
            'F': "=fid_brok_MWRR",                                  # MWRR (YTD)
            'G': "=fid_brok_cb_return",                              # CB Return
        }),
        ('Fidelity Roth IRA', 'roth_ira', {
            'B': "=roth_ira_monthly_begin",
            'C': "=roth_ira_holdings_total_mv",
            'D': "=roth_ira_monthly_totals_add-roth_ira_monthly_totals_sub",
            'E': "=roth_ira_TWR",
            'F': "=roth_ira_MWRR",
            'G': "=roth_ira_cb_return",
        }),
        ('Fidelity HSA', 'fid_hsa', {
            'B': "=fid_hsa_monthly_begin",
            'C': "=fid_hsa_holdings_total_mv",
            'D': "=fid_hsa_monthly_totals_add-fid_hsa_monthly_totals_sub",
            'E': "=fid_hsa_TWR",
            'F': "=fid_hsa_MWRR",
            'G': "=fid_hsa_cb_return",
        }),
        ('Robinhood', 'robinhood', {
            'B': "=robinhood_monthly_begin",
            'C': "=robinhood_net_portfolio",                         # NET PORTFOLIO VALUE
            'D': "=robinhood_monthly_totals_add-robinhood_monthly_totals_sub",
            'E': "=robinhood_TWR",
            'F': "=robinhood_MWRR",
            'G': "=robinhood_cb_return",
        }),
    ]
    liquid_accounts = [(name, prefix, refs) for name, prefix, refs in _all_liquid
                       if name in existing_tabs]

    acct_rows = {}
    acct_first = row
    liquid_first = row

    # Write liquid account rows
    for name, prefix, refs in liquid_accounts:
        cell(ws, row, 1, name, font=GREEN_FONT)
        for i, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], 2):
            val = refs[col_letter]
            cell(ws, row, i, val, font=GREEN_FONT, fmt=DOLLAR if col_letter in 'BCD' else PCT)
        acct_rows[name] = row
        row += 1

    liquid_last = row - 1

    # Liquid subtotal
    cell(ws, row, 1, 'LIQUID SUBTOTAL', font=BOLD_FONT)
    cell(ws, row, 2, f'=SUM(B{liquid_first}:B{liquid_last})', font=BOLD_FONT, fmt=DOLLAR)
    cell(ws, row, 3, f'=SUM(C{liquid_first}:C{liquid_last})', font=BOLD_FONT, fmt=DOLLAR)
    cell(ws, row, 4, f'=SUM(D{liquid_first}:D{liquid_last})', font=BOLD_FONT, fmt=DOLLAR)
    # TWR (YTD) = Liquid Portfolio TWR (filled later via liquid_twr_row)
    cell(ws, row, 5, None, fmt=PCT)  # Liquid TWR — filled in forward-ref section
    cell(ws, row, 6, None, fmt=PCT)  # MWRR — N/A at subtotal level
    # Cost Basis Return = total G/L / total CB — use formula referencing account tab totals
    cell(ws, row, 7, None, fmt=PCT)  # filled later
    cell(ws, row, 8, None, fmt=PCT)  # Alpha — filled later
    for col in range(2, 9):
        ws.cell(row=row, column=col).border = THIN_BORDER
    liquid_sub_row = row
    row += 2  # skip a row after Liquid Subtotal

    # Cash (separate line, same formatting as subtotals)
    has_cash = 'Cash' in existing_tabs
    if has_cash:
        cell(ws, row, 1, 'EXTERNAL CASH', font=BOLD_FONT)
        cell(ws, row, 2, "=IFERROR(cash_total_cash,0)", font=BOLD_FONT, fmt=DOLLAR)
        cell(ws, row, 3, "=IFERROR(cash_total_cash,0)", font=BOLD_FONT, fmt=DOLLAR)
        for col_idx in [4, 5, 6, 7]:
            cell(ws, row, col_idx, 'N/A', font=BLUE_FONT)
        for col in range(2, 9):
            ws.cell(row=row, column=col).border = THIN_BORDER
        acct_rows['Cash'] = row
        row += 2  # skip a row after Cash

    # --- Illiquid accounts (only include those with existing tabs) ---
    illiquid_first = row
    has_illiquid = False

    # 401(k) — has TWR/MWRR/CB Return
    if '401(k)' in existing_tabs:
        has_illiquid = True
        cell(ws, row, 1, '401(k)', font=GREEN_FONT)
        for i, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], 2):
            refs_401k = {
                'B': "=IFERROR(k401_quarterly_first,0)", 'C': "=IFERROR(k401_holdings_total,0)",
                'D': "=IFERROR(k401_ytd_totals,0)",
                'E': "=IFERROR(k401_TWR,\"N/A\")", 'F': "=IFERROR(k401_MWRR,\"N/A\")",
                'G': "=IFERROR(k401_cb_return,\"N/A\")",
            }
            cell(ws, row, i, refs_401k[col_letter], font=GREEN_FONT,
                 fmt=DOLLAR if col_letter in 'BCD' else PCT)
        acct_rows['401(k)'] = row
        row += 1

    # Angel Investments — only Cost Basis Return is meaningful
    if 'Angel Investments' in existing_tabs:
        has_illiquid = True
        cell(ws, row, 1, 'Angel Investments', font=GREEN_FONT)
        cell(ws, row, 2, "=IFERROR(angel_total_invested,0)", font=GREEN_FONT, fmt=DOLLAR)
        cell(ws, row, 3, "=IFERROR(angel_total_current,0)", font=GREEN_FONT, fmt=DOLLAR)
        cell(ws, row, 4, 'N/A', font=BLUE_FONT)
        cell(ws, row, 5, 'N/A', font=BLUE_FONT)  # TWR — N/A (no periodic cash flows)
        cell(ws, row, 6, 'N/A', font=BLUE_FONT)  # MWRR — N/A
        cell(ws, row, 7, "=IF(IFERROR(angel_total_invested,0)=0,0,(angel_total_current-angel_total_invested)/angel_total_invested)",
             font=GREEN_FONT, fmt=PCT)  # Cost Basis Return (MOIC-1)
        acct_rows['Angel Investments'] = row
        row += 1

    illiquid_last = row - 1
    illiquid_sub_row = None

    if has_illiquid:
        # Illiquid subtotal
        cell(ws, row, 1, 'ILLIQUID SUBTOTAL', font=BOLD_FONT)
        cell(ws, row, 2, f'=SUM(B{illiquid_first}:B{illiquid_last})', font=BOLD_FONT, fmt=DOLLAR)
        cell(ws, row, 3, f'=SUM(C{illiquid_first}:C{illiquid_last})', font=BOLD_FONT, fmt=DOLLAR)
        if '401(k)' in acct_rows:
            cell(ws, row, 4, f'=D{acct_rows["401(k)"]}', font=BOLD_FONT, fmt=DOLLAR)
        else:
            cell(ws, row, 4, 0, font=BOLD_FONT, fmt=DOLLAR)
        cell(ws, row, 5, 'N/A', font=BLUE_FONT)
        cell(ws, row, 6, 'N/A', font=BLUE_FONT)
        cell(ws, row, 7, f'=IF(B{row}=0,0,(C{row}-B{row})/B{row})', font=BLACK_FONT, fmt=PCT)
        for col in range(2, 9):
            ws.cell(row=row, column=col).border = THIN_BORDER
        illiquid_sub_row = row
        row += 2

    # Total Portfolio row
    total_parts_b = [f'B{liquid_sub_row}']
    total_parts_c = [f'C{liquid_sub_row}']
    total_parts_d = [f'D{liquid_sub_row}']
    if has_cash:
        total_parts_b.append(f'B{acct_rows["Cash"]}')
        total_parts_c.append(f'C{acct_rows["Cash"]}')
    if illiquid_sub_row:
        total_parts_b.append(f'B{illiquid_sub_row}')
        total_parts_c.append(f'C{illiquid_sub_row}')
        total_parts_d.append(f'D{illiquid_sub_row}')

    cell(ws, row, 1, 'TOTAL PORTFOLIO', font=BOLD_FONT)
    cell(ws, row, 2, '=' + '+'.join(total_parts_b), font=BOLD_FONT, fmt=DOLLAR)
    cell(ws, row, 3, '=' + '+'.join(total_parts_c), font=BOLD_FONT, fmt=DOLLAR)
    cell(ws, row, 4, '=' + '+'.join(total_parts_d), font=BOLD_FONT, fmt=DOLLAR)
    for col in [5, 6, 7, 8]:
        ws.cell(row=row, column=col).border = THIN_BORDER
    total_row = row
    row += 2

    # ==================================================================
    # Compute Liquid Portfolio TWR from account-tab monthly data
    # (aggregated Modified Dietz return across Fid Brok + Roth + HSA + Robinhood)
    # ==================================================================
    _all_liquid_tabs = {
        'Fidelity Brokerage': 44,
        'Fidelity Roth IRA':  34,
        'Fidelity HSA':       31,
        'Robinhood':          42,
    }
    liquid_tabs = {k: v for k, v in _all_liquid_tabs.items() if k in existing_tabs}
    growth_factors = []
    for i in range(12):
        cb = ce = ca = cs = 0
        for tab_name, first_row in liquid_tabs.items():
            tab_ws = wb[tab_name]
            r = first_row + i
            b = tab_ws.cell(r, 2).value
            a = tab_ws.cell(r, 3).value
            s = tab_ws.cell(r, 4).value
            e = tab_ws.cell(r, 7).value
            cb += b if isinstance(b, (int, float)) else 0
            ce += e if isinstance(e, (int, float)) else 0
            ca += a if isinstance(a, (int, float)) else 0
            cs += s if isinstance(s, (int, float)) else 0
        if cb == 0 and ce == 0:
            continue
        net_flow = ca - cs
        denom = cb + 0.5 * net_flow
        if denom == 0:
            continue
        ret = (ce - cb - net_flow) / denom
        growth_factors.append(1 + ret)

    liquid_twr = None
    if growth_factors:
        liquid_twr = 1
        for gf in growth_factors:
            liquid_twr *= gf
        liquid_twr -= 1

    # ==================================================================
    # Fill in benchmark alpha columns (now that acct_rows + liquid TWR are known)
    # ==================================================================
    # Map alpha columns to account names (only include present accounts)
    alpha_col_map = []  # list of (col_idx, account_name)
    col_idx = 4
    for name, prefix, refs in liquid_accounts:
        alpha_col_map.append((col_idx, name))
        col_idx += 1
    if '401(k)' in acct_rows:
        alpha_col_map.append((col_idx, '401(k)'))
        col_idx += 1

    # Update benchmark header to match present accounts
    alpha_headers = ['Benchmark', 'Return', 'Alpha: Liquid Portfolio']
    for _, name in alpha_col_map:
        liq_tag = '(liquid)' if name != '401(k)' else '(illiquid)'
        alpha_headers.append(f'Alpha: {name} {liq_tag}')
    header_row(ws, bench_hdr, alpha_headers)

    for bench_name, bench_row in bench_data_rows.items():
        bench_ret = ws.cell(bench_row, 2).value or 0
        # Col 3: Liquid Portfolio alpha
        if liquid_twr is not None and isinstance(bench_ret, (int, float)):
            cell(ws, bench_row, 3, round(liquid_twr - bench_ret, 6), font=BLACK_FONT, fmt=PCT)
        else:
            cell(ws, bench_row, 3, 'N/A', font=BLUE_FONT)
        # Per-account alpha columns
        for acol, aname in alpha_col_map:
            ar = acct_rows[aname]
            cell(ws, bench_row, acol,
                 f'=IF(OR(E{ar}="N/A",B{bench_row}=""),"N/A",E{ar}-B{bench_row})',
                 font=BLACK_FONT, fmt=PCT)

    # Liquid Subtotal TWR + Alpha
    if liquid_twr is not None:
        ws.cell(row=liquid_sub_row, column=5, value=round(liquid_twr, 8))
        ws.cell(row=liquid_sub_row, column=5).font = BLACK_FONT
        ws.cell(row=liquid_sub_row, column=5).number_format = PCT
        ws.cell(row=liquid_sub_row, column=5).border = THIN_BORDER
        sp_ret = ws.cell(sp_row, 2).value or 0
        ws.cell(row=liquid_sub_row, column=8, value=round(liquid_twr - sp_ret, 8))
        ws.cell(row=liquid_sub_row, column=8).font = BLACK_FONT
        ws.cell(row=liquid_sub_row, column=8).number_format = PCT
        ws.cell(row=liquid_sub_row, column=8).border = THIN_BORDER
    else:
        ws.cell(row=liquid_sub_row, column=5, value='N/A')
        ws.cell(row=liquid_sub_row, column=5).font = BLUE_FONT
        ws.cell(row=liquid_sub_row, column=5).border = THIN_BORDER
        ws.cell(row=liquid_sub_row, column=8, value='N/A')
        ws.cell(row=liquid_sub_row, column=8).font = BLUE_FONT
        ws.cell(row=liquid_sub_row, column=8).border = THIN_BORDER

    # Fix Alpha column (H) for individual accounts to reference S&P 500
    for name, acct_row in acct_rows.items():
        if name in ('Angel Investments', 'Cash'):
            continue
        e_col_ref = f'E{acct_row}'
        ws.cell(row=acct_row, column=8,
                value=f'=IF(OR({e_col_ref}="N/A",B{sp_row}=""),"N/A",{e_col_ref}-B{sp_row})')
        ws.cell(row=acct_row, column=8).font = BLACK_FONT
        ws.cell(row=acct_row, column=8).border = THIN_BORDER
        ws.cell(row=acct_row, column=8).number_format = PCT

    # ==================================================================
    # SECTION 5: SECTOR CONCENTRATION
    # ==================================================================
    ws.cell(row=row, column=1, value='SECTOR CONCENTRATION').font = SECTION_FONT
    row += 1
    header_row(ws, row, ['Sector', 'Total Value', '% of Portfolio', 'Fidelity Brokerage', 'Roth IRA',
                         '401(k)', 'HSA', 'Robinhood', 'Angel', '# Holdings'])
    row += 1

    sectors = [
        ('Defense Tech', 20000, 0.20, None, None, None, None, None, 20000, 2),
        ('Diversified/Index', 15000, 0.15, None, None, 15000, None, None, None, 5),
        ('Technology', 15000, 0.15, 5000, 3000, None, 2000, 3000, 2000, 17),
        ('Industrials', 10000, 0.10, 5000, 2000, None, None, None, 3000, 6),
        ('Other', 8000, 0.08, 8000, None, None, None, None, None, 1),
        ('Healthcare', 8000, 0.08, None, None, None, None, 3000, 5000, 4),
        ('Consumer Discretionary', 7000, 0.07, 3000, 2000, None, None, 2000, None, 5),
        ('Financials', 6000, 0.06, 6000, None, None, None, None, None, 2),
        ('Diversified/ETF', 5000, 0.05, None, 3000, None, 2000, None, None, 3),
        ('Communication Services', 3000, 0.03, 1500, None, None, 1500, None, None, 2),
        ('Utilities', 2000, 0.02, None, None, None, 2000, None, None, 1),
        ('Real Estate', 1000, 0.01, 500, None, None, 500, None, None, 3),
    ]

    sector_first = row
    for s in sectors:
        name, total, pct, fb, ri, k4, hsa, rh, angel, n = s
        cell(ws, row, 1, name, font=BLUE_FONT)
        cell(ws, row, 2, total, font=BLUE_FONT, fmt=DOLLAR)
        cell(ws, row, 3, pct, font=BLUE_FONT, fmt=PCT)
        for i, val in enumerate([fb, ri, k4, hsa, rh, angel], 4):
            if val is not None:
                cell(ws, row, i, val, font=BLUE_FONT, fmt=DOLLAR)
            else:
                ws.cell(row=row, column=i).border = THIN_BORDER
        cell(ws, row, 10, n, font=BLUE_FONT)
        row += 1
    sector_last = row - 1

    cell(ws, row, 1, 'TOTAL', font=BOLD_FONT)
    cell(ws, row, 2, f'=SUM(B{sector_first}:B{sector_last})', font=BOLD_FONT, fmt=DOLLAR)
    cell(ws, row, 3, 1, font=BOLD_FONT, fmt=PCT)
    cell(ws, row, 10, f'=SUM(J{sector_first}:J{sector_last})', font=BOLD_FONT)
    for col in range(4, 10):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1

    cell(ws, row, 1, 'Top sector: Defense Tech')
    cell(ws, row, 2, 0.20, fmt=PCT)
    row += 1
    cell(ws, row, 1, 'Top 3 sectors combined')
    cell(ws, row, 2, 0.50, fmt=PCT)
    row += 2

    # ==================================================================
    # SECTION 6: GEOGRAPHIC CONCENTRATION
    # ==================================================================
    ws.cell(row=row, column=1, value='GEOGRAPHIC CONCENTRATION').font = SECTION_FONT
    row += 1
    header_row(ws, row, ['Region', 'Value', '% of Portfolio'])
    row += 1

    geo = [
        ('United States', 80000, 0.80),
        ('International', 20000, 0.20),
        ('  — Diversified Intl', 8000, 0.08),
        ('  — Canada', 4000, 0.04),
        ('  — Japan (ETF)', 3000, 0.03),
        ('  — South Korea (ETF)', 2000, 0.02),
        ('  — United Kingdom', 1500, 0.015),
        ('  — Netherlands', 1000, 0.01),
        ('  — Taiwan', 500, 0.005),
    ]

    for name, val, pct in geo:
        cell(ws, row, 1, name, font=BLUE_FONT)
        cell(ws, row, 2, val, font=BLUE_FONT, fmt=DOLLAR)
        cell(ws, row, 3, pct, font=BLUE_FONT, fmt=PCT)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='Note: Robinhood holdings scaled to net portfolio value (margin adjustment).').font = NOTE_FONT
    row += 1
    ws.cell(row=row, column=1, value="Note: BR Intl Equity Index classified as 'Diversified Intl'.").font = NOTE_FONT
    row += 1
    ws.cell(row=row, column=1, value='Note: All angel investments are US-headquartered.').font = NOTE_FONT
    row += 2

    # ==================================================================
    # SECTION 7: RISK METRICS
    # ==================================================================
    ws.cell(row=row, column=1, value='RISK METRICS').font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value='Risk metrics require 12 months of monthly return data. Will be populated at year-end.').font = NOTE_FONT
    row += 1
    ws.cell(row=row, column=1, value='Scope: Fidelity Roth IRA + HSA + Robinhood (Fid Brokerage excluded — RSU distortion, 401k excluded — quarterly only).').font = NOTE_FONT
    row += 2

    # ==================================================================
    # SECTION 8: RETURN METRIC DEFINITIONS
    # ==================================================================
    ws.cell(row=row, column=1, value='RETURN METRIC DEFINITIONS').font = Font(
        name='Arial', size=12, bold=True,
    )
    row += 1

    glossary = [
        (
            'Time-Weighted Return',
            'Measures portfolio performance independent of cash flows (deposits/withdrawals). '
            'Calculated as the product of each period\'s growth factor minus 1: '
            '(1 + R\u2081) \u00d7 (1 + R\u2082) \u00d7 ... \u00d7 (1 + R\u2099) \u2212 1, '
            'where R\u2099 = (Ending + Withdrawals \u2212 Deposits) / Beginning \u2212 1. '
            'Best for comparing manager skill against benchmarks.',
        ),
        (
            'Money-Weighted Return',
            'Measures the internal rate of return (IRR) accounting for the timing and '
            'size of all cash flows. Solves for r in: '
            '\u03a3 CF\u209c / (1 + r)^t = 0. '
            'Reflects the investor\'s actual experience including deposit/withdrawal timing.',
        ),
        (
            'Cost Basis Return',
            'Unrealized gain or loss as a percentage of total cost basis: '
            'Cost Basis Return = (Market Value \u2212 Cost Basis) / Cost Basis. '
            'Shows how much current holdings have appreciated relative to what was paid.',
        ),
        (
            'Alpha',
            'Excess return over a benchmark (S&P 500 by default): '
            'Alpha = Account Time-Weighted Return \u2212 Benchmark Return. '
            'Positive alpha indicates outperformance; negative indicates underperformance.',
        ),
    ]

    label_font = Font(name='Arial', size=10, bold=True)
    body_font = Font(name='Arial', size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for label, definition in glossary:
        c = ws.cell(row=row, column=1, value=label)
        c.font = label_font
        c.border = border
        c = ws.cell(row=row, column=2, value=definition)
        c.font = body_font
        c.border = border
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 45
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row, column=1).alignment = Alignment(vertical='top')
        row += 1

    # ==================================================================
    # Define named ranges and disable gridlines
    # ==================================================================
    n = define_named_ranges(wb)
    for sheet_name in wb.sheetnames:
        wb[sheet_name].sheet_view.showGridLines = False

    wb.save(XLSX)
    print('Dashboard rebuilt successfully.')
    print('  Order: Daily Summary → Account Overview → Liquidity → Key Metrics → Benchmark → Sector → Geographic → Risk')

    # Validate after save
    from validate_workbook import validate_structural, format_findings
    findings = validate_structural(XLSX, "Dashboard")
    print(format_findings(findings))
    n_fail = sum(1 for f in findings if f.severity == "ERROR")
    if n_fail:
        print(f"  WARNING: {n_fail} validation error(s) detected!")


if __name__ == '__main__':
    main()
