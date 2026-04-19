#!/usr/bin/env python3
"""
build_portfolio.py — Portfolio Analysis Workbook Builder (2026)
Matches the 2025 workbook structure exactly per Portfolio Analysis Instructions.md.
Reads extraction JSONs from fidelity_csv.py and plaid_extract.py.
"""

import datetime
import json
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- File paths (defaults, overridable via build_workbook() or env vars) ---
FIDELITY_JSON = Path(os.environ.get("PORTFOLIO_FIDELITY_JSON",
    str(Path(__file__).parent / "extract_output" / "fidelity_pipeline_20260301_202139.json")))
ROBINHOOD_JSON = Path(os.environ.get("PORTFOLIO_ROBINHOOD_JSON",
    str(Path(__file__).parent / "extract_output" / "extract_raw_20260224_153857.json")))
OUTPUT = Path(os.environ.get("PORTFOLIO_OUTPUT",
    str(Path(__file__).parent / "2026_Portfolio_Analysis.xlsx")))

# --- Styles ---
BLUE = Font(name="Arial", color="0000FF", size=10)
BLACK = Font(name="Arial", size=10)
GREEN = Font(name="Arial", color="006100", size=10)
GRAY = Font(name="Arial", color="808080", size=10, italic=True)
RED = Font(name="Arial", color="FF0000", size=10)
BOLD = Font(name="Arial", bold=True, size=10)
BOLD_BLUE = Font(name="Arial", bold=True, color="0000FF", size=10)
BOLD_WHITE = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TITLE = Font(name="Arial", bold=True, size=14)
SECTION = Font(name="Arial", bold=True, size=12)

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")
DARK_FILL = PatternFill("solid", fgColor="1F3864")
YELLOW_FILL = PatternFill("solid", fgColor="FFF9C4")

C_CENTER = Alignment(horizontal="center", vertical="center")
C_RIGHT = Alignment(horizontal="right", vertical="center")
C_LEFT = Alignment(horizontal="left", vertical="center")

THIN = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

D_FMT = '$#,##0.00'
D_NEG = '$#,##0.00;($#,##0.00);"-"'
P_FMT = '0.0%'
N_FMT = '#,##0.000'
N2_FMT = '#,##0.00'

# --- Data ---
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Default data — override via manual_data.json or build_workbook(manual_json_path=...)
K401_DATA = {
    "quarterly": [
        {"period": "Q1 (Nov 1 - Jan 31)", "beginning": 0, "ee_contributions": 0,
         "er_contributions": 0, "fees": 0, "change_in_value": 0, "ending": 0},
    ],
    "holdings": [
        ("Example Fund", 0, 0, 0),
    ],
    "twr_provider_stated": 0,
}

CASH_BALANCES = {
    "fidelity_BROKERAGE": 0,
    "fidelity_ROTH_IRA": 0,
    "fidelity_HSA": 0,
}

ANGEL_DATA = [
    ("Example Company", "Technology", 2025, "Series A", 10000, 1e9, 1e9, "At cost"),
]

# Sector and geographic maps are loaded from each account's data/*.json file
# under the "sector_map" key. These are gitignored to avoid leaking portfolio
# composition. The maps are merged at runtime by portfolio_model.py.
SECTOR_MAP = {}
GEO_MAP = {}

def cl(n): return get_column_letter(n)

def hdr(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = C_CENTER; c.border = THIN
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[cl(i)].width = w

def brd(ws, r1, r2, c1, c2):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(r, c).border = THIN

def zb(ws, row, ncols):
    if row % 2 == 0:
        for c in range(1, ncols+1):
            if not ws.cell(row, c).fill or ws.cell(row, c).fill.fgColor.rgb == '00000000':
                ws.cell(row, c).fill = ZEBRA_FILL

def fmt_row(ws, row, ncols, font=BLUE, num_fmt=D_FMT, skip_first=True):
    for c in range(1, ncols+1):
        ws.cell(row, c).border = THIN
        if c == 1 and skip_first:
            ws.cell(row, c).alignment = C_LEFT
        else:
            ws.cell(row, c).alignment = C_RIGHT
            ws.cell(row, c).number_format = num_fmt
            if font: ws.cell(row, c).font = font


# ============================================================
# BROKERAGE / ROTH IRA TAB (build_monthly layout)
# Matches 2025: Month | Beginning | Deposits | Withdrawals | Dividends | Market Change | Ending | Monthly Return | Growth Factor
# ============================================================
def build_monthly_tab(wb, tab_name, label, holdings, cash, acct_type="standard"):
    ws = wb.create_sheet(tab_name)
    ws.cell(1, 1, f"{label} — 2026 Performance").font = TITLE
    ws.cell(2, 1, "Blue = hardcoded from statement | Black = formula | Green = cross-sheet ref").font = GRAY

    if acct_type == "fid_brokerage":
        # Fidelity Brokerage RSU layout: Month | Beginning | Subtractions | Change in Value | Ending | Notes
        headers = ["Month", "Beginning", "Subtractions", "Change in Value", "Ending", "Notes"]
        widths = [10, 16, 16, 16, 16, 30]
        hdr(ws, 3, headers, widths)
        for i in range(12):
            r = 4 + i
            ws.cell(r, 1, MONTHS[i]).font = BLUE
            for c in range(2, 6): ws.cell(r, c).font = BLUE; ws.cell(r, c).number_format = D_FMT
            ws.cell(r, 6).font = GRAY
            brd(ws, r, r, 1, 6); zb(ws, r, 6)

        r = 17
        ws.cell(r, 1, "Totals").font = BOLD
        ws.cell(r, 3).value = f"=SUM(C4:C15)"; ws.cell(r, 3).font = BOLD; ws.cell(r, 3).number_format = D_FMT
        ws.cell(r, 4).value = f"=SUM(D4:D15)"; ws.cell(r, 4).font = BOLD; ws.cell(r, 4).number_format = D_FMT
        brd(ws, r, r, 1, 6)

        ws.cell(19, 1, "Cost Basis Return").font = BLACK
        ws.cell(19, 2).font = BOLD_BLUE; ws.cell(19, 2).number_format = P_FMT
        ws.cell(19, 3, "(hardcoded — primary metric for RSU account)").font = GRAY
        ws.cell(20, 1, "YTD Change in Value").font = BLACK
        ws.cell(20, 2).value = "=D17"; ws.cell(20, 2).number_format = D_FMT
        ws.cell(21, 1, "YTD Subtractions").font = BLACK
        ws.cell(21, 2).value = "=C17"; ws.cell(21, 2).number_format = D_FMT

        ncols_h = 7
        h_start = 23
    elif acct_type == "fid_hsa":
        # Fidelity HSA layout: Month | Beginning | Additions | Subtractions | Change in Value | Ending | Monthly Return | Growth Factor
        headers = ["Month", "Beginning", "Additions", "Subtractions", "Change in Value", "Ending", "Monthly Return", "Growth Factor"]
        widths = [10, 16, 14, 14, 16, 16, 14, 14]
        hdr(ws, 3, headers, widths)
        for i in range(12):
            r = 4 + i
            ws.cell(r, 1, MONTHS[i]).font = BLUE
            for c in range(2, 7): ws.cell(r, c).font = BLUE; ws.cell(r, c).number_format = D_FMT
            # Monthly Return = (Ending + Subtractions - Additions) / Beginning - 1
            ws.cell(r, 7).value = f'=IF(B{r}=0,"",((F{r}+D{r}-C{r})/B{r})-1)'
            ws.cell(r, 7).number_format = P_FMT; ws.cell(r, 7).font = BLACK
            ws.cell(r, 8).value = f'=IF(G{r}="","",1+G{r})'
            ws.cell(r, 8).number_format = N2_FMT; ws.cell(r, 8).font = BLACK
            brd(ws, r, r, 1, 8); zb(ws, r, 8)

        r = 17
        ws.cell(r, 1, "Totals").font = BOLD
        for c in [3, 4, 5]: ws.cell(r, c).value = f"=SUM({cl(c)}4:{cl(c)}15)"; ws.cell(r, c).font = BOLD; ws.cell(r, c).number_format = D_FMT
        brd(ws, r, r, 1, 8)

        ws.cell(19, 1, "Time-Weighted Return (YTD)").font = BLACK
        ws.cell(19, 2).value = '=IFERROR(PRODUCT(H4:H15)-1,"")'; ws.cell(19, 2).number_format = P_FMT; ws.cell(19, 2).font = BOLD
        ws.cell(20, 1, "Annual Modified Dietz Return").font = BLACK
        ws.cell(20, 2).font = BOLD_BLUE; ws.cell(20, 2).number_format = P_FMT
        ws.cell(20, 3, "(computed from monthly cash flows)").font = GRAY
        ws.cell(21, 1, "Money-Weighted Return (YTD)").font = BLACK
        ws.cell(21, 2).font = BOLD_BLUE; ws.cell(21, 2).number_format = P_FMT
        ws.cell(21, 3, "(computed from monthly cash flows)").font = GRAY
        ws.cell(22, 1, "Total Contributions").font = BLACK
        ws.cell(22, 2).value = "=C17"; ws.cell(22, 2).number_format = D_FMT
        ws.cell(23, 1, "Total Distributions").font = BLACK
        ws.cell(23, 2).value = "=D17"; ws.cell(23, 2).number_format = D_FMT

        ncols_h = 7
        h_start = 25
    else:
        # Standard brokerage/Roth/Robinhood: Month | Beginning | Deposits | Withdrawals | Dividends | Market Change | Ending | Monthly Return | Growth Factor
        headers = ["Month", "Beginning Value", "Deposits", "Withdrawals", "Dividends", "Market Change", "Ending Value", "Monthly Return", "Growth Factor"]
        widths = [10, 16, 14, 14, 14, 16, 16, 14, 14]
        hdr(ws, 3, headers, widths)
        for i in range(12):
            r = 4 + i
            ws.cell(r, 1, MONTHS[i]).font = BLUE
            for c in range(2, 8): ws.cell(r, c).font = BLUE; ws.cell(r, c).number_format = D_FMT
            # Monthly Return = (Ending + Withdrawals - Deposits) / Beginning - 1
            ws.cell(r, 8).value = f'=IF(B{r}=0,"",((G{r}+D{r}-C{r})/B{r})-1)'
            ws.cell(r, 8).number_format = P_FMT; ws.cell(r, 8).font = BLACK
            ws.cell(r, 9).value = f'=IF(H{r}="","",1+H{r})'
            ws.cell(r, 9).number_format = N2_FMT; ws.cell(r, 9).font = BLACK
            brd(ws, r, r, 1, 9); zb(ws, r, 9)

        r = 17
        ws.cell(r, 1, "Totals").font = BOLD
        for c in [3, 4, 5, 6]:
            ws.cell(r, c).value = f"=SUM({cl(c)}4:{cl(c)}15)"; ws.cell(r, c).font = BOLD; ws.cell(r, c).number_format = D_FMT
        brd(ws, r, r, 1, 9)

        ws.cell(19, 1, "Time-Weighted Return (YTD)").font = BLACK
        ws.cell(19, 2).value = '=IFERROR(PRODUCT(I4:I15)-1,"")'; ws.cell(19, 2).number_format = P_FMT; ws.cell(19, 2).font = BOLD
        ws.cell(20, 1, "Money-Weighted Return (YTD)").font = BLACK
        ws.cell(20, 2).font = BOLD_BLUE; ws.cell(20, 2).number_format = P_FMT
        ws.cell(20, 3, "(computed from monthly cash flows)").font = GRAY
        ws.cell(21, 1, "Cost Basis Return").font = BLACK
        ws.cell(21, 2).font = BOLD_BLUE; ws.cell(21, 2).number_format = P_FMT
        ws.cell(21, 3, "(computed from monthly cash flows)").font = GRAY
        ws.cell(22, 1, "Total Dividends").font = BLACK
        ws.cell(22, 2).value = "=E17"; ws.cell(22, 2).number_format = D_FMT

        if acct_type == "robinhood":
            # Margin details section
            ws.cell(24, 1, "MARGIN ACCOUNT DETAILS").font = SECTION
            ws.cell(25, 1, "This is a margin account. Portfolio Value = Total Securities - Margin Debt.").font = GRAY
            ws.cell(26, 1, "Beginning Year Margin Debt").font = BLACK
            ws.cell(26, 2).font = BLUE; ws.cell(26, 2).number_format = D_FMT
            ws.cell(27, 1, "Ending Margin Debt").font = BLACK
            ws.cell(27, 2).font = BLUE; ws.cell(27, 2).number_format = D_FMT
            ws.cell(28, 1, "Annual Margin Interest (est.)").font = BLACK
            ws.cell(28, 2).font = BLUE; ws.cell(28, 2).number_format = D_FMT
            ws.cell(29, 1, "Market Change includes both stock appreciation and margin interest.").font = GRAY
            h_start = 31
        else:
            h_start = 24

        ncols_h = 7

    # --- Holdings section ---
    ws.cell(h_start, 1, "YEAR-END HOLDINGS" if acct_type != "robinhood" else "CURRENT HOLDINGS").font = SECTION
    h_row = h_start + 1

    if acct_type == "robinhood":
        # Robinhood: Security | Qty | Price | Market Value (no cost basis in 2025 layout)
        h_headers = ["Security", "Quantity", "Price", "Market Value"]
        h_widths = [20, 12, 14, 16]
        hdr(ws, h_row, h_headers)
        for i, w in enumerate(h_widths, 1): ws.column_dimensions[cl(i)].width = max(ws.column_dimensions[cl(i)].width, w)
    else:
        h_headers = ["Security", "Quantity", "Price", "Market Value", "Cost Basis", "Gain/Loss", "Return %"]
        hdr(ws, h_row, h_headers)

    h_row += 1
    first_h = h_row
    sorted_h = sorted(holdings.items(), key=lambda x: x[0])  # alphabetical per 2025 spec

    for ticker, h in sorted_h:
        ws.cell(h_row, 1, ticker).font = BLUE; ws.cell(h_row, 1).alignment = C_LEFT
        ws.cell(h_row, 2, h["qty"]).font = BLUE; ws.cell(h_row, 2).number_format = N_FMT; ws.cell(h_row, 2).alignment = C_RIGHT
        ws.cell(h_row, 3, h["price"]).font = BLUE; ws.cell(h_row, 3).number_format = D_FMT; ws.cell(h_row, 3).alignment = C_RIGHT
        ws.cell(h_row, 4, h["mv"]).font = BLUE; ws.cell(h_row, 4).number_format = D_FMT; ws.cell(h_row, 4).alignment = C_RIGHT
        if acct_type != "robinhood":
            ws.cell(h_row, 5, h["cb"]).font = BLUE; ws.cell(h_row, 5).number_format = D_FMT; ws.cell(h_row, 5).alignment = C_RIGHT
            ws.cell(h_row, 6).value = f"=D{h_row}-E{h_row}"; ws.cell(h_row, 6).number_format = D_NEG; ws.cell(h_row, 6).alignment = C_RIGHT
            ws.cell(h_row, 7).value = f'=IF(E{h_row}=0,"N/A",F{h_row}/E{h_row})'; ws.cell(h_row, 7).number_format = P_FMT; ws.cell(h_row, 7).alignment = C_RIGHT

        nc = 4 if acct_type == "robinhood" else 7
        brd(ws, h_row, h_row, 1, nc); zb(ws, h_row, nc)
        h_row += 1

    # Cash row
    if cash > 0:
        ws.cell(h_row, 1, "Cash").font = BLUE
        ws.cell(h_row, 4, cash).font = BLUE; ws.cell(h_row, 4).number_format = D_FMT
        nc = 4 if acct_type == "robinhood" else 7
        brd(ws, h_row, h_row, 1, nc)
        h_row += 1

    # TOTAL row
    nc = 4 if acct_type == "robinhood" else 7
    total_row = h_row
    ws.cell(h_row, 1, "TOTAL" if acct_type != "robinhood" else "TOTAL SECURITIES").font = BOLD
    ws.cell(h_row, 4).value = f"=SUM(D{first_h}:D{h_row-1})"; ws.cell(h_row, 4).font = BOLD; ws.cell(h_row, 4).number_format = D_FMT
    if acct_type != "robinhood":
        ws.cell(h_row, 5).value = f"=SUM(E{first_h}:E{h_row-1})"; ws.cell(h_row, 5).font = BOLD; ws.cell(h_row, 5).number_format = D_FMT
        ws.cell(h_row, 6).value = f"=SUM(F{first_h}:F{h_row-1})"; ws.cell(h_row, 6).font = BOLD; ws.cell(h_row, 6).number_format = D_NEG
        ws.cell(h_row, 7).value = f'=IF(E{h_row}=0,0,F{h_row}/E{h_row})'; ws.cell(h_row, 7).font = BOLD; ws.cell(h_row, 7).number_format = P_FMT
    brd(ws, h_row, h_row, 1, nc)

    if acct_type == "robinhood":
        # Margin Debt row
        h_row += 1
        rh_raw = json.loads(ROBINHOOD_JSON.read_text())
        margin_debt = rh_raw["robinhood"]["accounts"][0]["balances"]["current"]
        ws.cell(h_row, 1, "Margin Debt").font = RED
        ws.cell(h_row, 4, margin_debt).font = RED; ws.cell(h_row, 4).number_format = D_FMT
        brd(ws, h_row, h_row, 1, 4)
        debt_row = h_row

        h_row += 1
        ws.cell(h_row, 1, "NET PORTFOLIO VALUE").font = BOLD
        ws.cell(h_row, 4).value = f"=D{total_row}+D{debt_row}"; ws.cell(h_row, 4).font = BOLD; ws.cell(h_row, 4).number_format = D_FMT
        brd(ws, h_row, h_row, 1, 4)

    # Investment Gain Summary
    h_row += 2
    ws.cell(h_row, 1, "INVESTMENT GAIN SUMMARY").font = SECTION
    h_row += 1
    if acct_type == "robinhood":
        ws.cell(h_row, 1, "Market Value Change").font = BLACK
        ws.cell(h_row, 2).value = "=F17"; ws.cell(h_row, 2).font = GREEN; ws.cell(h_row, 2).number_format = D_FMT
        h_row += 1
        ws.cell(h_row, 1, "Dividends Received").font = BLACK
        ws.cell(h_row, 2).value = "=E17"; ws.cell(h_row, 2).font = GREEN; ws.cell(h_row, 2).number_format = D_FMT
        h_row += 1
        ws.cell(h_row, 1, "Total Investment Gain").font = BOLD
        ws.cell(h_row, 2).value = f"=B{h_row-2}+B{h_row-1}"; ws.cell(h_row, 2).font = BOLD; ws.cell(h_row, 2).number_format = D_FMT
    elif acct_type == "fid_brokerage":
        ws.cell(h_row, 1, "Unrealized Gain/Loss").font = BLACK
        ws.cell(h_row, 2).value = f"=F{total_row}"; ws.cell(h_row, 2).font = GREEN; ws.cell(h_row, 2).number_format = D_FMT
        h_row += 1
        ws.cell(h_row, 1, "Realized G/L (incl RSU)").font = BLACK
        ws.cell(h_row, 2).value = f"=D17-B{h_row-1}"; ws.cell(h_row, 2).font = BLACK; ws.cell(h_row, 2).number_format = D_FMT
        h_row += 1
        ws.cell(h_row, 1, "Total Investment Gain").font = BOLD
        ws.cell(h_row, 2).value = f"=B{h_row-2}+B{h_row-1}"; ws.cell(h_row, 2).font = BOLD; ws.cell(h_row, 2).number_format = D_FMT
    elif acct_type == "fid_hsa":
        ws.cell(h_row, 1, "Unrealized Gain/Loss").font = BLACK
        ws.cell(h_row, 2).value = f"=F{total_row}"; ws.cell(h_row, 2).font = GREEN; ws.cell(h_row, 2).number_format = D_FMT
        h_row += 1
        ws.cell(h_row, 1, "Dividends/Income").font = BLACK
        ws.cell(h_row, 2).value = f"=E17-B{h_row-1}"; ws.cell(h_row, 2).font = BLACK; ws.cell(h_row, 2).number_format = D_FMT
        h_row += 1
        ws.cell(h_row, 1, "Total Investment Gain").font = BOLD
        ws.cell(h_row, 2).value = f"=B{h_row-2}+B{h_row-1}"; ws.cell(h_row, 2).font = BOLD; ws.cell(h_row, 2).number_format = D_FMT
    else:
        # Standard brokerage/roth
        ws.cell(h_row, 1, "Unrealized Gain/Loss").font = BLACK
        ws.cell(h_row, 2).value = f"=F{total_row}"; ws.cell(h_row, 2).font = GREEN; ws.cell(h_row, 2).number_format = D_FMT
        h_row += 1
        ws.cell(h_row, 1, "Dividends Received").font = BLACK
        ws.cell(h_row, 2).value = "=E17"; ws.cell(h_row, 2).font = GREEN; ws.cell(h_row, 2).number_format = D_FMT
        h_row += 1
        ws.cell(h_row, 1, "Realized Gain/Loss").font = BLACK
        ws.cell(h_row, 2).value = f"=F17-B{h_row-2}-B{h_row-1}"; ws.cell(h_row, 2).font = BLACK; ws.cell(h_row, 2).number_format = D_FMT
        h_row += 1
        ws.cell(h_row, 1, "Total Investment Gain").font = BOLD
        ws.cell(h_row, 2).value = f"=B{h_row-3}+B{h_row-2}+B{h_row-1}"; ws.cell(h_row, 2).font = BOLD; ws.cell(h_row, 2).number_format = D_FMT

    return ws, total_row


# ============================================================
# 401(k) TAB
# ============================================================
def build_401k_tab(wb, live_merrill=None):
    """Build 401(k) tab. If live_merrill dict is provided, use it for current holdings."""
    ws = wb.create_sheet("401(k)")
    fiscal_year = K401_DATA.get("account", {}).get("fiscal_year", "Nov-Oct")
    provider_name = K401_DATA.get("account", {}).get("provider_name", "401(k) Provider")
    ws.cell(1, 1, f"401(k) — {provider_name} — {datetime.date.today().year} Performance (FY {fiscal_year})").font = TITLE
    ws.cell(2, 1, "Blue = hardcoded from statement | Black = formula | Green = cross-sheet ref").font = GRAY

    headers = ["Period", "Beginning", "EE Contrib", "ER Contrib", "Fees", "Change in Value", "Ending", "Total Contrib", "Mod. Dietz Return", "Growth Factor"]
    widths = [22, 14, 14, 14, 12, 16, 14, 14, 16, 14]
    hdr(ws, 4, headers, widths)

    # Q1 data
    q = K401_DATA["quarterly"][0]
    r = 5
    ws.cell(r, 1, q["period"]).font = BLUE
    ws.cell(r, 2, q["beginning"]).font = BLUE; ws.cell(r, 2).number_format = D_FMT
    ws.cell(r, 3, q["ee_contributions"]).font = BLUE; ws.cell(r, 3).number_format = D_FMT
    ws.cell(r, 4, q["er_contributions"]).font = BLUE; ws.cell(r, 4).number_format = D_FMT
    ws.cell(r, 5, q["fees"]).font = BLUE; ws.cell(r, 5).number_format = D_FMT
    ws.cell(r, 6, q["change_in_value"]).font = BLUE; ws.cell(r, 6).number_format = D_FMT
    ws.cell(r, 7, q["ending"]).font = BLUE; ws.cell(r, 7).number_format = D_FMT
    ws.cell(r, 8).value = f"=C{r}+D{r}"; ws.cell(r, 8).number_format = D_FMT
    # Modified Dietz = Change / (Beginning + 0.5*(EE+ER-Fees))
    ws.cell(r, 9).value = f"=F{r}/(B{r}+0.5*(C{r}+D{r}+E{r}))"; ws.cell(r, 9).number_format = P_FMT
    ws.cell(r, 10).value = f"=1+I{r}"; ws.cell(r, 10).number_format = N2_FMT
    brd(ws, r, r, 1, 10)

    # Q2-Q4 placeholder rows
    for qi in range(1, 4):
        r = 6 + qi - 1
        qnames = ["Q2 (Feb 1 - Apr 30)", "Q3 (May 1 - Jul 31)", "Q4 (Aug 1 - Oct 31)"]
        ws.cell(r, 1, qnames[qi-1]).font = BLUE
        for c in range(2, 8): ws.cell(r, c).font = BLUE; ws.cell(r, c).number_format = D_FMT
        ws.cell(r, 8).value = f"=C{r}+D{r}"; ws.cell(r, 8).number_format = D_FMT
        ws.cell(r, 9).value = f'=IF(B{r}=0,"",F{r}/(B{r}+0.5*(C{r}+D{r}+E{r})))'; ws.cell(r, 9).number_format = P_FMT
        ws.cell(r, 10).value = f'=IF(I{r}="","",1+I{r})'; ws.cell(r, 10).number_format = N2_FMT
        brd(ws, r, r, 1, 10); zb(ws, r, 10)

    # YTD Totals (row 10)
    r = 10
    ws.cell(r, 1, "YTD Totals").font = BOLD
    for c in [3, 4, 5, 6]:
        ws.cell(r, c).value = f"=SUM({cl(c)}5:{cl(c)}8)"; ws.cell(r, c).font = BOLD; ws.cell(r, c).number_format = D_FMT
    ws.cell(r, 8).value = "=SUM(H5:H8)"; ws.cell(r, 8).font = BOLD; ws.cell(r, 8).number_format = D_FMT
    brd(ws, r, r, 1, 10)

    # Verify ending (row 11)
    ws.cell(11, 1, "Verify Ending").font = BLACK
    ws.cell(11, 7).value = "=B5+C10+D10+E10+F10"; ws.cell(11, 7).number_format = D_FMT

    # TWR computed (row 13)
    ws.cell(13, 1, "Time-Weighted Return (Computed — Modified Dietz)").font = BLACK
    ws.cell(13, 2).value = '=IFERROR(PRODUCT(J5:J8)-1,"")'; ws.cell(13, 2).number_format = P_FMT; ws.cell(13, 2).font = BOLD

    # TWR provider stated (row 14)
    provider_name_short = K401_DATA.get("account", {}).get("provider_name", "Provider")
    ws.cell(14, 1, f"Time-Weighted Return ({provider_name_short} Stated)").font = BLACK
    ws.cell(14, 2, K401_DATA.get("twr_provider_stated", K401_DATA.get("twr_merrill_stated", 0))).font = BOLD_BLUE; ws.cell(14, 2).number_format = P_FMT
    ws.cell(14, 2).fill = YELLOW_FILL

    # MWRR (row 15)
    ws.cell(15, 1, "Money-Weighted Return (YTD)").font = BLACK
    ws.cell(15, 2).font = BOLD_BLUE; ws.cell(15, 2).number_format = P_FMT
    ws.cell(15, 3, "(computed from monthly cash flows)").font = GRAY

    # Cost Basis Return (row 16)
    ws.cell(16, 1, "Cost Basis Return").font = BLACK
    ws.cell(16, 2).value = "=IF(B5=0,0,(F10+E10)/(B5+0.5*(C10+D10)))"; ws.cell(16, 2).number_format = P_FMT

    # Fund Holdings (row 18+)
    ws.cell(18, 1, "FUND HOLDINGS").font = SECTION

    if live_merrill:
        # Live data from Plaid: Fund Name | Current Value | Cost Basis | Unrealized G/L | Return %
        fund_headers = ["Fund Name", "Current Value", "Cost Basis", "Unrealized Gain/Loss", "Return %"]
        hdr(ws, 19, fund_headers, [34, 18, 18, 18, 12])

        # Build security lookup
        sec_map = {}
        for s in live_merrill.get("securities", []):
            sec_map[s["security_id"]] = s.get("name") or s.get("ticker_symbol") or "Unknown"

        r = 20
        first_fund = r
        for h in sorted(live_merrill.get("holdings", []),
                        key=lambda x: x.get("institution_value", 0) or 0, reverse=True):
            name = sec_map.get(h.get("security_id", ""), "Unknown Fund")
            val = h.get("institution_value") or 0
            cb = h.get("cost_basis") or 0
            if val < 1:
                continue
            ws.cell(r, 1, name).font = BLUE; ws.cell(r, 1).alignment = C_LEFT
            ws.cell(r, 2, round(val, 2)).font = BLUE; ws.cell(r, 2).number_format = D_FMT
            ws.cell(r, 3, round(cb, 2)).font = BLUE; ws.cell(r, 3).number_format = D_FMT
            ws.cell(r, 4).value = f"=B{r}-C{r}"; ws.cell(r, 4).number_format = D_NEG
            ws.cell(r, 5).value = f'=IF(C{r}=0,0,D{r}/C{r})'; ws.cell(r, 5).number_format = P_FMT
            brd(ws, r, r, 1, 5); zb(ws, r, 5)
            r += 1
    else:
        # Manual data: Fund Name | Beginning Balance | Ending Balance | Investment Gain/Loss | Return %
        fund_headers = ["Fund Name", "Beginning Balance", "Ending Balance", "Investment Gain/Loss", "Return %"]
        hdr(ws, 19, fund_headers, [34, 18, 18, 18, 12])

        r = 20
        first_fund = r
        for name, beg, end, gain in K401_DATA["holdings"]:
            ws.cell(r, 1, name).font = BLUE; ws.cell(r, 1).alignment = C_LEFT
            ws.cell(r, 2, beg).font = BLUE; ws.cell(r, 2).number_format = D_FMT
            ws.cell(r, 3, end).font = BLUE; ws.cell(r, 3).number_format = D_FMT
            ws.cell(r, 4, gain).font = BLUE; ws.cell(r, 4).number_format = D_FMT
            ws.cell(r, 5).value = f'=IF(B{r}=0,0,D{r}/B{r})'; ws.cell(r, 5).number_format = P_FMT
            brd(ws, r, r, 1, 5); zb(ws, r, 5)
            r += 1

    # TOTAL row
    last_fund = r - 1
    ws.cell(r, 1, "TOTAL").font = BOLD
    ws.cell(r, 2).value = f"=SUM(B{first_fund}:B{last_fund})"; ws.cell(r, 2).font = BOLD; ws.cell(r, 2).number_format = D_FMT
    ws.cell(r, 3).value = f"=SUM(C{first_fund}:C{last_fund})"; ws.cell(r, 3).font = BOLD; ws.cell(r, 3).number_format = D_FMT
    ws.cell(r, 4).value = f"=SUM(D{first_fund}:D{last_fund})"; ws.cell(r, 4).font = BOLD; ws.cell(r, 4).number_format = D_FMT
    ws.cell(r, 5).value = f'=IF(B{r}=0,0,D{r}/B{r})'; ws.cell(r, 5).font = BOLD; ws.cell(r, 5).number_format = P_FMT
    brd(ws, r, r, 1, 5)
    fund_total_row = r

    # Investment Gain Summary
    r += 2
    ws.cell(r, 1, "INVESTMENT GAIN SUMMARY").font = SECTION
    r += 1
    ws.cell(r, 1, "Change in Investment Value").font = BLACK
    ws.cell(r, 2).value = "=F10"; ws.cell(r, 2).font = GREEN; ws.cell(r, 2).number_format = D_FMT
    r += 1
    ws.cell(r, 1, "Fees").font = BLACK
    ws.cell(r, 2).value = "=E10"; ws.cell(r, 2).font = GREEN; ws.cell(r, 2).number_format = D_FMT
    r += 1
    ws.cell(r, 1, "Total Investment Gain").font = BOLD
    ws.cell(r, 2).value = f"=B{r-2}+B{r-1}"; ws.cell(r, 2).font = BOLD; ws.cell(r, 2).number_format = D_FMT

    return ws, fund_total_row


# ============================================================
# ANGEL INVESTMENTS TAB
# ============================================================
def build_angel_tab(wb):
    ws = wb.create_sheet("Angel Investments")
    ws.cell(1, 1, "Angel Investments — Portfolio Summary").font = TITLE
    ws.cell(2, 1, "Blue = input data | Black = formula | Valuations as of Mar 2026").font = GRAY

    headers = ["Company", "Sector", "Year", "Series", "Amount Invested",
               "Post-Money at Investment", "Latest Valuation", "Valuation Source",
               "Est. Current Value", "Return %"]
    widths = [16, 14, 8, 12, 18, 22, 22, 22, 18, 12]
    hdr(ws, 4, headers, widths)

    r = 5
    first = r
    for co, sector, year, series, amt, pm_inv, pm_latest, source in ANGEL_DATA:
        ws.cell(r, 1, co).font = BLUE
        ws.cell(r, 2, sector).font = BLUE
        ws.cell(r, 3, year).font = BLUE
        ws.cell(r, 4, series).font = BLUE
        ws.cell(r, 5, amt).font = BLUE; ws.cell(r, 5).number_format = D_FMT
        ws.cell(r, 6, pm_inv).font = BLUE; ws.cell(r, 6).number_format = '$#,##0'
        ws.cell(r, 7, pm_latest).font = BLUE; ws.cell(r, 7).number_format = '$#,##0'
        ws.cell(r, 8, source).font = BLUE
        ws.cell(r, 9).value = f"=E{r}*(G{r}/F{r})"; ws.cell(r, 9).number_format = D_FMT
        ws.cell(r, 10).value = f"=(I{r}-E{r})/E{r}"; ws.cell(r, 10).number_format = P_FMT
        for c in range(1, 11):
            ws.cell(r, c).border = THIN
            ws.cell(r, c).alignment = C_RIGHT if c >= 5 else C_LEFT
        zb(ws, r, 10)
        r += 1

    last = r - 1
    ws.cell(r, 1, "TOTAL").font = BOLD
    ws.cell(r, 5).value = f"=SUM(E{first}:E{last})"; ws.cell(r, 5).font = BOLD; ws.cell(r, 5).number_format = D_FMT
    ws.cell(r, 9).value = f"=SUM(I{first}:I{last})"; ws.cell(r, 9).font = BOLD; ws.cell(r, 9).number_format = D_FMT
    ws.cell(r, 10).value = f"=(I{r}-E{r})/E{r}"; ws.cell(r, 10).font = BOLD; ws.cell(r, 10).number_format = P_FMT
    brd(ws, r, r, 1, 10)
    angel_total_row = r

    # Notes
    r += 2
    ws.cell(r, 1, "Valuation Methodology:").font = BOLD
    r += 1
    ws.cell(r, 1, "Est. Current Value = Amount Invested × (Latest Valuation / Post-Money at Investment)").font = GRAY
    r += 1
    ws.cell(r, 1, "Investments at cost: no subsequent priced round; pm_latest = pm_invest.").font = GRAY

    return ws, angel_total_row


# ============================================================
# CASH TAB
# ============================================================
def build_cash_tab(wb, cash_current, cash_history):
    """Build the Cash tab showing current balances and monthly history.

    cash_current: dict from extract_plaid_cash, e.g.
        {"chase": {"accounts": [...], "total": 15234.56}, "marcus": {...}}
    cash_history: list of snapshots, e.g.
        [{"date": "2026-04-04", "chase": 15234.56, "marcus": 42000.00, "total": 57234.56}]

    Returns (ws, total_row, first_data_row) where total_row is the TOTAL CASH row number.
    """
    ws = wb.create_sheet("Cash")

    ws.cell(1, 1, "Cash Accounts").font = TITLE
    ws.cell(2, 1, "Blue = hardcoded from Plaid | Black = formula").font = GRAY

    # --- Current Balances ---
    r = 4
    ws.cell(r, 1, "CURRENT BALANCES").font = SECTION
    r += 1
    hdr(ws, r, ["Account", "Institution", "Balance"], [30, 24, 16])
    r += 1

    acct_start = r
    inst_labels = {
        "chase": "Chase",
        "marcus": "Marcus (Goldman Sachs)",
    }

    if cash_current:
        for inst_key in sorted(cash_current.keys()):
            inst_data = cash_current[inst_key]
            inst_name = inst_labels.get(inst_key, inst_key)
            for acct in inst_data["accounts"]:
                ws.cell(r, 1, acct["name"]).font = BLUE
                ws.cell(r, 2, inst_name).font = BLUE
                ws.cell(r, 3, acct["balance"]).font = BLUE
                ws.cell(r, 3).number_format = D_FMT
                brd(ws, r, r, 1, 3); zb(ws, r, 3)
                r += 1

    # TOTAL row
    total_row = r
    ws.cell(r, 1, "TOTAL CASH").font = BOLD
    ws.cell(r, 2, "").font = BOLD
    if r > acct_start:
        ws.cell(r, 3).value = f"=SUM(C{acct_start}:C{r-1})"
    else:
        ws.cell(r, 3, 0)
    ws.cell(r, 3).font = BOLD
    ws.cell(r, 3).number_format = D_FMT
    brd(ws, r, r, 1, 3)

    # --- Monthly Balance History ---
    r += 2
    ws.cell(r, 1, "MONTHLY BALANCE HISTORY").font = SECTION
    r += 1

    # Determine which institutions appear in history
    inst_keys = sorted(set(
        k for entry in (cash_history or [])
        for k in entry.keys()
        if k not in ("date", "total")
    ))
    if not inst_keys and cash_current:
        inst_keys = sorted(cash_current.keys())

    hist_headers = ["Month"] + [inst_labels.get(k, k) for k in inst_keys] + ["Total"]
    hist_widths = [18] + [16] * len(inst_keys) + [16]
    hdr(ws, r, hist_headers, hist_widths)
    r += 1

    # Build monthly lookup: for each month, find the latest snapshot
    monthly = {}
    for entry in (cash_history or []):
        month_key = entry["date"][:7]  # "2026-04"
        if month_key not in monthly or entry["date"] > monthly[month_key]["date"]:
            monthly[month_key] = entry

    import datetime as _dt
    year = _dt.date.today().year
    history_start_row = r
    first_data_row = None

    for m in range(1, 13):
        month_key = f"{year}-{m:02d}"
        month_name = _dt.date(year, m, 1).strftime("%B %Y")
        ws.cell(r, 1, month_name).font = BLACK

        entry = monthly.get(month_key)
        if entry:
            if first_data_row is None:
                first_data_row = r
            for ci, inst_key in enumerate(inst_keys, 2):
                val = entry.get(inst_key, 0)
                ws.cell(r, ci, val).font = BLUE
                ws.cell(r, ci).number_format = D_FMT
            # Total column = sum of institution columns
            total_col = len(inst_keys) + 2
            start_col_letter = cl(2)
            end_col_letter = cl(total_col - 1)
            ws.cell(r, total_col).value = f"=SUM({start_col_letter}{r}:{end_col_letter}{r})"
            ws.cell(r, total_col).font = BLACK
            ws.cell(r, total_col).number_format = D_FMT
        else:
            for ci in range(2, len(inst_keys) + 3):
                ws.cell(r, ci, "--").font = GRAY
                ws.cell(r, ci).alignment = C_CENTER

        brd(ws, r, r, 1, len(inst_keys) + 2)
        zb(ws, r, len(inst_keys) + 2)
        r += 1

    return ws, total_row, first_data_row


# ============================================================
# DASHBOARD
# ============================================================
def build_dashboard(wb, acct_info, all_holdings, rh_raw, benchmarks=None):
    """benchmarks: dict like {"S&P 500": 0.05, "Dow Jones": 0.03, "NASDAQ": 0.08} or None"""
    ws = wb.create_sheet("Dashboard")
    wb.move_sheet("Dashboard", offset=-10)

    import datetime as _dt
    _today = _dt.date.today().strftime("%B %-d, %Y") if os.name != "nt" else _dt.date.today().strftime("%B %#d, %Y")
    ws.cell(1, 1, f"Portfolio Analysis — 2026 YTD (as of {_today})").font = TITLE
    ws.cell(2, 1, "Green = cross-sheet reference | Black = formula | Blue = hardcoded").font = GRAY

    # --- Section B: Account Overview (rows 4-14) ---
    ws.cell(4, 1, "ACCOUNT OVERVIEW").font = SECTION
    acct_headers = ["Account", "Beginning", "Ending", "Net Cash Flow", "Time-Weighted Return (YTD)", "Money-Weighted Return (YTD)", "Cost Basis Return", "Alpha"]
    widths = [22, 14, 14, 14, 10, 10, 10, 10]
    hdr(ws, 5, acct_headers, widths)
    ws.column_dimensions["A"].width = 24

    # Row ordering per 2025: Fid Brok, Roth IRA, 401k, HSA, Angel, Robinhood, Cash, then Total
    acct_rows = [
        ("Fidelity Brokerage", "'Fidelity Brokerage'", "fid_brokerage"),
        ("Fidelity Roth IRA", "'Fidelity Roth IRA'", "standard"),
        ("401(k)", "'401(k)'", "401k"),
        ("Fidelity HSA", "'Fidelity HSA'", "fid_hsa"),
        ("Angel Investments", "'Angel Investments'", "angel"),
        ("Robinhood", "'Robinhood'", "robinhood"),
        ("Cash", "'Cash'", "cash"),
    ]

    r = 6
    first_r = r
    for name, sref, atype in acct_rows:
        ws.cell(r, 1, name).font = BLACK

        if atype == "401k":
            ws.cell(r, 2).value = f"={sref}!B5"; ws.cell(r, 2).font = GREEN; ws.cell(r, 2).number_format = D_FMT
            ws.cell(r, 3).value = f"={sref}!G8"; ws.cell(r, 3).font = GREEN; ws.cell(r, 3).number_format = D_FMT
            ws.cell(r, 4).value = f"={sref}!H10"; ws.cell(r, 4).font = GREEN; ws.cell(r, 4).number_format = D_FMT
            ws.cell(r, 5).value = f"={sref}!B13"; ws.cell(r, 5).font = GREEN; ws.cell(r, 5).number_format = P_FMT
            ws.cell(r, 6).value = f"={sref}!B15"; ws.cell(r, 6).font = GREEN; ws.cell(r, 6).number_format = P_FMT
            ws.cell(r, 7).value = f"={sref}!B16"; ws.cell(r, 7).font = GREEN; ws.cell(r, 7).number_format = P_FMT
        elif atype == "angel":
            angel_tot = acct_info["angel_total_row"]
            ws.cell(r, 2).value = f"={sref}!E{angel_tot}"; ws.cell(r, 2).font = GREEN; ws.cell(r, 2).number_format = D_FMT
            ws.cell(r, 3).value = f"={sref}!I{angel_tot}"; ws.cell(r, 3).font = GREEN; ws.cell(r, 3).number_format = D_FMT
            ws.cell(r, 4, "N/A").font = GRAY
            ws.cell(r, 5, "N/A").font = GRAY
            ws.cell(r, 6, "N/A").font = GRAY
            ws.cell(r, 7).value = f"={sref}!J{angel_tot}"; ws.cell(r, 7).font = GREEN; ws.cell(r, 7).number_format = P_FMT
        elif atype == "fid_brokerage":
            ws.cell(r, 2).value = f"={sref}!B4"; ws.cell(r, 2).font = GREEN; ws.cell(r, 2).number_format = D_FMT
            ws.cell(r, 3).value = f"={sref}!E15"; ws.cell(r, 3).font = GREEN; ws.cell(r, 3).number_format = D_FMT
            ws.cell(r, 4).value = f"=C{r}-B{r}-{sref}!D17"; ws.cell(r, 4).font = BLACK; ws.cell(r, 4).number_format = D_FMT
            ws.cell(r, 5, "N/A").font = GRAY
            ws.cell(r, 6, "N/A").font = GRAY
            ws.cell(r, 7).value = f"={sref}!B19"; ws.cell(r, 7).font = GREEN; ws.cell(r, 7).number_format = P_FMT
        elif atype == "fid_hsa":
            ws.cell(r, 2).value = f"={sref}!B4"; ws.cell(r, 2).font = GREEN; ws.cell(r, 2).number_format = D_FMT
            ws.cell(r, 3).value = f"={sref}!F15"; ws.cell(r, 3).font = GREEN; ws.cell(r, 3).number_format = D_FMT
            ws.cell(r, 4).value = f"={sref}!C17-{sref}!D17"; ws.cell(r, 4).font = BLACK; ws.cell(r, 4).number_format = D_FMT
            ws.cell(r, 5).value = f"={sref}!B19"; ws.cell(r, 5).font = GREEN; ws.cell(r, 5).number_format = P_FMT
            ws.cell(r, 6).value = f"={sref}!B21"; ws.cell(r, 6).font = GREEN; ws.cell(r, 6).number_format = P_FMT
            ws.cell(r, 7, "N/A").font = GRAY
        elif atype == "robinhood":
            ws.cell(r, 2).value = f"={sref}!B4"; ws.cell(r, 2).font = GREEN; ws.cell(r, 2).number_format = D_FMT
            ws.cell(r, 3).value = f"={sref}!G15"; ws.cell(r, 3).font = GREEN; ws.cell(r, 3).number_format = D_FMT
            ws.cell(r, 4).value = f"={sref}!C17-{sref}!D17"; ws.cell(r, 4).font = BLACK; ws.cell(r, 4).number_format = D_FMT
            ws.cell(r, 5).value = f"={sref}!B19"; ws.cell(r, 5).font = GREEN; ws.cell(r, 5).number_format = P_FMT
            ws.cell(r, 6).value = f"={sref}!B20"; ws.cell(r, 6).font = GREEN; ws.cell(r, 6).number_format = P_FMT
            ws.cell(r, 7).value = f"={sref}!B21"; ws.cell(r, 7).font = GREEN; ws.cell(r, 7).number_format = P_FMT
        elif atype == "cash":
            cash_total_row = acct_info.get("cash_total_row")
            cash_first_data_row = acct_info.get("cash_first_data_row")
            if cash_total_row:
                ws.cell(r, 3).value = f"='Cash'!C{cash_total_row}"
                ws.cell(r, 3).font = GREEN; ws.cell(r, 3).number_format = D_FMT
            else:
                ws.cell(r, 3, 0).font = GRAY; ws.cell(r, 3).number_format = D_FMT
            # Beginning = earliest month's total from history tab
            if cash_first_data_row:
                total_col_letter = acct_info.get("cash_total_col_letter", "D")
                ws.cell(r, 2).value = f"='Cash'!{total_col_letter}{cash_first_data_row}"
                ws.cell(r, 2).font = GREEN; ws.cell(r, 2).number_format = D_FMT
            else:
                ws.cell(r, 2).value = ws.cell(r, 3).value  # same as ending if no history
                ws.cell(r, 2).font = GREEN; ws.cell(r, 2).number_format = D_FMT
            ws.cell(r, 4, "N/A").font = GRAY
            ws.cell(r, 5, "N/A").font = GRAY
            ws.cell(r, 6, "N/A").font = GRAY
            ws.cell(r, 7, "N/A").font = GRAY
        else:  # standard brokerage/roth
            ws.cell(r, 2).value = f"={sref}!B4"; ws.cell(r, 2).font = GREEN; ws.cell(r, 2).number_format = D_FMT
            ws.cell(r, 3).value = f"={sref}!G15"; ws.cell(r, 3).font = GREEN; ws.cell(r, 3).number_format = D_FMT
            ws.cell(r, 4).value = f"={sref}!C17-{sref}!D17"; ws.cell(r, 4).font = BLACK; ws.cell(r, 4).number_format = D_FMT
            ws.cell(r, 5).value = f"={sref}!B19"; ws.cell(r, 5).font = GREEN; ws.cell(r, 5).number_format = P_FMT
            ws.cell(r, 6).value = f"={sref}!B20"; ws.cell(r, 6).font = GREEN; ws.cell(r, 6).number_format = P_FMT
            ws.cell(r, 7).value = f"={sref}!B21"; ws.cell(r, 7).font = GREEN; ws.cell(r, 7).number_format = P_FMT

        # Alpha = TWR - benchmark (S&P 500) — row 8 col will be set after benchmarks
        ws.cell(r, 8).font = BLACK; ws.cell(r, 8).number_format = P_FMT
        brd(ws, r, r, 1, 8); zb(ws, r, 8)
        r += 1

    last_r = r - 1

    # Combined total (row 14 equiv → r)
    r += 1  # skip a row
    total_r = r
    ws.cell(r, 1, "TOTAL PORTFOLIO").font = BOLD_WHITE; ws.cell(r, 1).fill = DARK_FILL
    for c in range(2, 9):
        ws.cell(r, c).fill = DARK_FILL; ws.cell(r, c).font = BOLD_WHITE; ws.cell(r, c).border = THIN
    ws.cell(r, 2).value = f"=SUM(B{first_r}:B{last_r})"; ws.cell(r, 2).number_format = D_FMT
    ws.cell(r, 3).value = f"=SUM(C{first_r}:C{last_r})"; ws.cell(r, 3).number_format = D_FMT
    ws.cell(r, 4).value = f"=SUM(D{first_r}:D{last_r})"; ws.cell(r, 4).number_format = D_FMT
    ws.cell(r, 1).border = THIN

    # --- Section C: Benchmark Comparison ---
    r += 2
    ws.cell(r, 1, "YTD BENCHMARK COMPARISON").font = SECTION
    r += 1
    bench_headers = ["Benchmark", "Return", "Fidelity Brokerage", "Roth IRA", "401(k)", "HSA", "Robinhood"]
    hdr(ws, r, bench_headers)
    r += 1
    bench_list = [
        ("S&P 500", (benchmarks or {}).get("S&P 500")),
        ("Dow Jones", (benchmarks or {}).get("Dow Jones")),
        ("NASDAQ", (benchmarks or {}).get("NASDAQ")),
    ]
    sp500_row = r
    for bname, bret in bench_list:
        ws.cell(r, 1, bname).font = BLACK
        ws.cell(r, 2).font = BLUE; ws.cell(r, 2).number_format = P_FMT
        ws.cell(r, 2, bret)
        # Alpha per account = account TWR - benchmark
        for ci, acct_r in enumerate([first_r, first_r+1, first_r+2, first_r+3, first_r+5], 3):
            ws.cell(r, ci).value = f'=IF(OR(E{acct_r}="N/A",B{r}=""),"N/A",E{acct_r}-B{r})'
            ws.cell(r, ci).font = BLACK; ws.cell(r, ci).number_format = P_FMT
        brd(ws, r, r, 1, 7); zb(ws, r, 7)
        r += 1

    # Fill Alpha column in account summary (vs S&P 500)
    for i, acct_r in enumerate([first_r, first_r+1, first_r+2, first_r+3, first_r+5]):
        ws.cell(acct_r, 8).value = f'=IF(OR(E{acct_r}="N/A",B{sp500_row}=""),"N/A",E{acct_r}-B{sp500_row})'

    # --- Section D: Key Metrics ---
    r += 1
    ws.cell(r, 1, "KEY METRICS").font = SECTION
    r += 1
    ws.cell(r, 1, "Total Portfolio Value (Ending)").font = BLACK
    ws.cell(r, 2).value = f"=C{total_r}"; ws.cell(r, 2).font = GREEN; ws.cell(r, 2).number_format = D_FMT
    r += 1
    ws.cell(r, 1, "Capital Deployed (Beginning)").font = BLACK
    ws.cell(r, 2).value = f"=B{total_r}"; ws.cell(r, 2).font = GREEN; ws.cell(r, 2).number_format = D_FMT
    r += 1

    # Total Investment Gain = sum of all account investment gains via cross-sheet refs
    ws.cell(r, 1, "Total Investment Gain").font = BOLD
    # Build formula summing investment gain from each tab
    # These reference the "Total Investment Gain" row in each account tab
    # We'll hardcode the references based on known row positions
    ig_refs = []
    for name, sref, atype in acct_rows:
        ig_row = acct_info.get(f"ig_row_{atype}")
        if ig_row:
            ig_refs.append(f"{sref}!B{ig_row}")
    if ig_refs:
        ws.cell(r, 2).value = "=" + "+".join(ig_refs)
    ws.cell(r, 2).font = BOLD; ws.cell(r, 2).number_format = D_FMT
    r += 1
    ws.cell(r, 1, "YoY Growth").font = BLACK
    ws.cell(r, 2).value = f"=IF(B{r-2}=0,0,(B{r-3}-B{r-2})/B{r-2})"
    ws.cell(r, 2).font = BLACK; ws.cell(r, 2).number_format = P_FMT

    # --- Section D2: Liquidity Breakdown ---
    r += 2
    ws.cell(r, 1, "LIQUIDITY BREAKDOWN").font = SECTION
    r += 1
    ws.cell(r, 1, "Liquid").font = BLACK
    # Fid Brok + Roth IRA + HSA + Robinhood + Cash
    ws.cell(r, 2).value = f"=C{first_r}+C{first_r+1}+C{first_r+3}+C{first_r+5}+C{first_r+6}"
    ws.cell(r, 2).font = BLACK; ws.cell(r, 2).number_format = D_FMT
    ws.cell(r, 3).value = f"=B{r}/C{total_r}"; ws.cell(r, 3).number_format = P_FMT
    ws.cell(r, 4, "5 accounts").font = GRAY
    ws.cell(r, 5, "Fid Brok + Roth IRA + HSA + Robinhood + Cash").font = GRAY
    liq_r = r
    r += 1
    ws.cell(r, 1, "Illiquid").font = BLACK
    # 401(k) + Angel
    ws.cell(r, 2).value = f"=C{first_r+2}+C{first_r+4}"
    ws.cell(r, 2).font = BLACK; ws.cell(r, 2).number_format = D_FMT
    ws.cell(r, 3).value = f"=B{r}/C{total_r}"; ws.cell(r, 3).number_format = P_FMT
    ws.cell(r, 4, "2 accounts").font = GRAY
    ws.cell(r, 5, "401(k) + Angel Investments").font = GRAY
    illiq_r = r
    r += 1
    ws.cell(r, 1, "Total").font = BOLD
    ws.cell(r, 2).value = f"=B{liq_r}+B{illiq_r}"; ws.cell(r, 2).font = BOLD; ws.cell(r, 2).number_format = D_FMT

    # --- Section E: Sector Concentration ---
    r += 2
    ws.cell(r, 1, "SECTOR CONCENTRATION").font = SECTION
    r += 1

    # Build sector data from all holdings
    sector_vals = {}; sector_counts = {}; sector_by_acct = {}
    acct_short = {"fidelity_BROKERAGE": "Fidelity Brokerage", "fidelity_ROTH_IRA": "Roth IRA",
                  "fidelity_HSA": "HSA", "robinhood": "Robinhood"}

    for acct_key, holdings in all_holdings.items():
        scale = 1.0
        if acct_key == "robinhood":
            total_sec = sum(h["mv"] for h in holdings.values())
            md = rh_raw["robinhood"]["accounts"][0]["balances"]["current"]
            net_pv = total_sec + md
            scale = net_pv / total_sec if total_sec else 1
        short = acct_short.get(acct_key, acct_key)
        for ticker, h in holdings.items():
            sec = SECTOR_MAP.get(ticker, "Other")
            val = h["mv"] * scale
            sector_vals[sec] = sector_vals.get(sec, 0) + val
            sector_counts[sec] = sector_counts.get(sec, 0) + 1
            key = (sec, short, ticker)
            sector_by_acct[key] = val

    # 401k funds
    for name, beg, end, gain in K401_DATA["holdings"]:
        short_name = name.replace("BlackRock ", "BR ").replace(" Index", "").replace(" Trust", "").replace("Government ", "Gov ")
        sec = SECTOR_MAP.get(short_name, "Diversified/Index")
        sector_vals[sec] = sector_vals.get(sec, 0) + end
        sector_counts[sec] = sector_counts.get(sec, 0) + 1
        sector_by_acct[(sec, "401(k)", short_name)] = end

    # Angels
    for co, sector, year, series, amt, pm_inv, pm_latest, source in ANGEL_DATA:
        est = amt * (pm_latest / pm_inv)
        sector_vals[sector] = sector_vals.get(sector, 0) + est
        sector_counts[sector] = sector_counts.get(sector, 0) + 1
        sector_by_acct[(sector, "Angel", co)] = est

    all_acct_names = ["Fidelity Brokerage", "Roth IRA", "401(k)", "HSA", "Robinhood", "Angel"]
    sec_headers = ["Sector", "Total Value", "% of Portfolio"] + all_acct_names + ["# Holdings"]
    hdr(ws, r, sec_headers, [22, 14, 12] + [12]*6 + [10])

    sorted_sectors = sorted(sector_vals.items(), key=lambda x: x[1], reverse=True)
    total_sec_val = sum(v for _, v in sorted_sectors)

    r += 1
    sec_start = r
    for sec, val in sorted_sectors:
        ws.cell(r, 1, sec).font = BLACK
        ws.cell(r, 2, round(val, 2)).font = BLUE; ws.cell(r, 2).number_format = D_FMT
        ws.cell(r, 3, round(val/total_sec_val, 4) if total_sec_val else 0).number_format = P_FMT
        for ai, aname in enumerate(all_acct_names, 4):
            acct_val = sum(v for (s, a, t), v in sector_by_acct.items() if s == sec and a == aname)
            if acct_val > 0:
                ws.cell(r, ai, round(acct_val, 2)).font = BLACK; ws.cell(r, ai).number_format = D_FMT
        ws.cell(r, 10, sector_counts.get(sec, 0)).font = BLACK; ws.cell(r, 10).alignment = C_CENTER
        brd(ws, r, r, 1, 10); zb(ws, r, 10)
        r += 1

    # Sector total
    ws.cell(r, 1, "TOTAL").font = BOLD
    ws.cell(r, 2, round(total_sec_val, 2)).font = BOLD; ws.cell(r, 2).number_format = D_FMT
    ws.cell(r, 3, 1.0).font = BOLD; ws.cell(r, 3).number_format = P_FMT
    ws.cell(r, 10, sum(sector_counts.values())).font = BOLD; ws.cell(r, 10).alignment = C_CENTER
    brd(ws, r, r, 1, 10)

    # Concentration risk
    r += 1
    if sorted_sectors:
        ws.cell(r, 1, f"Top sector: {sorted_sectors[0][0]}").font = GRAY
        ws.cell(r, 2, round(sorted_sectors[0][1]/total_sec_val, 4)).number_format = P_FMT
    r += 1
    if len(sorted_sectors) >= 3:
        top3 = sum(v for _, v in sorted_sectors[:3])
        ws.cell(r, 1, "Top 3 sectors combined").font = GRAY
        ws.cell(r, 2, round(top3/total_sec_val, 4)).number_format = P_FMT

    # --- Section F: Geographic Concentration ---
    r += 2
    ws.cell(r, 1, "GEOGRAPHIC CONCENTRATION").font = SECTION
    r += 1

    us_total = 0; intl_breakdown = {}
    for acct_key, holdings in all_holdings.items():
        scale = 1.0
        if acct_key == "robinhood":
            total_sec = sum(h["mv"] for h in holdings.values())
            md = rh_raw["robinhood"]["accounts"][0]["balances"]["current"]
            scale = (total_sec + md) / total_sec if total_sec else 1
        for ticker, h in holdings.items():
            val = h["mv"] * scale
            country = GEO_MAP.get(ticker, "United States")
            if country == "United States": us_total += val
            else: intl_breakdown[country] = intl_breakdown.get(country, 0) + val

    # 401k
    for name, beg, end, gain in K401_DATA["holdings"]:
        short_name = name.replace("BlackRock ", "BR ").replace(" Index", "").replace(" Trust", "").replace("Government ", "Gov ")
        country = GEO_MAP.get(short_name, "United States")
        if country == "United States": us_total += end
        else: intl_breakdown[country] = intl_breakdown.get(country, 0) + end

    # Angels (all US)
    for co, sector, year, series, amt, pm_inv, pm_latest, source in ANGEL_DATA:
        us_total += amt * (pm_latest / pm_inv)

    intl_total = sum(intl_breakdown.values())
    geo_total = us_total + intl_total

    geo_headers = ["Region", "Value", "% of Portfolio"]
    hdr(ws, r, geo_headers, [30, 16, 14])

    r += 1
    ws.cell(r, 1, "United States").font = BLACK
    ws.cell(r, 2, round(us_total, 2)).font = BLUE; ws.cell(r, 2).number_format = D_FMT
    ws.cell(r, 3, round(us_total/geo_total, 4) if geo_total else 0).number_format = P_FMT
    brd(ws, r, r, 1, 3)

    r += 1
    ws.cell(r, 1, "International").font = BOLD
    ws.cell(r, 2, round(intl_total, 2)).font = BLUE; ws.cell(r, 2).number_format = D_FMT
    ws.cell(r, 3, round(intl_total/geo_total, 4) if geo_total else 0).number_format = P_FMT
    brd(ws, r, r, 1, 3)

    for country, val in sorted(intl_breakdown.items(), key=lambda x: x[1], reverse=True):
        r += 1
        ws.cell(r, 1, f"  — {country}").font = GRAY
        ws.cell(r, 2, round(val, 2)).font = BLUE; ws.cell(r, 2).number_format = D_FMT
        ws.cell(r, 3, round(val/geo_total, 4) if geo_total else 0).number_format = P_FMT
        brd(ws, r, r, 1, 3)

    r += 2
    ws.cell(r, 1, "Note: Robinhood holdings scaled to net portfolio value (margin adjustment).").font = GRAY
    r += 1
    ws.cell(r, 1, "Note: BR Intl Equity Index classified as 'Diversified Intl'.").font = GRAY
    r += 1
    ws.cell(r, 1, "Note: All angel investments are US-headquartered.").font = GRAY

    # --- Section G: Risk Metrics placeholder ---
    r += 2
    ws.cell(r, 1, "RISK METRICS").font = SECTION
    r += 1
    ws.cell(r, 1, "Risk metrics require 12 months of monthly return data. Will be populated at year-end.").font = GRAY
    r += 1
    ws.cell(r, 1, "Scope: Fidelity Roth IRA + HSA + Robinhood (Fid Brokerage excluded — RSU distortion, 401k excluded — quarterly only).").font = GRAY

    return ws


# ============================================================
# LOAD MANUAL DATA
# ============================================================
def load_manual_data(manual_json_path=None):
    """Load 401(k), Angel, and cash balance data from external JSON.
    Falls back to module-level defaults if no file provided."""
    if manual_json_path and Path(manual_json_path).exists():
        md = json.loads(Path(manual_json_path).read_text())
        k401 = md.get("k401_data", {})
        # Convert angel list-of-dicts to tuple format expected by build functions
        angel_list = []
        for a in md.get("angel_data", []):
            angel_list.append((
                a["company"], a["sector"], a["year"], a["series"],
                a["amount"], a["pm_invest"], a["pm_latest"], a["source"]
            ))
        cash = md.get("cash_balances", {})
        return k401, angel_list, cash
    return K401_DATA, ANGEL_DATA, CASH_BALANCES


# ============================================================
# BUILD WORKBOOK (parameterized entry point)
# ============================================================
def build_workbook(fid_json_path=None, rh_json_path=None, output_path=None,
                   manual_json_path=None, benchmarks=None,
                   fid_data_dict=None, rh_raw_dict=None, merrill_raw=None,
                   cash_current=None, cash_history=None):
    """Build the portfolio analysis workbook.

    Data can be provided either as file paths or as pre-loaded dicts.
    If dicts are provided, they take precedence over file paths.
    merrill_raw: raw Plaid extraction dict for Merrill (live 401k holdings).
    """
    global K401_DATA, ANGEL_DATA, CASH_BALANCES

    fid_json = Path(fid_json_path) if fid_json_path else FIDELITY_JSON
    rh_json = Path(rh_json_path) if rh_json_path else ROBINHOOD_JSON
    out = Path(output_path) if output_path else OUTPUT

    # Load manual data (401k, angels, cash)
    k401, angel_data, cash_balances = load_manual_data(manual_json_path)

    # Temporarily override module globals for build functions that reference them
    orig_k401, orig_angel, orig_cash = K401_DATA, ANGEL_DATA, CASH_BALANCES
    K401_DATA = k401
    ANGEL_DATA = angel_data
    CASH_BALANCES = cash_balances

    try:
        # Load extraction data
        fid_data = fid_data_dict if fid_data_dict is not None else json.loads(fid_json.read_text())
        rh_raw = rh_raw_dict if rh_raw_dict is not None else json.loads(rh_json.read_text())

        # Parse Robinhood holdings
        rh_holdings = {}
        for h in rh_raw["robinhood"]["holdings"]:
            if h["institution_value"] > 1:
                rh_holdings[h["ticker"]] = {
                    "qty": h["quantity"], "price": h["institution_price"],
                    "mv": h["institution_value"], "cb": h.get("cost_basis", 0),
                    "gl": h.get("gain_loss", 0),
                }

        # Parse Fidelity holdings
        fid_holdings = {}
        for acct_key, acct_data in fid_data.items():
            if acct_key == "fidelity_EXCLUDED": continue
            for date_key, holdings in acct_data.get("holdings", {}).items():
                if holdings: fid_holdings[acct_key] = holdings

        all_holdings = {**fid_holdings, "robinhood": rh_holdings}

        wb = Workbook()
        wb.remove(wb.active)

        acct_info = {}

        # Build account tabs
        ws, total_row = build_monthly_tab(wb, "Fidelity Brokerage", "Fidelity Brokerage",
            fid_holdings.get("fidelity_BROKERAGE", {}), cash_balances.get("fidelity_BROKERAGE", 0), "fid_brokerage")

        ws2, total_row2 = build_monthly_tab(wb, "Fidelity Roth IRA", "Fidelity Roth IRA",
            fid_holdings.get("fidelity_ROTH_IRA", {}), cash_balances.get("fidelity_ROTH_IRA", 0), "standard")

        ws3, fund_total = build_401k_tab(wb, live_merrill=merrill_raw)

        ws4, total_row4 = build_monthly_tab(wb, "Fidelity HSA", "Fidelity HSA",
            fid_holdings.get("fidelity_HSA", {}), cash_balances.get("fidelity_HSA", 0), "fid_hsa")

        ws5, total_row5 = build_monthly_tab(wb, "Robinhood", "Robinhood Brokerage (Margin)",
            rh_holdings, 0, "robinhood")

        ws6, angel_total = build_angel_tab(wb)

        # Build Cash tab
        cash_total_row = None
        cash_first_data_row = None
        if cash_current or cash_history:
            ws7, cash_total_row, cash_first_data_row = build_cash_tab(
                wb, cash_current, cash_history or [])

        # Track investment gain rows for dashboard
        for sheet_name in wb.sheetnames:
            ws_check = wb[sheet_name]
            for row in range(1, ws_check.max_row + 1):
                val = ws_check.cell(row, 1).value
                if val == "Total Investment Gain":
                    if sheet_name == "Fidelity Brokerage": acct_info["ig_row_fid_brokerage"] = row
                    elif sheet_name == "Fidelity Roth IRA": acct_info["ig_row_standard"] = row
                    elif sheet_name == "Fidelity HSA": acct_info["ig_row_fid_hsa"] = row
                    elif sheet_name == "Robinhood": acct_info["ig_row_robinhood"] = row
                    elif sheet_name == "401(k)": acct_info["ig_row_401k"] = row

        acct_info["angel_total_row"] = angel_total
        acct_info["ig_row_angel"] = None
        acct_info["cash_total_row"] = cash_total_row
        acct_info["cash_first_data_row"] = cash_first_data_row
        # Determine total column letter for Cash history tab
        if cash_current:
            n_inst = len(cash_current)
            from openpyxl.utils import get_column_letter as _gcl
            acct_info["cash_total_col_letter"] = _gcl(n_inst + 2)
        else:
            acct_info["cash_total_col_letter"] = "D"

        # Dashboard with optional benchmarks
        build_dashboard(wb, acct_info, all_holdings, rh_raw, benchmarks=benchmarks)

        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(out))
        print(f"Saved: {out}")
        print(f"  Tabs: {wb.sheetnames}")
        return str(out)
    finally:
        K401_DATA, ANGEL_DATA, CASH_BALANCES = orig_k401, orig_angel, orig_cash


# ============================================================
# MAIN
# ============================================================
def main():
    build_workbook()

if __name__ == "__main__":
    main()
