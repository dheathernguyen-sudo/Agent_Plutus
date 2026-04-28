"""Write the Recommendations tab into the portfolio Excel workbook.

Executive Summary format:
  1. Title + disclaimer
  2. Overall State (LLM headline + performance context)
  3. Immediate Priority table  (urgent findings)
  4. Active Concerns table     (attention findings)
  5. Opportunities             (context + positive findings)
  6. Suggested Sequencing      (ordered action list)
  7. Structured findings table (machine-readable)
"""
from __future__ import annotations

import logging
import re
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .observations import Finding

logger = logging.getLogger(__name__)

TAB_NAME = "Recommendations"


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
_TITLE_FONT      = Font(name="Calibri", size=16, bold=True,  color="1F3864")
_SECTION_FONT    = Font(name="Calibri", size=11, bold=True,  color="FFFFFF")
_TBL_HDR_FONT    = Font(name="Calibri", size=10, bold=True,  color="1F3864")
_LABEL_FONT      = Font(name="Calibri", size=10, bold=True)
_BODY_FONT       = Font(name="Calibri", size=10)
_NOTE_FONT       = Font(name="Calibri", size=9,  italic=True, color="808080")
_SEQ_NUM_FONT    = Font(name="Calibri", size=10, bold=True,  color="1F3864")

_SECTION_FILL    = PatternFill("solid", fgColor="1F3864")   # dark navy — all sections
_TBL_HDR_FILL    = PatternFill("solid", fgColor="BDD7EE")   # light blue
_URGENT_FILL     = PatternFill("solid", fgColor="FFCCCC")   # light red
_ATTENTION_FILL  = PatternFill("solid", fgColor="FFF2CC")   # light amber
_CONTEXT_FILL    = PatternFill("solid", fgColor="DDEBF7")   # light blue
_POSITIVE_FILL   = PatternFill("solid", fgColor="E2EFDA")   # light green
_ALT_FILL        = PatternFill("solid", fgColor="F9F9F9")   # off-white
_OVERALL_FILL    = PatternFill("solid", fgColor="EEF3FA")   # very light blue

_THIN  = Side(style="thin",   color="CCCCCC")
_MED   = Side(style="medium", color="AAAAAA")
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_BORDER  = Border(left=_MED,  right=_MED,  top=_MED,  bottom=_MED)

_WRAP_TOP   = Alignment(wrap_text=True, vertical="top")
_CENTER_MID = Alignment(horizontal="center", vertical="center")
_LEFT_MID   = Alignment(horizontal="left",   vertical="center")
_LEFT_TOP   = Alignment(horizontal="left",   vertical="top")

# ---------------------------------------------------------------------------
# Metadata maps
# ---------------------------------------------------------------------------
_DISPLAY_NAMES: Dict[str, str] = {
    "margin_leverage":             "Margin Leverage (Robinhood)",
    "sector_concentration":        "Sector Concentration",
    "single_position_concentration": "Single-Position Concentration",
    "employer_stock_concentration": "Employer Stock",
    "cash_vs_target":              "Emergency Cash Shortfall",
    "illiquid_ratio":              "Illiquid Asset Ratio",
    "international_equity_share":  "International Equity Underweight",
    "tax_loss_harvest_candidate":  "Tax-Loss Harvest Opportunities",
    "ytd_vs_benchmark":            "YTD Performance vs S&P 500",
    "ytd_investment_gain":         "YTD Investment Gain",
    "glide_path_drift":            "Glide Path Drift",
    "asset_location_inefficiency": "Asset Location Inefficiency",
    "pre_retirement_equity_risk":  "Pre-Retirement Equity Risk",
    "inflation_hedge_exposure":    "Inflation Hedge Exposure",
    "upcoming_expense_coverage":   "Upcoming Expense Coverage",
    "cash_vs_target":              "Emergency Cash Shortfall",
}

_SEQUENCE_LABELS: Dict[str, Optional[str]] = {
    "margin_leverage":              "Pay down margin debt",
    "cash_vs_target":               "Rebuild emergency cash buffer",
    "employer_stock_concentration": "Trim employer stock on vest schedule",
    "single_position_concentration":"Reduce single-name concentration",
    "tax_loss_harvest_candidate":   "Harvest tax losses before triggering gains",
    "sector_concentration":         "Tilt new contributions away from concentrated sector",
    "illiquid_ratio":               "Grow liquid sleeve via ongoing contributions",
    "international_equity_share":   "Add international equity via new contributions",
    "glide_path_drift":             "Rebalance toward glide-path target",
    "asset_location_inefficiency":  "Relocate tax-inefficient holdings to sheltered accounts",
    "pre_retirement_equity_risk":   "Reduce equity toward retirement target",
    "inflation_hedge_exposure":     "Add inflation hedges (TIPS / REITs / commodities)",
    "ytd_vs_benchmark":             None,   # context only
    "ytd_investment_gain":          None,   # positive only
    "upcoming_expense_coverage":    "Fund upcoming expense target",
}

_SEVERITY_FILLS = {
    "urgent":    _URGENT_FILL,
    "attention": _ATTENTION_FILL,
    "context":   _CONTEXT_FILL,
    "positive":  _POSITIVE_FILL,
}

_SEVERITY_BADGES = {
    "urgent":    "🔴 URGENT",
    "attention": "⚠ ATTENTION",
    "context":   "ℹ CONTEXT",
    "positive":  "✅ POSITIVE",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _display_name(category: str) -> str:
    if category in _DISPLAY_NAMES:
        return _DISPLAY_NAMES[category]
    return category.replace("_", " ").title()


def _format_detail(f: Finding) -> str:
    d = f.detail or {}
    c = f.category
    if c == "margin_leverage":
        return f"{d.get('ratio', 0):.0%} debt/equity"
    if c == "cash_vs_target":
        delta = d.get("delta", 0)
        return f"${abs(delta):,.0f} {'below' if delta < 0 else 'above'} target"
    if c in ("sector_concentration", "single_position_concentration",
             "employer_stock_concentration"):
        pct   = d.get("pct", 0)
        limit = d.get("limit")
        return f"{pct:.0%} vs {limit:.0%} limit" if limit else f"{pct:.0%}"
    if c == "illiquid_ratio":
        return f"{d.get('ratio', 0):.0%} illiquid"
    if c == "international_equity_share":
        return f"{d.get('intl_pct', 0):.0%} (15–50% band)"
    if c == "tax_loss_harvest_candidate":
        total = sum(i.get("loss", 0) for i in (d.get("items") or []))
        return f"${total:,.0f} harvestable"
    if c == "ytd_vs_benchmark":
        return f"{d.get('alpha', 0):+.1%} alpha"
    if c == "ytd_investment_gain":
        return f"${d.get('total', 0):,.0f} YTD"
    if c == "glide_path_drift":
        devs = d.get("leg_deviations", [])
        if devs:
            worst = max(devs, key=lambda x: abs(x.get("gap", 0)))
            return f"{worst['leg']}: {worst['actual']:.0%} vs {worst['target']:.0%}"
    return ""


def _parse_brief(brief_md: str) -> Tuple[str, Dict[str, str]]:
    """Return (headline, {category: narrative})."""
    headline   = ""
    narratives: Dict[str, str] = {}
    in_headline = in_new = in_standing = False

    for line in (brief_md or "").splitlines():
        s = line.strip()
        if s == "## Headline":
            in_headline, in_new, in_standing = True, False, False
        elif s in ("## New / Changed observations", "## New/Changed observations"):
            in_headline, in_new, in_standing = False, True, False
        elif s == "## Standing concerns":
            in_headline, in_new, in_standing = False, False, True
        elif s.startswith("## "):
            in_headline, in_new, in_standing = False, False, False
        elif in_headline and s and not s.startswith("#"):
            headline = s
            in_headline = False
        elif in_new and s.startswith("- **"):
            m = re.match(r"- \*\*(.+?)\*\*\s*[—\-]\s*(.+)", s)
            if m:
                narratives[m.group(1)] = m.group(2)
        elif in_standing and s.startswith("- "):
            m = re.match(r"- (.+?):\s*(.+)", s)
            if m:
                narratives[m.group(1)] = m.group(2)

    return headline, narratives


def _perf_context(findings: List[Finding]) -> str:
    """Build a one-liner performance context from ytd findings."""
    parts = []
    for f in findings:
        if f.category == "ytd_vs_benchmark":
            d = f.detail or {}
            parts.append(
                f"Liquid portfolio YTD: {d.get('portfolio_pct', 0):+.1%}  "
                f"vs S&P 500 {d.get('benchmark_pct', 0):+.1%}  "
                f"(alpha {d.get('alpha', 0):+.1%})"
            )
        elif f.category == "ytd_investment_gain":
            d = f.detail or {}
            parts.append(
                f"YTD gain: ${d.get('total', 0):,.0f}  "
                f"(unrealized ${d.get('unrealized', 0):,.0f}  "
                f"| dividends ${d.get('dividends', 0):,.0f}  "
                f"| realized ${d.get('realized', 0):,.0f})"
            )
    return "   •   ".join(parts)


# ---------------------------------------------------------------------------
# Low-level cell writer
# ---------------------------------------------------------------------------
def _w(ws, row: int, col: int, value, font=None, fill=None,
       alignment=None, border=None, height: Optional[int] = None) -> None:
    cell = ws.cell(row, col, value=value)
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border
    if height:
        ws.row_dimensions[row].height = height


def _section_header(ws, row: int, label: str, last_col: int = 4) -> None:
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=last_col)
    _w(ws, row, 1, label,
       font=_SECTION_FONT, fill=_SECTION_FILL,
       alignment=_LEFT_MID, height=22)


def _table_headers(ws, row: int) -> None:
    for col, (label, width) in enumerate([
        ("#", None), ("Finding", None), ("Key Metric", None), ("Analysis", None)
    ], start=1):
        _w(ws, row, col, label,
           font=_TBL_HDR_FONT, fill=_TBL_HDR_FILL,
           alignment=_LEFT_MID, border=_HDR_BORDER)
    ws.row_dimensions[row].height = 18


def _finding_row(ws, row: int, num: int, f: Finding, narrative: str) -> None:
    fill   = _SEVERITY_FILLS.get(f.severity, _ALT_FILL)
    detail = _format_detail(f)
    text   = narrative or f.headline
    _w(ws, row, 1, num,                        font=_SEQ_NUM_FONT, fill=fill, alignment=_CENTER_MID, border=_CELL_BORDER)
    _w(ws, row, 2, _display_name(f.category),  font=_LABEL_FONT,  fill=fill, alignment=_LEFT_TOP,   border=_CELL_BORDER)
    _w(ws, row, 3, detail,                     font=_BODY_FONT,   fill=fill, alignment=_LEFT_TOP,   border=_CELL_BORDER)
    _w(ws, row, 4, text,                       font=_BODY_FONT,   fill=fill, alignment=_WRAP_TOP,   border=_CELL_BORDER)
    ws.row_dimensions[row].height = 80


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def write_recommendations_tab(workbook_path, brief_md: str,
                               findings: List[Finding]) -> None:
    """Write (or overwrite) the Recommendations tab.

    Args:
        workbook_path: Target .xlsx path.
        brief_md:      LLM-generated markdown brief.
        findings:      Structured Finding list.
    """
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        logger.warning(f"writer: workbook missing at {workbook_path}; skipping")
        return
    try:
        wb = load_workbook(workbook_path)
        if TAB_NAME in wb.sheetnames:
            del wb[TAB_NAME]
        ws = wb.create_sheet(TAB_NAME, 0)   # insert as first tab

        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 78

        headline, narratives = _parse_brief(brief_md)
        today_str = date.today().strftime("%B %d, %Y").replace(" 0", " ")

        # Segregate findings
        urgent    = [f for f in findings if f.severity == "urgent"]
        attention = [f for f in findings if f.severity == "attention"]
        context   = [f for f in findings if f.severity == "context"]
        positive  = [f for f in findings if f.severity == "positive"]

        row = 1

        # ── Title ────────────────────────────────────────────────────────────
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        _w(ws, row, 1, f"Portfolio Executive Summary — {today_str}",
           font=_TITLE_FONT, alignment=_LEFT_MID, height=34)
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        _w(ws, row, 1,
           "Past performance does not guarantee future results. All investments carry risk, "
           "including potential loss of principal. This is general educational information, "
           "not personalized financial advice. Consult a licensed CFP for tailored guidance.",
           font=_NOTE_FONT, alignment=Alignment(wrap_text=True, vertical="top"), height=28)
        row += 2

        # ── Overall State ─────────────────────────────────────────────────────
        _section_header(ws, row, "  OVERALL STATE")
        row += 1

        perf = _perf_context(findings)
        body = headline or "Portfolio summary unavailable."
        if perf:
            body = body + "\n\n" + perf

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        _w(ws, row, 1, body,
           font=_BODY_FONT, fill=_OVERALL_FILL,
           alignment=Alignment(wrap_text=True, vertical="top"), height=52)
        row += 2

        # ── Immediate Priority ────────────────────────────────────────────────
        if urgent:
            _section_header(ws, row, f"  IMMEDIATE PRIORITY  ({len(urgent)})")
            row += 1
            _table_headers(ws, row)
            row += 1
            for i, f in enumerate(urgent, start=1):
                _finding_row(ws, row, i, f, narratives.get(f.category, ""))
                row += 1
            row += 1

        # ── Active Concerns ───────────────────────────────────────────────────
        if attention:
            _section_header(ws, row, f"  ACTIVE CONCERNS  ({len(attention)})")
            row += 1
            _table_headers(ws, row)
            row += 1
            for i, f in enumerate(attention, start=1):
                _finding_row(ws, row, i, f, narratives.get(f.category, ""))
                row += 1
            row += 1

        # ── Opportunities ─────────────────────────────────────────────────────
        opp = context + positive
        if opp:
            _section_header(ws, row, f"  OPPORTUNITIES  ({len(opp)})")
            row += 1
            _table_headers(ws, row)
            row += 1
            for i, f in enumerate(opp, start=1):
                _finding_row(ws, row, i, f, narratives.get(f.category, ""))
                row += 1
            row += 1

        # ── Suggested Sequencing ──────────────────────────────────────────────
        seq = [
            (f, _SEQUENCE_LABELS.get(f.category))
            for f in (urgent + attention + context)
            if _SEQUENCE_LABELS.get(f.category)
        ]
        if seq:
            _section_header(ws, row, "  SUGGESTED SEQUENCING")
            row += 1
            for step, (f, label) in enumerate(seq, start=1):
                fill = _SEVERITY_FILLS.get(f.severity, _ALT_FILL)
                ws.merge_cells(start_row=row, start_column=2,
                               end_row=row,   end_column=4)
                _w(ws, row, 1, step, font=_SEQ_NUM_FONT, fill=fill,
                   alignment=_CENTER_MID, border=_CELL_BORDER)
                _w(ws, row, 2, label, font=_BODY_FONT, fill=fill,
                   alignment=_LEFT_MID, border=_CELL_BORDER)
                ws.row_dimensions[row].height = 18
                row += 1
            row += 1

        # ── Footer disclaimer ─────────────────────────────────────────────────
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        _w(ws, row, 1,
           "Tax, legal, and estate planning recommendations require licensed professionals. "
           "This output is generated automatically and is not a substitute for IRS Form 1099 "
           "or personalized financial advice from a licensed CFP.",
           font=_NOTE_FONT, alignment=Alignment(wrap_text=True, vertical="top"), height=28)
        row += 2

        # ── Machine-readable findings table ───────────────────────────────────
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        _w(ws, row, 1, "Structured Findings (machine-readable)",
           font=Font(name="Calibri", size=10, bold=True, color="808080"),
           alignment=_LEFT_MID, height=18)
        row += 1

        for col, label in enumerate(["Category", "Severity", "Key", "Headline"], start=1):
            _w(ws, row, col, label, font=_TBL_HDR_FONT, fill=_TBL_HDR_FILL,
               alignment=_LEFT_MID, border=_CELL_BORDER)
        ws.row_dimensions[row].height = 16
        row += 1

        for f in findings:
            fill = _SEVERITY_FILLS.get(f.severity, _ALT_FILL)
            _w(ws, row, 1, f.category,  font=_BODY_FONT, fill=fill, alignment=_LEFT_TOP, border=_CELL_BORDER)
            _w(ws, row, 2, f.severity,  font=_BODY_FONT, fill=fill, alignment=_LEFT_TOP, border=_CELL_BORDER)
            _w(ws, row, 3, f.key,       font=_BODY_FONT, fill=fill, alignment=_LEFT_TOP, border=_CELL_BORDER)
            _w(ws, row, 4, f.headline,  font=_BODY_FONT, fill=fill, alignment=_WRAP_TOP, border=_CELL_BORDER)
            ws.row_dimensions[row].height = 28
            row += 1

        wb.save(workbook_path)
        logger.info(f"writer: Recommendations tab written ({row} rows)")

    except Exception as exc:
        logger.warning(f"writer: failed to update Recommendations tab ({exc}); skipping")
