#!/usr/bin/env python3
"""Create a redacted copy of the portfolio workbook for public screenshots.

Redaction policy:
  REDACT  -> All dollar amounts, share quantities, cost bases, account numbers,
             angel company names/amounts, margin debt, cash balances,
             monthly beginning/ending values.
  KEEP    -> Stock tickers, return percentages (TWR, MWRR, monthly return,
             cost basis return, alpha), benchmark returns, sector/geo percentages,
             column headers, section titles, formatting/colors, sheet tab names,
             number-of-holdings counts, sold-position dates & action labels.
"""

import copy
import re
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

SRC = Path("2026_Portfolio_Analysis.xlsx")
DST = Path("2026_Portfolio_Analysis_REDACTED.xlsx")

REDACTED_FONT = Font(name="Arial", size=10, color="999999")
REDACTED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
PLACEHOLDER = "REDACTED"

# Number formats that indicate dollar amounts
DOLLAR_FMTS = {"$#,##0.00", '$#,##0', "#,##0.00", "$#,##0.00_"}
QTY_FMTS = {"#,##0.000", "#,##0"}
PCT_FMTS = {"0.00%", "0.0000"}


def is_dollar_format(fmt: str) -> bool:
    if not fmt:
        return False
    return "$" in fmt or fmt in DOLLAR_FMTS


def is_qty_format(fmt: str) -> bool:
    if not fmt:
        return False
    return fmt in QTY_FMTS


def is_pct_format(fmt: str) -> bool:
    if not fmt:
        return False
    return fmt in PCT_FMTS or "%" in fmt


def is_formula(val) -> bool:
    return isinstance(val, str) and val.startswith("=")


def cell_has_dollar_formula(val: str) -> bool:
    """Check if a formula likely produces a dollar value (SUM, references to dollar cols)."""
    if not is_formula(val):
        return False
    upper = val.upper()
    # SUM formulas are almost always dollar
    if "SUM" in upper:
        return True
    # References to columns B, C, D, E, F, G (typical dollar columns in account tabs)
    # but NOT if it's clearly a percentage formula
    if "%" in upper or "PRODUCT" in upper:
        return False
    return True


def redact_cell(cell, reason="dollar"):
    """Replace cell value with placeholder and style it."""
    cell.value = PLACEHOLDER
    cell.font = REDACTED_FONT
    cell.fill = REDACTED_FILL


def compute_holdings_returns(ws):
    """Recompute Gain/Loss and Return % from hardcoded Market Value and Cost Basis.

    After baking formulas (which may yield None when Excel's cache is empty),
    this fills in the derived values that we want to keep visible.
    Works for any account tab by finding columns from the header row.
    """
    # Find the header row that contains "Return %"
    ret_col = mv_col = cb_col = gl_col = None
    header_row = None

    for r in range(1, ws.max_row + 1):
        row_vals = {}
        for c in range(1, (ws.max_column or 10) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                row_vals[v] = c
        if "Return %" in row_vals:
            header_row = r
            ret_col = row_vals["Return %"]
            mv_col = row_vals.get("Market Value")
            cb_col = row_vals.get("Cost Basis")
            gl_col = row_vals.get("Gain/Loss")
            break

    if not all([header_row, ret_col, mv_col, cb_col]):
        return 0

    # Walk data rows after header, accumulating totals
    total_mv = 0.0
    total_cb = 0.0
    computed = 0

    for r in range(header_row + 1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value

        # Stop at next section
        if isinstance(label, str) and any(kw in label.upper() for kw in [
            "RETURN CALC", "MARGIN ACCOUNT", "MONTHLY CALC", "SOLD POSITIONS",
            "YTD INVESTMENT",
        ]):
            break

        mv = ws.cell(row=r, column=mv_col).value
        cb = ws.cell(row=r, column=cb_col).value
        is_total = isinstance(label, str) and "TOTAL" in label.upper()

        if is_total:
            # Compute TOTAL row from accumulated sums
            if total_cb != 0:
                total_gl = total_mv - total_cb
                if gl_col:
                    ws.cell(row=r, column=gl_col).value = total_gl
                ws.cell(row=r, column=ret_col).value = total_gl / total_cb
                computed += 1
            continue  # Don't break — Robinhood has TOTAL SECURITIES then Margin then NET

        if isinstance(mv, (int, float)) and isinstance(cb, (int, float)):
            gl = mv - cb
            if gl_col and ws.cell(row=r, column=gl_col).value is None:
                ws.cell(row=r, column=gl_col).value = gl
            if cb != 0 and ws.cell(row=r, column=ret_col).value is None:
                ws.cell(row=r, column=ret_col).value = gl / cb
                computed += 1
            total_mv += mv
            total_cb += cb
        elif isinstance(mv, (int, float)):
            # Cash rows (no cost basis) — include in MV total
            total_mv += mv

    return computed


def append_glossary(ws):
    """Append a return-metric glossary at the bottom of the Dashboard tab."""
    row = ws.max_row + 3

    ws.cell(row=row, column=1, value="RETURN METRIC DEFINITIONS").font = Font(
        name="Arial", size=12, bold=True,
    )
    row += 1

    glossary = [
        (
            "Time-Weighted Return",
            "Measures portfolio performance independent of cash flows (deposits/withdrawals). "
            "Calculated as the product of each period's growth factor minus 1: "
            "(1 + R₁) × (1 + R₂) × ... × (1 + Rₙ) − 1, "
            "where Rₙ = (Ending + Withdrawals − Deposits) / Beginning − 1. "
            "Best for comparing manager skill against benchmarks.",
        ),
        (
            "Money-Weighted Return",
            "Measures the internal rate of return (IRR) accounting for the timing and "
            "size of all cash flows. Solves for r in: "
            "Σ CFₜ / (1 + r)^t = 0. "
            "Reflects the investor's actual experience including deposit/withdrawal timing.",
        ),
        (
            "Cost Basis Return",
            "Unrealized gain or loss as a percentage of total cost basis: "
            "Cost Basis Return = (Market Value − Cost Basis) / Cost Basis. "
            "Shows how much current holdings have appreciated relative to what was paid.",
        ),
        (
            "Alpha",
            "Excess return over a benchmark (S&P 500 by default): "
            "Alpha = Account Time-Weighted Return − Benchmark Return. "
            "Positive alpha indicates outperformance; negative indicates underperformance.",
        ),
    ]

    note_font = Font(name="Arial", size=9, italic=True, color="666666")
    label_font = Font(name="Arial", size=10, bold=True)
    body_font = Font(name="Arial", size=10)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for label, definition in glossary:
        c = ws.cell(row=row, column=1, value=label)
        c.font = label_font
        c.border = border
        c = ws.cell(row=row, column=2, value=definition)
        c.font = body_font
        c.border = border
        # Merge columns B-H for readable definition text
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 45
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=1).alignment = Alignment(vertical="top")
        row += 1


def redact_dashboard(ws):
    """Redact Dashboard tab — keep percentages, redact dollar amounts."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue

            val = cell.value
            fmt = cell.number_format or ""

            # Always keep header rows (white font on blue fill)
            if cell.font and cell.font.color and hasattr(cell.font.color, "rgb"):
                rgb = str(cell.font.color.rgb)
                if rgb in ("00FFFFFF", "FFFFFF"):
                    continue

            # Keep section titles, notes, labels in column A
            if cell.column == 1:
                # Rename "Defense Tech" -> "Hardware"
                if isinstance(val, str) and "Defense Tech" in val:
                    cell.value = val.replace("Defense Tech", "Hardware")

                # Redact angel investment company names if they appear
                if isinstance(val, str) and any(
                    name in val
                    for name in [
                        "Anduril", "Saronic", "Deel", "GSBacker",
                        "Z23-889908", "266-209863",
                    ]
                ):
                    redact_cell(cell, "account/company")
                continue

            # Keep percentages
            if is_pct_format(fmt):
                continue

            # Keep count columns (# Holdings = column J in sector section)
            if isinstance(val, (int, float)) and fmt in ("#,##0", "General", "0") and cell.column == 10:
                continue

            # Keep "N/A" text
            if val == "N/A":
                continue

            # Redact dollar amounts (numeric or formula)
            if is_dollar_format(fmt):
                redact_cell(cell)
                continue

            # Redact numeric values that look like dollars
            if isinstance(val, (int, float)) and not is_pct_format(fmt) and abs(val) > 1:
                redact_cell(cell)
                continue

            # Redact formulas that produce dollar values
            if is_formula(val) and cell_has_dollar_formula(val):
                redact_cell(cell)
                continue


def find_ticker_rows(ws):
    """Find row ranges inside CURRENT HOLDINGS and SOLD POSITIONS sections.

    Returns a set of row numbers where security tickers should be redacted.
    """
    ticker_rows = set()
    in_section = False
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        cell = row[0]
        val = cell.value
        if isinstance(val, str):
            upper = val.upper().strip()
            if upper in ("CURRENT HOLDINGS", "SOLD POSITIONS"):
                in_section = True
                continue
            # Section headers that end the ticker sections
            if in_section and any(kw in upper for kw in [
                "RETURN CALCULATIONS", "MARGIN ACCOUNT", "MONTHLY CALCULATIONS",
                "YTD INVESTMENT", "KEY METRICS", "BENCHMARK", "SECTOR",
                "GEOGRAPHIC", "RISK METRICS", "LIQUIDITY",
            ]):
                in_section = False
                continue
        if in_section and cell.row:
            ticker_rows.add(cell.row)
    return ticker_rows


def redact_account_tab(ws):
    """Redact an account tab (Fidelity Brokerage, Roth IRA, HSA, Robinhood).

    Keep: return % (col G/H), dates, action labels.
    Redact: tickers in holdings, quantities, prices, market values, cost bases,
            gains, monthly values.
    """
    ticker_rows = find_ticker_rows(ws)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue

            val = cell.value
            fmt = cell.number_format or ""

            # Keep header rows
            if cell.font and cell.font.color and hasattr(cell.font.color, "rgb"):
                rgb = str(cell.font.color.rgb)
                if rgb in ("00FFFFFF", "FFFFFF"):
                    continue

            # Column A: keep labels and section titles, but redact tickers in holdings
            if cell.column == 1:
                # Redact account numbers
                if isinstance(val, str) and re.search(r"\b\d{3}-\d{6}\b", val):
                    cell.value = re.sub(r"\d{3}-\d{6}", "XXX-XXXXXX", val)
                # Redact tickers inside holdings and sold positions sections
                if cell.row in ticker_rows and isinstance(val, str):
                    if val.upper() not in ("TOTAL", "TOTAL SECURITIES", "NET PORTFOLIO VALUE",
                                           "MARGIN DEBT"):
                        redact_cell(cell, "ticker")
                continue

            # Keep percentages (Return %, Monthly Return, Growth Factor)
            if is_pct_format(fmt):
                continue

            # Keep growth factors (format 0.0000)
            if fmt == "0.0000":
                continue

            # Keep date strings
            if isinstance(val, str) and re.match(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)", val):
                continue

            # Keep action labels and notes (text in non-A columns)
            if isinstance(val, str) and not is_formula(val):
                # Italic notes — redact if they mention specific tickers
                if cell.font and cell.font.italic:
                    # Scrub ticker references from notes (e.g. "WMT, HUT, JPM sold")
                    scrubbed = re.sub(
                        r'\b[A-Z]{2,5}\b(?:\s*,\s*[A-Z]{2,5}\b)*',
                        '***',
                        val,
                    )
                    if scrubbed != val:
                        cell.value = scrubbed
                    continue
                # Keep non-numeric text (action descriptions like "Full exit")
                if not any(c.isdigit() for c in val[:3]):
                    continue

            # Redact everything else: dollar amounts, quantities, formulas
            if is_dollar_format(fmt) or is_qty_format(fmt):
                redact_cell(cell)
                continue

            if isinstance(val, (int, float)):
                # Keep small integers that might be counts
                if isinstance(val, int) and 0 <= val <= 100 and fmt in ("General", "0", "#,##0"):
                    # But redact if it's in a dollar column (B-G typically)
                    if cell.column >= 2 and cell.column <= 7:
                        redact_cell(cell)
                    continue
                redact_cell(cell)
                continue

            if is_formula(val):
                redact_cell(cell)
                continue


def redact_angel_tab(ws):
    """Redact Angel Investments tab — redact company names, all amounts."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue

            val = cell.value
            fmt = cell.number_format or ""

            # Keep headers
            if cell.font and cell.font.color and hasattr(cell.font.color, "rgb"):
                rgb = str(cell.font.color.rgb)
                if rgb in ("00FFFFFF", "FFFFFF"):
                    continue

            # Keep section titles and static labels in column A
            if cell.column == 1:
                if isinstance(val, str) and val.isupper():
                    continue  # Section headers
                # Redact company names (non-header text in col A that isn't a label)
                if isinstance(val, str) and val not in (
                    "Company Name", "TOTAL", "Status", "Notes",
                    "Angel Investments", "Angel Investments — 2026 Performance",
                ):
                    # Keep labels like "TOTAL INVESTED", "TOTAL CURRENT VALUE"
                    if "TOTAL" in val.upper() or "Return" in val:
                        continue
                    redact_cell(cell, "company name")
                continue

            # Keep Status column text (Active/Exited)
            if isinstance(val, str) and val in ("Active", "Exited"):
                continue

            # Keep return multiples (format like 1.5x) — these are percentages essentially
            if isinstance(val, str) and val.endswith("x"):
                continue

            # Keep percentage formats
            if is_pct_format(fmt):
                continue

            # Redact all dollar amounts, dates, and other data
            if isinstance(val, (int, float)):
                redact_cell(cell)
                continue

            if is_formula(val):
                redact_cell(cell)
                continue

            if is_dollar_format(fmt):
                redact_cell(cell)
                continue


def redact_cash_tab(ws):
    """Redact Cash tab — all balances are sensitive."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue

            val = cell.value

            # Keep headers
            if cell.font and cell.font.color and hasattr(cell.font.color, "rgb"):
                rgb = str(cell.font.color.rgb)
                if rgb in ("00FFFFFF", "FFFFFF"):
                    continue

            # Keep labels in column A, section titles
            if cell.column == 1:
                continue

            # Keep percentages
            if is_pct_format(cell.number_format or ""):
                continue

            # Redact all values
            if isinstance(val, (int, float)) or is_formula(val) or is_dollar_format(cell.number_format or ""):
                redact_cell(cell)


def redact_401k_tab(ws):
    """Redact 401(k) tab — keep return %, redact dollar amounts and fund names."""
    ticker_rows = find_ticker_rows(ws)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue

            val = cell.value
            fmt = cell.number_format or ""

            # Keep headers
            if cell.font and cell.font.color and hasattr(cell.font.color, "rgb"):
                rgb = str(cell.font.color.rgb)
                if rgb in ("00FFFFFF", "FFFFFF"):
                    continue

            # Column A: keep section titles but redact fund names in holdings
            if cell.column == 1:
                if cell.row in ticker_rows and isinstance(val, str):
                    if val.upper() not in ("TOTAL", "TOTAL SECURITIES"):
                        redact_cell(cell, "fund name")
                continue

            # Keep percentages
            if is_pct_format(fmt):
                continue

            # Keep growth factors
            if fmt == "0.0000":
                continue

            # Redact dollar amounts, quantities, and formulas
            if is_dollar_format(fmt) or is_qty_format(fmt):
                redact_cell(cell)
                continue

            if isinstance(val, (int, float)) and abs(val) > 1:
                redact_cell(cell)
                continue

            if is_formula(val):
                redact_cell(cell)
                continue


def compute_all_formulas(wb):
    """Evaluate ALL formulas in the workbook using Python arithmetic.

    Replaces the old bake_all_formulas() which relied on Excel's cached
    formula results (empty when the file was saved by openpyxl, not Excel).

    Handles the formula subset used in this workbook:
      SUM, PRODUCT, IF, OR, IFERROR, cell/range references (including
      cross-sheet), and basic arithmetic (+, -, *, /).

    Evaluates in multiple passes so that formulas depending on other
    formulas resolve once their dependencies are ready.
    """
    from openpyxl.utils import column_index_from_string

    _UNRESOLVED = object()

    # -- helpers ----------------------------------------------------------

    def _get(sheet_name, col_str, row_num):
        """Read a cell value; return _UNRESOLVED if still a formula."""
        v = wb[sheet_name].cell(
            row=row_num, column=column_index_from_string(col_str)
        ).value
        if isinstance(v, str) and v.startswith("="):
            return _UNRESOLVED
        return v

    def _get_range(sheet_name, c1, r1, c2, r2):
        """Collect numeric values in a rectangular range.

        Returns _UNRESOLVED if *any* cell in the range is still a formula.
        Non-numeric cells (None, empty strings) are silently skipped so
        that SUM/PRODUCT ignore blank months.
        """
        ws = wb[sheet_name]
        ic1 = column_index_from_string(c1)
        ic2 = column_index_from_string(c2)
        vals = []
        for r in range(r1, r2 + 1):
            for c in range(ic1, ic2 + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and v.startswith("="):
                    return _UNRESOLVED
                if isinstance(v, (int, float)):
                    vals.append(v)
        return vals

    def _split_args(s):
        """Split comma-separated args respecting nested parens and strings."""
        args, depth, curr, in_str = [], 0, [], False
        for ch in s:
            if ch == '"':
                in_str = not in_str
                curr.append(ch)
                continue
            if in_str:
                curr.append(ch)
                continue
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            if ch == "," and depth == 0:
                args.append("".join(curr).strip())
                curr = []
            else:
                curr.append(ch)
        if curr:
            args.append("".join(curr).strip())
        return args

    def _find_close(s, start):
        """Index of the ')' matching the '(' at *start*."""
        depth, in_str = 0, False
        for i in range(start, len(s)):
            if s[i] == '"':
                in_str = not in_str
            if in_str:
                continue
            if s[i] == "(":
                depth += 1
            elif s[i] == ")":
                depth -= 1
                if depth == 0:
                    return i
        return len(s) - 1

    # -- build named-range lookup from workbook ----------------------------
    _named_ranges = {}
    for dn in wb.defined_names.values():
        ref = dn.attr_text
        if ref and "!" in ref:
            _named_ranges[dn.name] = ref.replace("$", "")

    def _resolve_named_ranges(f):
        """Replace named ranges with their cell references.

        e.g. 'fid_brok_TWR' -> "'Fidelity Brokerage'!B6"
        """
        # Sort by length descending so longer names match first
        # (prevents 'fid_brok_TWR' partially matching 'fid_brok_TW')
        for name in sorted(_named_ranges, key=len, reverse=True):
            if name in f:
                f = f.replace(name, _named_ranges[name])
        return f

    # -- per-cell evaluator -----------------------------------------------

    def _eval_formula(formula, sheet_name):
        """Evaluate one formula string.  Returns _UNRESOLVED on failure."""
        f = formula[1:]  # strip leading '='

        # 0. Resolve named ranges to cell references ----------------------
        f = _resolve_named_ranges(f)

        unresolved = False

        # 1. Replace SUM(range) / PRODUCT(range) with computed value ------
        def _repl_func(m):
            nonlocal unresolved
            func = m.group(1).upper()
            inner = m.group(2)
            # cross-sheet range  'Sheet'!A1:B5
            rm = re.match(r"'([^']+)'!([A-Z]+)(\d+):([A-Z]+)(\d+)$", inner)
            if rm:
                vals = _get_range(rm.group(1), rm.group(2), int(rm.group(3)),
                                  rm.group(4), int(rm.group(5)))
            else:
                # same-sheet range  A1:B5
                rm = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)$", inner)
                if rm:
                    vals = _get_range(sheet_name,
                                      rm.group(1), int(rm.group(2)),
                                      rm.group(3), int(rm.group(4)))
                else:
                    unresolved = True
                    return m.group(0)
            if vals is _UNRESOLVED:
                unresolved = True
                return m.group(0)
            if func == "SUM":
                return repr(sum(vals))
            if func == "PRODUCT":
                p = 1.0
                for v in vals:
                    p *= v
                return repr(p)
            return m.group(0)

        f = re.sub(r"(SUM|PRODUCT)\(([^)]+)\)", _repl_func, f, flags=re.I)
        if unresolved:
            return _UNRESOLVED

        # 2. Replace cross-sheet cell refs --------------------------------
        def _repl_xref(m):
            nonlocal unresolved
            sn, col, row = m.group(1), m.group(2), int(m.group(3))
            v = _get(sn, col, row)
            if v is _UNRESOLVED:
                unresolved = True
                return m.group(0)
            if v is None:
                return "0"
            return repr(v)

        # quoted:   'Fidelity Brokerage'!B44
        f = re.sub(r"'([^']+)'!([A-Z]+)(\d+)", _repl_xref, f)
        if unresolved:
            return _UNRESOLVED
        # unquoted: Robinhood!B42  Cash!C10
        f = re.sub(r"([A-Za-z][A-Za-z0-9_ ]*)!([A-Z]+)(\d+)", _repl_xref, f)
        if unresolved:
            return _UNRESOLVED

        # 3. Replace same-sheet cell refs  (B44, E13, …) -----------------
        def _repl_local(m):
            nonlocal unresolved
            col, row = m.group(1), int(m.group(2))
            v = _get(sheet_name, col, row)
            if v is _UNRESOLVED:
                unresolved = True
                return m.group(0)
            if v is None:
                return "0"
            return repr(v)

        f = re.sub(r"(?<![A-Za-z!])([A-Z]{1,3})(\d+)(?![\d:])", _repl_local, f)
        if unresolved:
            return _UNRESOLVED

        # 4. Translate Excel functions to Python --------------------------
        # comparison  =  →  ==   (but not <=  >=  <>)
        f = re.sub(r"(?<![<>!])=(?!=)", "==", f)

        def _xlate(expr):
            """Recursively translate IF / OR / IFERROR to Python."""
            result = expr
            for _ in range(30):          # safety cap
                changed = False

                # IFERROR(expr, fallback)  →  _ie(lambda: expr, fallback)
                idx = result.upper().find("IFERROR(")
                if idx >= 0:
                    ps = idx + 7         # '(' position
                    pe = _find_close(result, ps)
                    args = _split_args(result[ps + 1 : pe])
                    if len(args) == 2:
                        a0 = _xlate(args[0])
                        a1 = _xlate(args[1])
                        repl = f"_ie(lambda: ({a0}), {a1})"
                        result = result[:idx] + repl + result[pe + 1 :]
                        changed = True

                # IF(cond, T, F)  →  ((T) if (cond) else (F))
                # must not match the "IF" inside "IFERROR"
                for pos in range(len(result) - 3):
                    up = result[pos : pos + 3].upper()
                    if up == "IF(" and (pos == 0 or not result[pos - 1].isalpha()):
                        ps = pos + 2
                        pe = _find_close(result, ps)
                        args = _split_args(result[ps + 1 : pe])
                        if len(args) == 3:
                            cond = _xlate(args[0])
                            tv = _xlate(args[1])
                            fv = _xlate(args[2])
                            repl = f"(({tv}) if ({cond}) else ({fv}))"
                            result = result[:pos] + repl + result[pe + 1 :]
                            changed = True
                            break  # restart scan after replacement

                # OR(a, b, …)  →  (a or b or …)
                idx = result.upper().find("OR(")
                if idx >= 0 and (idx == 0 or not result[idx - 1].isalpha()):
                    ps = idx + 2
                    pe = _find_close(result, ps)
                    args = _split_args(result[ps + 1 : pe])
                    parts = [f"({_xlate(a)})" for a in args]
                    repl = "(" + " or ".join(parts) + ")"
                    result = result[:idx] + repl + result[pe + 1 :]
                    changed = True

                if not changed:
                    break
            return result

        f = _xlate(f)

        # 5. eval() with restricted builtins ------------------------------
        def _ie(fn, fallback):
            try:
                r = fn()
                if r is None:
                    return fallback
                return r
            except Exception:
                return fallback

        try:
            result = eval(f, {"__builtins__": {}}, {"_ie": _ie})
            return result
        except Exception:
            return _UNRESOLVED

    # -- collect every formula cell ---------------------------------------
    formula_cells = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(1, ws.max_row + 1):
            for c in range(1, (ws.max_column or 10) + 1):
                if is_formula(ws.cell(row=r, column=c).value):
                    formula_cells.append((sn, r, c))
    total = len(formula_cells)

    # -- multi-pass evaluation --------------------------------------------
    for pass_num in range(15):
        resolved = 0
        for sn, r, c in formula_cells:
            cell = wb[sn].cell(row=r, column=c)
            if not is_formula(cell.value):
                continue
            result = _eval_formula(cell.value, sn)
            if result is not _UNRESOLVED:
                cell.value = result
                resolved += 1
        if resolved == 0:
            break

    remaining = sum(
        1 for sn, r, c in formula_cells
        if is_formula(wb[sn].cell(row=r, column=c).value)
    )
    return total, remaining


def main():
    print(f"Source:  {SRC}")
    print(f"Output: {DST}")
    print()

    # Copy the file first
    shutil.copy2(SRC, DST)

    # Step 1: Open workbook, then bake ALL formulas into static values
    wb = openpyxl.load_workbook(str(DST))
    print(f"Sheets: {wb.sheetnames}")
    print()

    print("Computing all formulas using Python arithmetic...")
    total, remaining = compute_all_formulas(wb)
    print(f"  Evaluated {total - remaining}/{total} formula cells")
    if remaining:
        print(f"  WARNING: {remaining} formula(s) could not be evaluated")
    print()

    # Step 3: Redact
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if sheet_name == "Dashboard":
            redact_dashboard(ws)
            append_glossary(ws)
        elif sheet_name in ("Fidelity Brokerage", "Fidelity Roth IRA", "Fidelity HSA", "Robinhood"):
            redact_account_tab(ws)
        elif sheet_name == "Angel Investments":
            redact_angel_tab(ws)
        elif sheet_name == "Cash":
            redact_cash_tab(ws)
        elif sheet_name == "401(k)":
            redact_401k_tab(ws)
        else:
            # Unknown tab — redact all numeric values to be safe
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        redact_cell(cell)

        print(f"  Redacted: {sheet_name}")

    wb.save(str(DST))
    print()
    print(f"Redacted workbook saved to: {DST}")
    print()
    print("What's visible:")
    print("  - Return percentages (TWR, MWRR, alpha, monthly returns)")
    print("  - Benchmark returns (S&P 500, DJIA, NASDAQ)")
    print("  - Sector/geographic concentration percentages")
    print("  - Column headers and section structure")
    print("  - Sheet tab names and formatting")
    print("  - Sold position dates and action labels")
    print()
    print("What's redacted:")
    print("  - All dollar amounts (balances, values, gains, costs)")
    print("  - Stock tickers / security names in holdings & sold positions")
    print("  - Share quantities")
    print("  - Account numbers")
    print("  - Angel investment company names")
    print("  - Margin debt figures")
    print("  - Cash balances")
    print("  - Monthly beginning/ending values")
    print()
    print("Renamed: 'Defense Tech' -> 'Hardware'")


if __name__ == "__main__":
    main()
