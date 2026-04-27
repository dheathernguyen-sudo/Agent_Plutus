#!/usr/bin/env python3
"""Rebuild the Cash tab with current balances from manual_data.json."""

import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

XLSX = "2026_Portfolio_Analysis.xlsx"
MANUAL_DATA = "manual_data.json"

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HEADER_ALIGN = Alignment(horizontal='center')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
TITLE_FONT = Font(name='Arial', size=14, bold=True)
SECTION_FONT = Font(name='Arial', size=12, bold=True)
BLUE_FONT = Font(name='Arial', size=10, color='0000FF')
BLACK_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', size=10, bold=True)
NOTE_FONT = Font(name='Arial', size=9, italic=True, color='666666')
DOLLAR = '$#,##0.00'


def header_row(ws, row, labels, col_start=1):
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = THIN_BORDER


def cell(ws, row, col, value, font=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or BLUE_FONT
    c.border = THIN_BORDER
    if fmt:
        c.number_format = fmt
    return c


def main():
    wb = openpyxl.load_workbook(XLSX)

    # Delete old Cash tab if it exists, recreate
    if 'Cash' in wb.sheetnames:
        old_idx = wb.sheetnames.index('Cash')
        del wb['Cash']
        ws = wb.create_sheet('Cash', old_idx)
    else:
        ws = wb.create_sheet('Cash')

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16

    # Load Fidelity core cash from manual_data.json
    md = json.loads(open(MANUAL_DATA).read())
    fid_cash = md.get('cash_balances', {})

    # Load Plaid cash (Chase, Marcus) from latest extraction
    plaid_cash = {}
    try:
        import importlib.util
        from pathlib import Path as _Path
        _repo_plaid = _Path(__file__).parent / "repo" / "plaid_extract.py"
        _spec = importlib.util.spec_from_file_location("repo_plaid_extract", str(_repo_plaid))
        _mod = importlib.util.module_from_spec(_spec)
        _spec.loader.exec_module(_mod)
        config = _mod.load_config()
        plaid_cash = _mod.extract_plaid_cash(config)
    except Exception as e:
        print(f"  Could not load Plaid cash (using manual only): {e}")

    # Account labels
    acct_labels = {
        'fidelity_Z23889908': ('Fidelity Brokerage Core', 'Fidelity'),
        'fidelity_266209863': ('Fidelity Roth IRA Core', 'Fidelity'),
        'fidelity_249509651': ('Fidelity HSA Core', 'Fidelity'),
    }

    import datetime

    row = 1
    ws.cell(row=row, column=1, value='Cash Accounts').font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value='Blue = hardcoded from Plaid | Black = formula').font = NOTE_FONT
    row += 2

    # ==================================================================
    # EXTERNAL CASH (not embedded in any investment account)
    # These are the only balances that count toward portfolio total.
    # ==================================================================
    ws.cell(row=row, column=1, value='EXTERNAL CASH ACCOUNTS').font = SECTION_FONT
    row += 1
    header_row(ws, row, ['Account', 'Institution', 'Balance'])
    row += 1

    ext_start = row
    plaid_labels = {'chase': 'Chase', 'marcus': 'Marcus (Goldman Sachs)'}
    for label, cdata in sorted(plaid_cash.items()):
        inst_name = plaid_labels.get(label, label)
        for acct in cdata.get('accounts', []):
            cell(ws, row, 1, acct['name'])
            cell(ws, row, 2, inst_name)
            cell(ws, row, 3, acct['balance'], fmt=DOLLAR)
            row += 1

    if row == ext_start:
        ws.cell(row=row, column=1, value='No external cash data available.').font = NOTE_FONT
        row += 1

    cell(ws, row, 1, 'TOTAL EXTERNAL CASH', font=BOLD_FONT)
    cell(ws, row, 2, '', font=BOLD_FONT)
    cell(ws, row, 3, f'=SUM(C{ext_start}:C{row-1})', font=BOLD_FONT, fmt=DOLLAR)
    total_ext_cash_row = row
    row += 2

    # ==================================================================
    # EMBEDDED CASH (already included in investment account balances)
    # Shown for reference only — NOT added to portfolio totals.
    # ==================================================================
    ws.cell(row=row, column=1, value='EMBEDDED CASH (for reference only — included in account balances)').font = SECTION_FONT
    row += 1
    header_row(ws, row, ['Account', 'Institution', 'Balance'])
    row += 1

    emb_start = row
    for acct_key, balance in fid_cash.items():
        label_info = acct_labels.get(acct_key, (acct_key, 'Unknown'))
        cell(ws, row, 1, label_info[0])
        cell(ws, row, 2, label_info[1])
        cell(ws, row, 3, balance, fmt=DOLLAR)
        row += 1

    cell(ws, row, 1, 'TOTAL EMBEDDED CASH', font=BOLD_FONT)
    cell(ws, row, 2, '', font=BOLD_FONT)
    cell(ws, row, 3, f'=SUM(C{emb_start}:C{row-1})', font=BOLD_FONT, fmt=DOLLAR)
    row += 2

    # ==================================================================
    # SUMMARY
    # ==================================================================
    ws.cell(row=row, column=1, value='CASH SUMMARY').font = SECTION_FONT
    row += 1
    cell(ws, row, 1, 'External Cash (counts toward portfolio)')
    cell(ws, row, 2, f'=C{total_ext_cash_row}', font=BLACK_FONT, fmt=DOLLAR)
    summary_ext_row = row
    row += 1
    cell(ws, row, 1, 'Embedded Cash (already in account balances)')
    cell(ws, row, 2, f'=C{emb_start + len(fid_cash)}', font=BLACK_FONT, fmt=DOLLAR)
    row += 1
    cell(ws, row, 1, 'Total Cash (all sources)', font=BOLD_FONT)
    cell(ws, row, 2, f'=B{row-2}+B{row-1}', font=BOLD_FONT, fmt=DOLLAR)
    total_all_cash_row = row

    ws.sheet_view.showGridLines = False
    wb.save(XLSX)

    from registry import update_registry
    update_registry("Cash", rows={
        "total_cash": total_ext_cash_row,
    })

    print('Cash tab rebuilt successfully.')


if __name__ == '__main__':
    main()
