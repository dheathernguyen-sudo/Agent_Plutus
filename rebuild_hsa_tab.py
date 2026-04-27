#!/usr/bin/env python3
"""Rebuild the Fidelity HSA tab with same structure as Robinhood tab."""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

XLSX = "2026_Portfolio_Analysis.xlsx"


def compute_mwrr(monthly_data, month_labels):
    """Compute annualised Money-Weighted Return (IRR) from monthly cash flows."""
    populated = [m for m in month_labels if m in monthly_data]
    if not populated:
        return None
    first = monthly_data[populated[0]]
    last = monthly_data[populated[-1]]
    cfs = []
    t = 0
    cfs.append((t, -first['begin']))
    for m in populated:
        d = monthly_data[m]
        cfs.append((t + 0.5, -(d['add'] - d['sub'])))
        t += 1
    cfs.append((t, last['end']))
    r = 0.01
    for _ in range(200):
        npv = sum(cf / (1 + r) ** ti for ti, cf in cfs)
        dnpv = sum(-ti * cf / (1 + r) ** (ti + 1) for ti, cf in cfs)
        if abs(dnpv) < 1e-14:
            break
        r_new = r - npv / dnpv
        if abs(r_new - r) < 1e-12:
            r = r_new
            break
        r = r_new
    return (1 + r) ** 12 - 1

# Styles (same as Robinhood)
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HEADER_ALIGN = Alignment(horizontal='center')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
BLUE_FONT = Font(name='Arial', size=10, color='0000FF')
BLACK_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', size=10, bold=True)
SECTION_FONT = Font(name='Arial', size=12, bold=True)
TITLE_FONT = Font(name='Arial', size=14, bold=True)
NOTE_FONT = Font(name='Arial', size=9, italic=True, color='666666')

DOLLAR = '$#,##0.00'
PCT = '0.00%'
NUM = '#,##0'
QTY_FMT = '#,##0.000'


def header_row(ws, row, labels, col_start=1):
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = THIN_BORDER


def data_cell(ws, row, col, value=None, font=None, fmt=None, formula=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or (BLACK_FONT if formula else BLUE_FONT)
    c.border = THIN_BORDER
    if fmt:
        c.number_format = fmt
    return c


def main():
    wb = openpyxl.load_workbook(XLSX)

    if 'Fidelity HSA' in wb.sheetnames:
        old_idx = wb.sheetnames.index('Fidelity HSA')
        del wb['Fidelity HSA']
        ws = wb.create_sheet('Fidelity HSA', old_idx)
    else:
        ws = wb.create_sheet('Fidelity HSA')

    # Column widths
    widths = {'A': 26, 'B': 16, 'C': 14, 'D': 16, 'E': 14, 'F': 19, 'G': 16, 'H': 16, 'I': 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    row = 1

    # ==================================================================
    # TITLE
    # ==================================================================
    ws.cell(row=row, column=1, value='Fidelity HSA — 2026 Performance').font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value='Blue = hardcoded from statement | Black = formula').font = NOTE_FONT
    row += 2

    # ==================================================================
    # SECTION 1: YTD RETURN CALCULATIONS
    # ==================================================================
    ws.cell(row=row, column=1, value='YTD RETURN CALCULATIONS').font = SECTION_FONT
    row += 1
    header_row(ws, row, ['Metric', 'Value', 'Note'])
    row += 1

    twr_row = row
    data_cell(ws, row, 1, 'Time-Weighted Return (YTD)')
    data_cell(ws, row, 2, None, fmt=PCT)  # placeholder
    row += 1

    mwrr_row = row
    data_cell(ws, row, 1, 'Money-Weighted Return (YTD)')
    data_cell(ws, row, 2, None, fmt=PCT)  # filled after monthly data is defined
    data_cell(ws, row, 3, '(computed from monthly cash flows)', font=NOTE_FONT)
    row += 1

    cost_basis_return_row = row
    data_cell(ws, row, 1, 'Cost Basis Return')
    data_cell(ws, row, 2, None, fmt=PCT)  # placeholder — needs total_row, fixed in forward references
    data_cell(ws, row, 3, 'Unrealized G/L / Cost Basis', font=NOTE_FONT)
    row += 1

    div_display_row = row
    data_cell(ws, row, 1, 'Total Contributions')
    data_cell(ws, row, 2, None, fmt=DOLLAR)  # placeholder
    row += 1

    data_cell(ws, row, 1, 'Total Distributions')
    data_cell(ws, row, 2, None, fmt=DOLLAR)  # placeholder
    dist_row = row
    row += 2

    # ==================================================================
    # SECTION 2: YTD INVESTMENT GAIN SUMMARY
    # ==================================================================
    ws.cell(row=row, column=1, value='YTD INVESTMENT GAIN SUMMARY').font = SECTION_FONT
    row += 1
    header_row(ws, row, ['Metric', 'Value', 'Note'])
    row += 1

    gain_div_row = row
    data_cell(ws, row, 1, 'Dividends/Income')
    data_cell(ws, row, 2, None, fmt=DOLLAR)  # placeholder
    row += 1

    gain_unrealized_row = row
    data_cell(ws, row, 1, 'Unrealized Gain/Loss')
    data_cell(ws, row, 2, None, fmt=DOLLAR)  # placeholder — holdings total
    data_cell(ws, row, 3, 'Current holdings vs. cost basis (all-time)', font=NOTE_FONT)
    row += 1

    gain_realized_row = row
    data_cell(ws, row, 1, 'Realized Gain/Loss (2026)')
    data_cell(ws, row, 2, 230.91, fmt=DOLLAR)  # from Fidelity statement
    data_cell(ws, row, 3, 'SPY, VB, QQQ sold Feb 2026 (from statement)', font=NOTE_FONT)
    row += 1

    gain_total_row = row
    data_cell(ws, row, 1, 'Total YTD Gain', font=BOLD_FONT)
    data_cell(ws, row, 2, f'=B{gain_div_row}+B{gain_unrealized_row}+B{gain_realized_row}',
              font=BOLD_FONT, fmt=DOLLAR, formula=True)
    data_cell(ws, row, 3, 'Unrealized + Realized + Dividends', font=NOTE_FONT)
    row += 2

    # ==================================================================
    # SECTION 3: CURRENT HOLDINGS
    # ==================================================================
    ws.cell(row=row, column=1, value='CURRENT HOLDINGS').font = SECTION_FONT
    row += 1
    header_row(ws, row, ['Security', 'Quantity', 'Price', 'Market Value', 'Cost Basis', 'Gain/Loss', 'Return %'])
    row += 1

    # Holdings from Mar 2026 statement
    holdings = [
        ('CEG', 12, 279.25, 3351.00, 2187.16),
        ('DLR', 10, 180.21, 1802.10, 1480.30),
        ('KDEF', 49.636, 52.73, 2617.30, 3000.00),
        ('NFLX', 20, 96.15, 1923.00, 2363.16),
        ('STX', 12, 391.76, 4701.12, 1112.16),
        ('Cash (FDRXX)', None, None, 4623.49, None),
    ]

    hold_first = row
    for item in holdings:
        ticker = item[0]
        qty = item[1]
        price = item[2]
        mv = item[3]
        cb = item[4]

        data_cell(ws, row, 1, ticker)
        if qty is not None:
            data_cell(ws, row, 2, qty, fmt=QTY_FMT)
        else:
            data_cell(ws, row, 2, None)
        if price is not None:
            data_cell(ws, row, 3, price, fmt=DOLLAR)
        else:
            data_cell(ws, row, 3, None)
        data_cell(ws, row, 4, mv, fmt=DOLLAR)
        if cb is not None:
            data_cell(ws, row, 5, cb, fmt=DOLLAR)
            data_cell(ws, row, 6, f'=D{row}-E{row}', font=BLACK_FONT, fmt=DOLLAR, formula=True)
            data_cell(ws, row, 7, f'=IF(E{row}=0,"",F{row}/E{row})', font=BLACK_FONT, fmt=PCT, formula=True)
        else:
            data_cell(ws, row, 5, None)
            data_cell(ws, row, 6, None)
            data_cell(ws, row, 7, None)
        row += 1
    hold_last = row - 1

    # Total
    data_cell(ws, row, 1, 'TOTAL', font=BOLD_FONT)
    data_cell(ws, row, 4, f'=SUM(D{hold_first}:D{hold_last})', font=BOLD_FONT, fmt=DOLLAR, formula=True)
    data_cell(ws, row, 5, f'=SUM(E{hold_first}:E{hold_last})', font=BOLD_FONT, fmt=DOLLAR, formula=True)
    data_cell(ws, row, 6, f'=SUM(F{hold_first}:F{hold_last})', font=BOLD_FONT, fmt=DOLLAR, formula=True)
    data_cell(ws, row, 7, f'=IF(E{row}=0,"",F{row}/E{row})', font=BOLD_FONT, fmt=PCT, formula=True)
    for col in [2, 3]:
        ws.cell(row=row, column=col).border = THIN_BORDER
    total_row = row
    row += 2

    # ==================================================================
    # SECTION 4: MONTHLY CALCULATIONS (unchanged)
    # ==================================================================
    ws.cell(row=row, column=1, value='MONTHLY CALCULATIONS').font = SECTION_FONT
    row += 1
    header_row(ws, row, ['Month', 'Beginning', 'Additions', 'Subtractions', 'Dividends',
                         'Market Change', 'Ending', 'Monthly Return', 'Growth Factor'])
    monthly_hdr = row
    row += 1

    # Monthly data from statements
    # Market Change = Change in Value - Dividends
    monthly_data = {
        'Jan': {'begin': 14036.28, 'add': 0, 'sub': 0, 'div': 24.25, 'change': 705.33, 'end': 14741.61},
        'Feb': {'begin': 14741.61, 'add': 3950.00, 'sub': 0, 'div': 7.64, 'change': 950.93, 'end': 19642.54},
        'Mar': {'begin': 19642.54, 'add': 500.00, 'sub': 0, 'div': 30.39, 'change': -1124.53, 'end': 19018.01},
    }

    month_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    monthly_first = row

    for month_name in month_labels:
        r = row
        data_cell(ws, r, 1, month_name)

        if month_name in monthly_data:
            m = monthly_data[month_name]
            mkt_change = m['change'] - m['div']
            data_cell(ws, r, 2, m['begin'], fmt=DOLLAR)
            data_cell(ws, r, 3, m['add'], fmt=DOLLAR)
            data_cell(ws, r, 4, m['sub'], fmt=DOLLAR)
            data_cell(ws, r, 5, m['div'], fmt=DOLLAR)
            data_cell(ws, r, 6, round(mkt_change, 2), fmt=DOLLAR)
            data_cell(ws, r, 7, m['end'], fmt=DOLLAR)
        else:
            for col in range(2, 8):
                data_cell(ws, r, col, None, fmt=DOLLAR)

        # H: Monthly Return
        data_cell(ws, r, 8, f'=IF(B{r}=0,"",((G{r}+D{r}-C{r})/B{r})-1)',
                  font=BLACK_FONT, fmt=PCT, formula=True)
        # I: Growth Factor
        data_cell(ws, r, 9, f'=IF(H{r}="","",1+H{r})',
                  font=BLACK_FONT, fmt='0.0000', formula=True)
        row += 1

    monthly_last = row - 1
    row += 1

    # Totals
    data_cell(ws, row, 1, 'Totals', font=BOLD_FONT)
    for col in [3, 4, 5, 6]:
        data_cell(ws, row, col,
                  f'=SUM({get_column_letter(col)}{monthly_first}:{get_column_letter(col)}{monthly_last})',
                  font=BOLD_FONT, fmt=DOLLAR, formula=True)
    monthly_totals_row = row
    row += 2

    # ==================================================================
    # SECTION 5: SOLD POSITIONS (unchanged)
    # ==================================================================
    ws.cell(row=row, column=1, value='SOLD POSITIONS').font = SECTION_FONT
    row += 1
    header_row(ws, row, ['Security', 'Date', 'Quantity', 'Cost Basis', 'Proceeds', 'Realized Gain/Loss', 'Action'])
    row += 1

    ws.cell(row=row, column=1, value='2026').font = BOLD_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    for col in range(2, 8):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1

    # Sold in Feb 2026 (from statement)
    sold_2026 = [
        ('SPY', 'Feb 2026', 1.551, 999.51, 1075.32, 'Sold to buy KDEF'),
        ('VB', 'Feb 2026', 2.0, 401.32, 466.99, 'Sold to buy KDEF'),
        ('QQQ', 'Feb 2026', 2.562, 1464.50, 1553.93, 'Sold to buy KDEF'),
    ]

    sold_first = row
    for ticker, date, qty, cb, proceeds, action in sold_2026:
        data_cell(ws, row, 1, ticker)
        data_cell(ws, row, 2, date)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        data_cell(ws, row, 3, qty, fmt=QTY_FMT)
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        data_cell(ws, row, 4, cb, fmt=DOLLAR)
        data_cell(ws, row, 5, proceeds, fmt=DOLLAR)
        data_cell(ws, row, 6, f'=E{row}-D{row}', font=BLACK_FONT, fmt=DOLLAR, formula=True)
        data_cell(ws, row, 7, action)
        row += 1
    sold_last = row - 1

    data_cell(ws, row, 1, '2026 TOTAL', font=BOLD_FONT)
    data_cell(ws, row, 4, f'=SUM(D{sold_first}:D{sold_last})', font=BOLD_FONT, fmt=DOLLAR, formula=True)
    data_cell(ws, row, 5, f'=SUM(E{sold_first}:E{sold_last})', font=BOLD_FONT, fmt=DOLLAR, formula=True)
    data_cell(ws, row, 6, f'=SUM(F{sold_first}:F{sold_last})', font=BOLD_FONT, fmt=DOLLAR, formula=True)
    for col in [2, 3, 7]:
        ws.cell(row=row, column=col).border = THIN_BORDER
    sold_total_row = row

    # ==================================================================
    # Fix forward references
    # ==================================================================
    # Dividends/Income = sum of monthly Dividends column (E)
    ws.cell(row=gain_div_row, column=2, value=f'=E{monthly_totals_row}')
    ws.cell(row=gain_div_row, column=2).font = BLACK_FONT
    ws.cell(row=gain_div_row, column=2).number_format = DOLLAR

    # Unrealized G/L = holdings total
    ws.cell(row=gain_unrealized_row, column=2, value=f'=F{total_row}')
    ws.cell(row=gain_unrealized_row, column=2).font = BLACK_FONT
    ws.cell(row=gain_unrealized_row, column=2).number_format = DOLLAR

    # Cost Basis Return (moved before holdings, needs forward ref)
    ws.cell(row=cost_basis_return_row, column=2, value=f'=F{total_row}/E{total_row}')
    ws.cell(row=cost_basis_return_row, column=2).font = BLACK_FONT
    ws.cell(row=cost_basis_return_row, column=2).number_format = PCT

    # TWR (growth factor is now column I)
    ws.cell(row=twr_row, column=2, value=f'=IFERROR(PRODUCT(I{monthly_first}:I{monthly_last})-1,"")')
    ws.cell(row=twr_row, column=2).font = BLACK_FONT
    ws.cell(row=twr_row, column=2).number_format = PCT

    mwrr = compute_mwrr(monthly_data, month_labels)
    if mwrr is not None:
        ws.cell(row=mwrr_row, column=2, value=round(mwrr, 8))
        ws.cell(row=mwrr_row, column=2).font = BLACK_FONT
        ws.cell(row=mwrr_row, column=2).number_format = PCT

    # Contributions = sum of additions
    ws.cell(row=div_display_row, column=2, value=f'=C{monthly_totals_row}')
    ws.cell(row=div_display_row, column=2).font = BLACK_FONT
    ws.cell(row=div_display_row, column=2).number_format = DOLLAR

    # Distributions = sum of subtractions
    ws.cell(row=dist_row, column=2, value=f'=D{monthly_totals_row}')
    ws.cell(row=dist_row, column=2).font = BLACK_FONT
    ws.cell(row=dist_row, column=2).number_format = DOLLAR

    ws.sheet_view.showGridLines = False
    wb.save(XLSX)

    from registry import update_registry
    update_registry("Fidelity HSA", rows={
        "TWR": twr_row, "MWRR": mwrr_row, "cb_return": cost_basis_return_row,
        "dividends": gain_div_row, "unrealized": gain_unrealized_row,
        "realized": gain_realized_row, "total_ytd": gain_total_row,
        "holdings_total": total_row,
        "monthly_jan": monthly_first, "monthly_dec": monthly_last,
        "monthly_totals": monthly_totals_row, "sold_2026_total": sold_total_row,
    }, holdings={
        "first": hold_first, "last": hold_last, "total": total_row,
        "mv_col": "D", "cb_col": "E", "gl_col": "F",
    })

    print('Fidelity HSA tab rebuilt successfully.')

    # Validate after save
    from validate_workbook import validate_structural, format_findings
    findings = validate_structural(str(XLSX), "Fidelity HSA")
    print(format_findings(findings))
    n_fail = sum(1 for f in findings if f.severity == "ERROR")
    if n_fail:
        print(f"  WARNING: {n_fail} validation error(s) detected!")


if __name__ == '__main__':
    main()
