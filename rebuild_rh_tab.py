#!/usr/bin/env python3
"""Rebuild the Robinhood tab in the portfolio workbook with new layout.

New section order:
1. Return Calculations (TWR, MWRR, Cost Basis Return, Dividends)
2. Investment Gain Summary
3. Current Holdings (with cost basis, gain/loss, return %)
4. Margin Account Details
5. Monthly Calculations (from statement data)
6. Sold Positions
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pdfplumber
import re
import json
from pathlib import Path

XLSX = Path("2026_Portfolio_Analysis.xlsx")


def compute_mwrr_from_stmts(stmts):
    """Compute annualised Money-Weighted Return (IRR) from Robinhood statement data."""
    sorted_keys = sorted(k for k in stmts.keys() if '/2026' in k)
    if not sorted_keys:
        return None
    first = stmts[sorted_keys[0]]
    last = stmts[sorted_keys[-1]]
    cfs = []
    t = 0
    opening = first.get('opening', 0) or 0
    cfs.append((t, -opening))
    for key in sorted_keys:
        s = stmts[key]
        deposits = s.get('deposits', 0) or 0
        withdrawals = s.get('withdrawals', 0) or 0
        cfs.append((t + 0.5, -(deposits - withdrawals)))
        t += 1
    closing = last.get('closing', 0) or 0
    cfs.append((t, closing))
    r = 0.01
    for _ in range(200):
        npv = sum(cf / (1 + r) ** ti for ti, cf in cfs)
        dnpv = sum(-ti * cf / (1 + r) ** (ti + 1) for ti, cf in cfs)
        if abs(dnpv) < 1e-14:
            break
        r_new = r - npv / dnpv
        if abs(r_new - r) < 1e-12:
            r = r_new
            break
        r = r_new
    return (1 + r) ** 12 - 1
STMT_DIR = Path("account statement/Robinhood")

# ==========================================================================
# Styles
# ==========================================================================
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HEADER_ALIGN = Alignment(horizontal='center')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
BLUE_FONT = Font(name='Arial', size=10, color='0000FF')       # hardcoded
BLACK_FONT = Font(name='Arial', size=10)                       # formula
GREEN_FONT = Font(name='Arial', size=10, color='008000')       # cross-sheet
BOLD_FONT = Font(name='Arial', size=10, bold=True)
SECTION_FONT = Font(name='Arial', size=12, bold=True)
TITLE_FONT = Font(name='Arial', size=14, bold=True)
NOTE_FONT = Font(name='Arial', size=9, italic=True, color='666666')

DOLLAR = '$#,##0.00'
PCT = '0.00%'
QTY = '#,##0.000'
NUM = '#,##0'


def header_row(ws, row, labels, col_start=1):
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = THIN_BORDER


def data_cell(ws, row, col, value=None, font=None, fmt=None, formula=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or (BLACK_FONT if formula else BLUE_FONT)
    c.border = THIN_BORDER
    if fmt:
        c.number_format = fmt
    return c


# ==========================================================================
# Parse statements
# ==========================================================================
def parse_dollar_str(s):
    s = s.strip()
    neg = "(" in s or s.startswith("-")
    s = re.sub(r'[$() ,]', '', s)
    try:
        val = float(s)
        return -val if neg else val
    except ValueError:
        return None


def parse_statement(pdf_path):
    """Extract monthly data from a Robinhood statement PDF."""
    with pdfplumber.open(str(pdf_path)) as pdf:
        page1 = pdf.pages[0].extract_text() or ''
        if 'robinhood' not in page1.lower():
            return None

        dm = re.search(r'(\d{2}/\d{2}/\d{4})\s+to\s+(\d{2}/\d{2}/\d{4})', page1)
        if not dm:
            return None

        data = {'start': dm.group(1), 'end': dm.group(2)}
        lines = page1.split('\n')

        for line in lines:
            if 'Portfolio Value' in line:
                nums = re.findall(r'\$[\d,]+\.\d+', line)
                if len(nums) >= 2:
                    data['opening'] = parse_dollar_str(nums[0])
                    data['closing'] = parse_dollar_str(nums[1])
            if line.strip().startswith('Dividends') and 'Capital' not in line:
                nums = re.findall(r'\$[\d,]+\.\d+', line)
                if nums:
                    data['dividends'] = parse_dollar_str(nums[0])

        # Parse all pages for deposits/withdrawals
        all_text = ''
        for page in pdf.pages:
            all_text += (page.extract_text() or '') + '\n'

        deposits = 0.0
        withdrawals = 0.0
        for line in all_text.split('\n'):
            if re.search(r'ACH\s+(Deposit|Transfer)', line, re.IGNORECASE):
                if 'withdrawal' not in line.lower():
                    nums = re.findall(r'\$[\d,]+\.\d+', line)
                    if nums:
                        amt = parse_dollar_str(nums[-1])
                        if amt and amt > 0:
                            deposits += amt
            if re.search(r'ACH\s+Withdrawal', line, re.IGNORECASE):
                nums = re.findall(r'\$[\d,]+\.\d+', line)
                if nums:
                    amt = parse_dollar_str(nums[-1])
                    if amt and amt > 0:
                        withdrawals += amt

        data['deposits'] = deposits
        data['withdrawals'] = withdrawals
        return data


def load_all_statements():
    """Load all statements from the organized folder."""
    statements = {}
    pdf_files = list((STMT_DIR / '2025').glob('*.pdf')) + list(STMT_DIR.glob('*.pdf'))

    for pdf_path in pdf_files:
        try:
            data = parse_statement(pdf_path)
            if data:
                end = data['end']
                key = end[:2] + '/' + end[6:10]
                if key not in statements:
                    statements[key] = data
        except Exception:
            pass

    return statements


# ==========================================================================
# Main rebuild
# ==========================================================================
def main():
    wb = openpyxl.load_workbook(str(XLSX))

    # Delete old Robinhood tab and recreate
    if 'Robinhood' in wb.sheetnames:
        old_idx = wb.sheetnames.index('Robinhood')
        del wb['Robinhood']
        ws = wb.create_sheet('Robinhood', old_idx)
    else:
        ws = wb.create_sheet('Robinhood')

    # Column widths
    widths = {'A': 26, 'B': 16, 'C': 16, 'D': 16, 'E': 16, 'F': 16, 'G': 19, 'H': 16, 'I': 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Parse statements
    stmts = load_all_statements()

    # Load cost basis from saved JSON
    cb_data = json.loads(Path('rh_cost_basis.json').read_text())

    row = 1

    # ==================================================================
    # TITLE
    # ==================================================================
    ws.cell(row=row, column=1, value='Robinhood Brokerage (Margin) — 2026 Performance').font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value='Blue = hardcoded from statement | Black = formula | Green = cross-sheet ref').font = NOTE_FONT
    row += 2  # blank row

    # ==================================================================
    # SECTION 1: YTD RETURN CALCULATIONS
    # ==================================================================
    ws.cell(row=row, column=1, value='YTD RETURN CALCULATIONS').font = SECTION_FONT
    row += 1
    header_row(ws, row, ['Metric', 'Value', 'Note'])
    row += 1

    twr_row = row
    data_cell(ws, row, 1, 'Time-Weighted Return (YTD)')
    data_cell(ws, row, 2, None, fmt=PCT)  # placeholder — will reference monthly section
    row += 1

    mwrr_row = row
    data_cell(ws, row, 1, 'Money-Weighted Return (YTD)')
    data_cell(ws, row, 2, None, fmt=PCT)  # filled after monthly data is defined
    data_cell(ws, row, 3, '(computed from monthly cash flows)', font=NOTE_FONT)
    row += 1

    cost_basis_return_row = row
    data_cell(ws, row, 1, 'Cost Basis Return')
    data_cell(ws, row, 2, None, fmt=PCT)  # placeholder — needs total_sec_row, fixed in forward references
    data_cell(ws, row, 3, 'Unrealized G/L / Cost Basis', font=NOTE_FONT)
    row += 1

    data_cell(ws, row, 1, 'Total Dividends')
    data_cell(ws, row, 2, None, fmt=DOLLAR)  # placeholder
    div_display_row = row
    row += 2

    # ==================================================================
    # SECTION 2: YTD INVESTMENT GAIN SUMMARY
    # ==================================================================
    sec1_start = row
    ws.cell(row=row, column=1, value='YTD INVESTMENT GAIN SUMMARY').font = SECTION_FONT
    row += 1
    header_row(ws, row, ['Metric', 'Value', 'Note'])
    row += 1

    # These will reference the monthly calc section below — use placeholder rows for now
    # We'll fix the formulas after we know the monthly section row numbers
    gain_div_row = row
    data_cell(ws, row, 1, 'Dividends Received')
    data_cell(ws, row, 2, None, fmt=DOLLAR)  # placeholder — monthly totals
    row += 1

    gain_unrealized_row = row
    data_cell(ws, row, 1, 'Unrealized Gain/Loss')
    data_cell(ws, row, 2, None, fmt=DOLLAR)  # placeholder — holdings total G/L
    data_cell(ws, row, 3, 'Current holdings vs. cost basis (all-time)', font=NOTE_FONT)
    row += 1

    gain_realized_row = row
    data_cell(ws, row, 1, 'Realized Gain/Loss (2026)')
    data_cell(ws, row, 2, None, fmt=DOLLAR)  # placeholder — will reference sold section
    data_cell(ws, row, 3, 'NVO + SPOT (proceeds - cost basis)', font=NOTE_FONT)
    row += 1

    gain_total_row = row
    data_cell(ws, row, 1, 'Total YTD Gain', font=BOLD_FONT)
    data_cell(ws, row, 2, f'=B{gain_unrealized_row}+B{gain_realized_row}+B{gain_div_row}', font=BOLD_FONT, fmt=DOLLAR, formula=True)
    data_cell(ws, row, 3, 'Unrealized + Realized + Dividends', font=NOTE_FONT)
    row += 2

    # ==================================================================
    # SECTION 3: CURRENT HOLDINGS
    # ==================================================================
    ws.cell(row=row, column=1, value='CURRENT HOLDINGS').font = SECTION_FONT
    row += 1
    header_row(ws, row, ['Security', 'Quantity', 'Price', 'Market Value', 'Average Cost', 'Cost Basis', 'Gain/Loss', 'Return %'])
    hdr_row = row
    row += 1

    holdings = [
        ('AGIO', 50, 35.19, 1759.50, 35.14, 1757.00),
        ('ISRG', 5, 452.06, 2260.30, 576.20, 2881.00),
        ('MCK', 8, 884.35, 7074.80, 510.70, 4085.60),
        ('MRVL', 20, 107.11, 2142.20, 118.62, 2372.40),
        ('NVDA', 50, 177.40, 8870.00, 77.75, 3887.50),
        ('RCL', 10, 273.63, 2736.30, 242.99, 2429.90),
        ('RDDT', 15, 136.05, 2040.75, 155.82, 2337.30),
        ('TSM', 14, 338.945, 4745.23, 194.69, 2725.66),
    ]

    hold_first = row
    for ticker, qty, price, mv, avg_cost, cb in holdings:
        data_cell(ws, row, 1, ticker)
        data_cell(ws, row, 2, qty, fmt=NUM)
        data_cell(ws, row, 3, price, fmt=DOLLAR)
        data_cell(ws, row, 4, mv, fmt=DOLLAR)
        data_cell(ws, row, 5, avg_cost, fmt=DOLLAR)
        data_cell(ws, row, 6, cb, fmt=DOLLAR)
        data_cell(ws, row, 7, f'=D{row}-F{row}', font=BLACK_FONT, fmt=DOLLAR, formula=True)
        data_cell(ws, row, 8, f'=IF(F{row}=0,"",G{row}/F{row})', font=BLACK_FONT, fmt=PCT, formula=True)
        row += 1
    hold_last = row - 1

    # Totals
    data_cell(ws, row, 1, 'TOTAL SECURITIES', font=BOLD_FONT)
    data_cell(ws, row, 4, f'=SUM(D{hold_first}:D{hold_last})', font=BOLD_FONT, fmt=DOLLAR, formula=True)
    data_cell(ws, row, 6, f'=SUM(F{hold_first}:F{hold_last})', font=BOLD_FONT, fmt=DOLLAR, formula=True)
    data_cell(ws, row, 7, f'=SUM(G{hold_first}:G{hold_last})', font=BOLD_FONT, fmt=DOLLAR, formula=True)
    data_cell(ws, row, 8, f'=IF(F{row}=0,"",G{row}/F{row})', font=BOLD_FONT, fmt=PCT, formula=True)
    total_sec_row = row
    row += 1

    data_cell(ws, row, 1, 'Margin Debt')
    data_cell(ws, row, 4, -14111.01, fmt=DOLLAR)
    margin_debt_row = row
    row += 1

    net_portfolio_row = row
    data_cell(ws, row, 1, 'NET PORTFOLIO VALUE', font=BOLD_FONT)
    data_cell(ws, row, 4, f'=D{total_sec_row}+D{margin_debt_row}', font=BOLD_FONT, fmt=DOLLAR, formula=True)
    row += 2

    # ==================================================================
    # SECTION 4: MARGIN ACCOUNT DETAILS
    # ==================================================================
    ws.cell(row=row, column=1, value='MARGIN ACCOUNT DETAILS').font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value='This is a margin account. Portfolio Value = Total Securities - Margin Debt.').font = NOTE_FONT
    row += 1
    header_row(ws, row, ['Detail', 'Amount'])
    row += 1

    data_cell(ws, row, 1, 'Beginning Year Margin Debt')
    data_cell(ws, row, 2, 14787.43, fmt=DOLLAR)
    row += 1

    data_cell(ws, row, 1, 'Ending Margin Debt')
    data_cell(ws, row, 2, 14111.01, fmt=DOLLAR)
    row += 1

    data_cell(ws, row, 1, 'Annual Margin Interest (est.)')
    data_cell(ws, row, 2, None, fmt=DOLLAR)
    row += 1

    ws.cell(row=row, column=1, value='Market Change includes both stock appreciation and margin interest.').font = NOTE_FONT
    row += 2

    # ==================================================================
    # SECTION 5: MONTHLY CALCULATIONS (unchanged)
    # ==================================================================
    ws.cell(row=row, column=1, value='MONTHLY CALCULATIONS').font = SECTION_FONT
    row += 1
    header_row(ws, row, ['Month', 'Beginning Value', 'Deposits', 'Withdrawals', 'Dividends',
                         'Market Change', 'Ending Value', 'Monthly Return', 'Growth Factor'])
    monthly_hdr = row
    row += 1

    months_2026 = ['01/2026', '02/2026', '03/2026']
    month_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    monthly_first = row
    for i, month_name in enumerate(month_labels):
        r = row + i
        key = f'{i+1:02d}/2026'

        data_cell(ws, r, 1, month_name)

        if key in stmts:
            s = stmts[key]
            opening = s.get('opening', 0) or 0
            closing = s.get('closing', 0) or 0
            deposits = s.get('deposits', 0) or 0
            withdrawals = s.get('withdrawals', 0) or 0
            dividends = s.get('dividends', 0) or 0
            market_chg = closing - opening - deposits + withdrawals - dividends

            data_cell(ws, r, 2, opening, fmt=DOLLAR)       # B: Beginning
            data_cell(ws, r, 3, deposits, fmt=DOLLAR)       # C: Deposits
            data_cell(ws, r, 4, withdrawals, fmt=DOLLAR)    # D: Withdrawals
            data_cell(ws, r, 5, dividends, fmt=DOLLAR)      # E: Dividends
            data_cell(ws, r, 6, round(market_chg, 2), fmt=DOLLAR)  # F: Market Change
            data_cell(ws, r, 7, closing, fmt=DOLLAR)        # G: Ending
        else:
            for col in range(2, 8):
                data_cell(ws, r, col, None, fmt=DOLLAR)

        # H: Monthly Return (formula)
        data_cell(ws, r, 8, f'=IF(B{r}=0,"",((G{r}+D{r}-C{r})/B{r})-1)', font=BLACK_FONT, fmt=PCT, formula=True)
        # I: Growth Factor (formula)
        data_cell(ws, r, 9, f'=IF(H{r}="","",1+H{r})', font=BLACK_FONT, fmt='0.0000', formula=True)

    monthly_last = row + 11  # Dec row
    row = monthly_last + 1

    # Totals row
    row += 1
    data_cell(ws, row, 1, 'Totals', font=BOLD_FONT)
    for col in [3, 4, 5, 6]:
        data_cell(ws, row, col, f'=SUM({get_column_letter(col)}{monthly_first}:{get_column_letter(col)}{monthly_last})',
                  font=BOLD_FONT, fmt=DOLLAR, formula=True)
    monthly_totals_row = row
    row += 2

    # ==================================================================
    # Now fix forward references in earlier sections
    # ==================================================================
    # Dividends Received = sum of monthly dividends
    ws.cell(row=gain_div_row, column=2, value=f'=E{monthly_totals_row}')
    ws.cell(row=gain_div_row, column=2).font = BLACK_FONT
    ws.cell(row=gain_div_row, column=2).number_format = DOLLAR

    # Unrealized G/L = total holdings G/L
    ws.cell(row=gain_unrealized_row, column=2, value=f'=G{total_sec_row}')
    ws.cell(row=gain_unrealized_row, column=2).font = BLACK_FONT
    ws.cell(row=gain_unrealized_row, column=2).number_format = DOLLAR

    # Realized Proceeds (2026 only) — reference will be set after sold section
    # Use placeholder, fix below

    # Cost Basis Return (moved before holdings, needs forward ref)
    ws.cell(row=cost_basis_return_row, column=2, value=f'=G{total_sec_row}/F{total_sec_row}')
    ws.cell(row=cost_basis_return_row, column=2).font = BLACK_FONT
    ws.cell(row=cost_basis_return_row, column=2).number_format = PCT

    # TWR = product of growth factors - 1
    ws.cell(row=twr_row, column=2, value=f'=IFERROR(PRODUCT(I{monthly_first}:I{monthly_last})-1,"")')
    ws.cell(row=twr_row, column=2).font = BLACK_FONT
    ws.cell(row=twr_row, column=2).number_format = PCT

    mwrr = compute_mwrr_from_stmts(stmts)
    if mwrr is not None:
        ws.cell(row=mwrr_row, column=2, value=round(mwrr, 8))
        ws.cell(row=mwrr_row, column=2).font = BLACK_FONT
        ws.cell(row=mwrr_row, column=2).number_format = PCT

    # Total Dividends in return section
    ws.cell(row=div_display_row, column=2, value=f'=E{monthly_totals_row}')
    ws.cell(row=div_display_row, column=2).font = BLACK_FONT
    ws.cell(row=div_display_row, column=2).number_format = DOLLAR

    # ==================================================================
    # SECTION 6: SOLD POSITIONS (unchanged)
    # ==================================================================
    ws.cell(row=row, column=1, value='SOLD POSITIONS').font = SECTION_FONT
    row += 1
    header_row(ws, row, ['Security', 'Date', 'Quantity', 'Cost Basis', 'Proceeds', 'Realized Gain/Loss', 'Action'])
    row += 1

    # 2026 sales first (these are YTD)
    ws.cell(row=row, column=1, value='2026').font = BOLD_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    for col in range(2, 8):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1

    # (ticker, date, qty, cost_basis, proceeds, action)
    sold_2026 = [
        ('NVO', 'Feb 2026', 40, 3556.00, 1955.20, 'Full exit'),
        ('SPOT', 'Mar 2026', 6, 2829.60, 2875.14, 'Full exit'),
    ]

    sold_2026_first = row
    for ticker, date, qty, cb, proceeds, action in sold_2026:
        data_cell(ws, row, 1, ticker)
        data_cell(ws, row, 2, date)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        data_cell(ws, row, 3, qty, fmt=NUM)
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        data_cell(ws, row, 4, cb, fmt=DOLLAR)
        data_cell(ws, row, 5, proceeds, fmt=DOLLAR)
        data_cell(ws, row, 6, f'=E{row}-D{row}', font=BLACK_FONT, fmt=DOLLAR, formula=True)
        data_cell(ws, row, 7, action)
        row += 1
    sold_2026_last = row - 1

    data_cell(ws, row, 1, '2026 TOTAL', font=BOLD_FONT)
    data_cell(ws, row, 4, f'=SUM(D{sold_2026_first}:D{sold_2026_last})', font=BOLD_FONT, fmt=DOLLAR, formula=True)
    data_cell(ws, row, 5, f'=SUM(E{sold_2026_first}:E{sold_2026_last})', font=BOLD_FONT, fmt=DOLLAR, formula=True)
    data_cell(ws, row, 6, f'=SUM(F{sold_2026_first}:F{sold_2026_last})', font=BOLD_FONT, fmt=DOLLAR, formula=True)
    for col in [2, 3, 7]:
        ws.cell(row=row, column=col).border = THIN_BORDER
    sold_2026_total_row = row
    row += 2

    # 2025 sales (prior year, for reference)
    ws.cell(row=row, column=1, value='2025 (Prior Year)').font = BOLD_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    for col in range(2, 8):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 1

    # (ticker, date, qty, cost_basis, proceeds, action)
    # Cost basis from SnapTrade: ARM bought pre-2024, PPLT bought pre-2024, MCK avg cost $510.70
    sold_2025 = [
        ('ARM', 'Jan 2025', 1, None, 143.00, 'Full exit'),
        ('PPLT', 'Jan 2025', 40, None, 3439.91, 'Full exit'),
        ('MCK', 'Mar 2025', 1, 510.70, 656.02, 'Trim (9 to 8)'),
    ]

    sold_2025_first = row
    for ticker, date, qty, cb, proceeds, action in sold_2025:
        data_cell(ws, row, 1, ticker)
        data_cell(ws, row, 2, date)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        data_cell(ws, row, 3, qty, fmt=NUM)
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        if cb is not None:
            data_cell(ws, row, 4, cb, fmt=DOLLAR)
            data_cell(ws, row, 6, f'=E{row}-D{row}', font=BLACK_FONT, fmt=DOLLAR, formula=True)
        else:
            data_cell(ws, row, 4, None, fmt=DOLLAR)
            data_cell(ws, row, 6, None, fmt=DOLLAR)
        data_cell(ws, row, 5, proceeds, fmt=DOLLAR)
        data_cell(ws, row, 7, action)
        row += 1
    sold_2025_last = row - 1

    data_cell(ws, row, 1, '2025 TOTAL', font=BOLD_FONT)
    data_cell(ws, row, 5, f'=SUM(E{sold_2025_first}:E{sold_2025_last})', font=BOLD_FONT, fmt=DOLLAR, formula=True)
    data_cell(ws, row, 6, f'=SUM(F{sold_2025_first}:F{sold_2025_last})', font=BOLD_FONT, fmt=DOLLAR, formula=True)
    for col in [2, 3, 4, 7]:
        ws.cell(row=row, column=col).border = THIN_BORDER

    # Realized G/L (2026 only) — reference the 2026 total realized G/L column
    ws.cell(row=gain_realized_row, column=2, value=f'=F{sold_2026_total_row}')
    ws.cell(row=gain_realized_row, column=2).font = BLACK_FONT
    ws.cell(row=gain_realized_row, column=2).number_format = DOLLAR

    ws.sheet_view.showGridLines = False
    wb.save(str(XLSX))

    from registry import update_registry
    update_registry("Robinhood", rows={
        "TWR": twr_row, "MWRR": mwrr_row, "cb_return": cost_basis_return_row,
        "dividends": gain_div_row, "unrealized": gain_unrealized_row,
        "realized": gain_realized_row, "total_ytd": gain_total_row,
        "holdings_total_mv": total_sec_row, "holdings_total_cb": total_sec_row,
        "holdings_total_gl": total_sec_row,
        "margin_debt": margin_debt_row, "net_portfolio": net_portfolio_row,
        "monthly_jan": monthly_first, "monthly_dec": monthly_last,
        "monthly_totals": monthly_totals_row, "sold_2026_total": sold_2026_total_row,
    }, holdings={
        "first": hold_first, "last": hold_last, "total": total_sec_row,
        "mv_col": "D", "cb_col": "F", "gl_col": "G",
    })

    print(f'Robinhood tab rebuilt successfully.')
    print(f'  Sections: Investment Gain Summary, Current Holdings, Return Calculations,')
    print(f'            Margin Account Details, Monthly Calculations, Sold Positions')
    print(f'  Monthly data populated: {", ".join(sorted(k for k in stmts if "2026" in k))}')

    # Validate after save
    from validate_workbook import validate_structural, format_findings
    findings = validate_structural(str(XLSX), "Robinhood")
    print(format_findings(findings))
    n_fail = sum(1 for f in findings if f.severity == "ERROR")
    if n_fail:
        print(f"  WARNING: {n_fail} validation error(s) detected!")


if __name__ == '__main__':
    main()
